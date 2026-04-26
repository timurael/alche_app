#!/usr/bin/env python3
"""
alche P&L Excel Generator — All 11 Decisions Applied
Generates a comprehensive 24-month P&L in .xlsx format.

Decisions Applied:
  D1: Doctor SaaS → EUR 99/mo, cap 20 doctors (was EUR 150)
  D2: Restaurant → EUR 0 revenue (500+ user trigger not reached in 24mo)
  D3: CGM Hardware → REMOVED (Software-Only BYOD)
  D4: Smoothie Bar → NEW revenue line (EUR 9.67 avg, 65% margin)
  D5: CapEx → NEW section added
  D6: AI → Hybrid Claude+Gemini, EUR 0.35/user/mo COGS
  D7: Waste → Category-weighted 5.5% blended (capsules 4%, powders 7%)
  D8: Terra API → REMOVED (defer, use free HealthKit/Health Connect)
  D9: SaaS Stack → Itemized (PostHog free, Crisp EUR 88/mo)
  D10: Menu Analysis → EUR 0 (founders do it)
  D11: Retail Products → EUR 49/unit, COGS EUR 23.50 (R10 validated)
"""

import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ─── COLOR PALETTE (alche brand-ish, Excel-friendly) ───
CREAM = "F5F0E8"
DEEP = "2C2418"
AMBER = "C4956A"
SAGE = "8B9E7C"
TERRA = "B86B4A"
ROSE = "C47A8A"
STONE = "9E948A"
WHITE = "FFFFFF"
LIGHT_AMBER = "FFF3E8"
LIGHT_SAGE = "F0F5ED"
LIGHT_TERRA = "FFF0EB"
LIGHT_CREAM = "FEFCF9"
NEG_RED = "D4463A"
POS_GREEN = "4A7C59"
HEADER_BG = "3D3225"
SECTION_BG = "EDE7DC"
TOTAL_BG = "E5DDD3"
GRAND_TOTAL_BG = "D8CFC3"

# Fonts
font_title = Font(name="Calibri", size=18, bold=True, color=DEEP)
font_section = Font(name="Calibri", size=11, bold=True, color=DEEP)
font_header = Font(name="Calibri", size=10, bold=True, color=WHITE)
font_sub = Font(name="Calibri", size=10, color=DEEP)
font_sub_grey = Font(name="Calibri", size=10, color="7A7068")
font_total = Font(name="Calibri", size=10, bold=True, color=DEEP)
font_grand_total = Font(name="Calibri", size=11, bold=True, color=DEEP)
font_neg = Font(name="Calibri", size=10, color=NEG_RED)
font_pos = Font(name="Calibri", size=10, color=POS_GREEN)
font_note = Font(name="Calibri", size=9, italic=True, color=STONE)
font_assumption_label = Font(name="Calibri", size=10, color="5C5244")
font_assumption_val = Font(name="Calibri", size=10, bold=True, color=DEEP)
font_decision = Font(name="Calibri", size=10, color=TERRA)

# Fills
fill_header = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
fill_section = PatternFill(start_color=SECTION_BG, end_color=SECTION_BG, fill_type="solid")
fill_total = PatternFill(start_color=TOTAL_BG, end_color=TOTAL_BG, fill_type="solid")
fill_grand = PatternFill(start_color=GRAND_TOTAL_BG, end_color=GRAND_TOTAL_BG, fill_type="solid")
fill_cream = PatternFill(start_color=LIGHT_CREAM, end_color=LIGHT_CREAM, fill_type="solid")
fill_light_amber = PatternFill(start_color=LIGHT_AMBER, end_color=LIGHT_AMBER, fill_type="solid")
fill_light_sage = PatternFill(start_color=LIGHT_SAGE, end_color=LIGHT_SAGE, fill_type="solid")
fill_light_terra = PatternFill(start_color=LIGHT_TERRA, end_color=LIGHT_TERRA, fill_type="solid")
fill_capex = PatternFill(start_color="FFF8F0", end_color="FFF8F0", fill_type="solid")

# Borders
thin_border = Border(
    bottom=Side(style="thin", color="D8CFC3"),
)
thick_border = Border(
    top=Side(style="medium", color=DEEP),
    bottom=Side(style="medium", color=DEEP),
)

# Alignment
align_right = Alignment(horizontal="right", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")
align_center = Alignment(horizontal="center", vertical="center")

# ─── ASSUMPTIONS (with all decisions applied) ───
A = {
    # Subscription pricing
    "core_price": 19,
    "pro_price": 49,
    "prem_price": 99,
    "churn": 0.08,
    "core_pct": 0.52,
    "pro_pct": 0.38,
    "prem_pct": 0.10,
    # New subs per month by period
    "new_subs": [3, 15, 22, 28, 32, 38],  # M1-3, M4-6, M7-9, M10-12, M13-18, M19-24

    # Product pricing
    "pot_price": 49,      # EUR per potion
    "pot_cogs": 23.50,    # EUR per potion COGS (R10: wholesale 40-60% of retail, midpoint EUR 22-25)
    "led_price": 45,      # EUR per LED session
    "led_cost": 25,       # EUR practitioner cost per session
    "event_price": 35,    # EUR per event ticket
    "event_cost": 250,    # EUR fixed cost per event
    "ticket_fee": 0.07,   # 7% ticketing platform fee
    "smoothie_price": 9.67,  # EUR average smoothie (from R2 Cost Analysis PDF)
    "smoothie_cogs_pct": 0.35,  # 35% COGS (from R2)

    # D1: Doctor SaaS — EUR 99/mo, was EUR 150
    "doc_rev": 99,
    "doc_cogs": 25,  # verification cost per new doctor

    # D2: Restaurant — EUR 0 (500+ user trigger not met in 24mo)
    "rest_rev": 0,

    # D3: CGM — REMOVED entirely
    # (no cgm_price, cgm_cogs, cgm_rma)

    # D6: AI costs
    "ai_per_user": 0.35,  # EUR/user/month (from R9 hybrid recommendation)
    "claude_internal": 210,  # EUR/mo Claude Max for team

    # D7: Waste — category weighted
    "waste_capsules": 0.04,
    "waste_powders": 0.07,
    "waste_blended": 0.055,  # blended for P&L

    # 3PL & Stripe
    "tpl": 2.50,          # EUR per physical unit
    "stripe": 0.029,      # 2.9%

    # Payroll (gross × 1.22 employer mult)
    "empl_mult": 1.22,
    "founder_gross": 4166,
    "mgr_gross": 2000,
    "growth_gross": 4500,
    "cto_gross": 6500,
    "space_staff_gross": 1230,  # part-time EUR 1,500 loaded

    # R&D
    "app_maint": 1500,
    "ux_design": 800,  # flagged as underbudgeted per R6

    # G&A
    "rent_cowork": 750,      # M1-3 co-working
    "rent_space": 2500,      # M4+ physical wellness space
    "steuerberater": 400,
    "insurance": 200,

    # D9: SaaS Stack (itemized from R8)
    # Google Workspace ~EUR 150, Figma ~EUR 50, Notion ~EUR 50, Crisp ~EUR 88
    # PostHog: free, Cloudflare: free
    "saas_phase1": 338,     # M1-6
    "saas_phase2": 500,     # M7-12 (add tools as team grows)
    "saas_phase3": 750,     # M13-24

    # D8: Terra API — REMOVED (EUR 0)

    # D10: Menu Analysis — EUR 0

    # CapEx (D5)
    "capex_gmbh": 2500,          # M1
    "capex_legal_hwg": 3500,     # M1
    "capex_rd_testing": 5000,    # M1 (moved from OpEx)
    "capex_space_buildout": 15000,  # M3
    "capex_kitchen": 20000,      # M3
    "capex_led_devices": 5000,   # M3
}

# ─── GROWTH CURVES ───
# Potions (monthly units shipped)
pot_units = [0, 0, 0, 50, 100, 150, 250, 350, 450, 600, 750, 900,
             1100, 1300, 1500, 1800, 2100, 2400, 2700, 3000, 3400, 3800, 4200, 4600]

# LED Therapy sessions (capped at 300/mo by capacity)
led_sessions = [0, 0, 0, 10, 20, 35, 50, 65, 80, 100, 120, 140,
                160, 180, 200, 220, 240, 260, 280, 300, 300, 300, 300, 300]

# Events [count, avg attendees per event]
events = [[0,0],[0,0],[0,0],[1,12],[1,15],[2,15],[2,18],[3,18],[3,18],[3,20],[3,20],[3,20],
          [3,22],[3,22],[4,22],[4,22],[4,24],[4,24],[4,25],[5,25],[5,25],[5,25],[5,25],[5,25]]

# D1: Doctor SaaS — capped at 20 (was 865 at M24)
doc_saas = [0, 0, 0, 0, 0, 2, 5, 8, 10, 12, 15, 17, 18, 19, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20]

# D2: Restaurant — EUR 0 entire 24mo (500+ subs not reached)
rest_covers = [0] * 24

# D4: Smoothie Bar — NEW (physical space opens M4)
smoothie_units = [0, 0, 0, 100, 150, 250, 350, 450, 550, 650, 750, 850,
                  950, 1050, 1150, 1250, 1350, 1450, 1550, 1650, 1750, 1850, 1950, 2000]

# Marketing spend (unchanged)
marketing = [50, 50, 100, 200, 400, 600, 800, 800, 1000, 1200, 1400, 1400,
             1800, 1800, 2000, 2000, 2200, 2200, 2500, 2500, 2800, 2800, 3000, 3000]


# ─── COMPUTE MODEL ───
months = []
total_subs = 0

for m in range(24):
    d = {}

    # Subscriber growth with churn
    if m < 3:
        new_rate = A["new_subs"][0]
    elif m < 6:
        new_rate = A["new_subs"][1]
    elif m < 9:
        new_rate = A["new_subs"][2]
    elif m < 12:
        new_rate = A["new_subs"][3]
    elif m < 18:
        new_rate = A["new_subs"][4]
    else:
        new_rate = A["new_subs"][5]

    churned = round(total_subs * A["churn"])
    total_subs = total_subs - churned + new_rate
    if total_subs < 0:
        total_subs = 0
    d["total_subs"] = total_subs
    d["new_subs"] = new_rate
    d["churned"] = churned

    # Tier split
    d["core_subs"] = round(total_subs * A["core_pct"])
    d["pro_subs"] = round(total_subs * A["pro_pct"])
    d["prem_subs"] = total_subs - d["core_subs"] - d["pro_subs"]
    if d["prem_subs"] < 0:
        d["prem_subs"] = 0

    # ═══ INCOME ═══
    # Potions
    d["pot_units"] = pot_units[m]
    d["pot_revenue"] = d["pot_units"] * A["pot_price"]

    # LED Therapy
    d["led_sessions"] = led_sessions[m]
    d["led_revenue"] = d["led_sessions"] * A["led_price"]

    # Smoothie Bar (NEW — D4)
    d["smoothie_units"] = smoothie_units[m]
    d["smoothie_revenue"] = round(d["smoothie_units"] * A["smoothie_price"])

    # Events
    d["event_count"] = events[m][0]
    d["event_attendees"] = events[m][0] * events[m][1]
    d["event_revenue"] = d["event_attendees"] * A["event_price"]

    # Doctor SaaS (D1: EUR 99, cap 20)
    d["doc_subs"] = doc_saas[m]
    d["doc_new"] = doc_saas[m] - (doc_saas[m-1] if m > 0 else 0)
    if d["doc_new"] < 0:
        d["doc_new"] = 0
    d["doc_revenue"] = d["doc_subs"] * A["doc_rev"]

    # Restaurant (D2: EUR 0)
    d["rest_covers"] = 0
    d["rest_revenue"] = 0

    # Subscriptions
    d["sub_core"] = d["core_subs"] * A["core_price"]
    d["sub_pro"] = d["pro_subs"] * A["pro_price"]
    d["sub_prem"] = d["prem_subs"] * A["prem_price"]
    d["sub_total"] = d["sub_core"] + d["sub_pro"] + d["sub_prem"]

    # Total Income
    d["total_products"] = (d["pot_revenue"] + d["led_revenue"] + d["smoothie_revenue"]
                           + d["event_revenue"] + d["doc_revenue"] + d["rest_revenue"])
    d["total_income"] = d["total_products"] + d["sub_total"]

    # ═══ COGS ═══
    d["cogs_potions"] = round(d["pot_units"] * A["pot_cogs"])
    d["cogs_waste"] = round(d["pot_revenue"] * A["waste_blended"])  # only potions now (no CGM)
    d["cogs_3pl"] = round(d["pot_units"] * A["tpl"]) if m >= 3 else 0  # 3PL from M4 (potions only)
    d["cogs_led"] = round(d["led_sessions"] * A["led_cost"])
    d["cogs_smoothie"] = round(d["smoothie_revenue"] * A["smoothie_cogs_pct"])  # NEW
    d["cogs_events"] = round((d["event_count"] * A["event_cost"]) + (d["event_revenue"] * A["ticket_fee"]))
    d["cogs_docs"] = round(d["doc_new"] * A["doc_cogs"])
    d["cogs_ai"] = round(total_subs * A["ai_per_user"])  # NEW (D6)
    d["cogs_stripe"] = round(d["total_income"] * A["stripe"])

    d["total_cogs"] = (d["cogs_potions"] + d["cogs_waste"] + d["cogs_3pl"] + d["cogs_led"]
                       + d["cogs_smoothie"] + d["cogs_events"] + d["cogs_docs"]
                       + d["cogs_ai"] + d["cogs_stripe"])

    # Gross Profit
    d["gross_profit"] = d["total_income"] - d["total_cogs"]
    d["gross_margin"] = (d["gross_profit"] / d["total_income"] * 100) if d["total_income"] > 0 else 0

    # ═══ OPERATIONS ═══
    # Payroll
    d["pay_founder1"] = round(A["founder_gross"] * A["empl_mult"])
    d["pay_founder2"] = round(A["founder_gross"] * A["empl_mult"])
    d["pay_space_staff"] = round(A["space_staff_gross"] * A["empl_mult"]) if m >= 3 else 0  # NEW from M4
    d["pay_mgr"] = round(A["mgr_gross"] * A["empl_mult"]) if m >= 5 else 0
    d["pay_growth"] = round(A["growth_gross"] * A["empl_mult"]) if m >= 11 else 0
    d["pay_cto"] = round(A["cto_gross"] * A["empl_mult"]) if m >= 17 else 0
    d["pay_total"] = (d["pay_founder1"] + d["pay_founder2"] + d["pay_space_staff"]
                      + d["pay_mgr"] + d["pay_growth"] + d["pay_cto"])

    # R&D
    d["rd_app_maint"] = A["app_maint"]
    d["rd_ux"] = A["ux_design"]
    d["rd_total"] = d["rd_app_maint"] + d["rd_ux"]

    # S&M
    d["sm_paid_ads"] = marketing[m]
    d["sm_content"] = 150
    d["sm_influencer"] = 200 if m >= 6 else 0
    d["sm_menu"] = 0  # D10: EUR 0
    d["sm_total"] = d["sm_paid_ads"] + d["sm_content"] + d["sm_influencer"] + d["sm_menu"]

    # G&A
    d["ga_rent"] = A["rent_cowork"] if m < 3 else A["rent_space"]  # M1-3 co-working, M4+ space
    d["ga_steuer"] = A["steuerberater"]
    d["ga_insurance"] = A["insurance"]
    d["ga_permits"] = 500 if m == 2 else 0  # M3 one-off permits
    d["ga_total"] = d["ga_rent"] + d["ga_steuer"] + d["ga_insurance"] + d["ga_permits"]

    # Management & APIs
    if m < 6:
        d["mg_saas"] = A["saas_phase1"]
    elif m < 12:
        d["mg_saas"] = A["saas_phase2"]
    else:
        d["mg_saas"] = A["saas_phase3"]
    d["mg_claude"] = A["claude_internal"]
    d["mg_terra"] = 0  # D8: removed
    d["mg_ai_api"] = round(total_subs * A["ai_per_user"])  # User-facing AI (now in COGS, but keep line for visibility)
    # NOTE: AI API is already counted in COGS. To avoid double-counting, set OpEx AI to 0.
    # The AI API cost is a COGS item (variable, scales with subscribers).
    d["mg_ai_api_opex"] = 0  # moved to COGS per R9 recommendation
    d["mg_total"] = d["mg_saas"] + d["mg_claude"] + d["mg_ai_api_opex"]

    # Travel
    d["tr_investor"] = 150
    d["tr_conference"] = 100 if m >= 6 else 0
    d["tr_local"] = 50
    d["tr_total"] = d["tr_investor"] + d["tr_conference"] + d["tr_local"]

    # Total OpEx
    d["total_opex"] = (d["pay_total"] + d["rd_total"] + d["sm_total"]
                       + d["ga_total"] + d["mg_total"] + d["tr_total"])

    # ═══ CapEx (D5) ═══
    d["capex_gmbh"] = A["capex_gmbh"] if m == 0 else 0
    d["capex_legal"] = A["capex_legal_hwg"] if m == 0 else 0
    d["capex_rd"] = A["capex_rd_testing"] if m == 0 else 0
    d["capex_buildout"] = A["capex_space_buildout"] if m == 2 else 0
    d["capex_kitchen"] = A["capex_kitchen"] if m == 2 else 0
    d["capex_led_devices"] = A["capex_led_devices"] if m == 2 else 0
    d["total_capex"] = (d["capex_gmbh"] + d["capex_legal"] + d["capex_rd"]
                        + d["capex_buildout"] + d["capex_kitchen"] + d["capex_led_devices"])

    # ═══ TOTAL P&L ═══
    d["ebitda"] = d["gross_profit"] - d["total_opex"]
    d["total_pnl"] = d["ebitda"] - d["total_capex"]

    months.append(d)


# ─── HELPER: Write a P&L row ───
def write_row(ws, row, label, key_or_values, style="sub", number_format="#,##0"):
    """Write a labeled row with 24 monthly values + total."""
    ws.cell(row=row, column=1, value=label)

    if isinstance(key_or_values, str):
        values = [months[m][key_or_values] for m in range(24)]
    else:
        values = key_or_values

    total = sum(values)

    for col, val in enumerate(values, start=2):
        cell = ws.cell(row=row, column=col, value=val)
        cell.number_format = number_format
        cell.alignment = align_right

    # Total column
    total_cell = ws.cell(row=row, column=26, value=total)
    total_cell.number_format = number_format
    total_cell.alignment = align_right

    # Apply style
    if style == "header":
        for c in range(1, 27):
            cell = ws.cell(row=row, column=c)
            cell.font = font_header
            cell.fill = fill_header
    elif style == "section":
        for c in range(1, 27):
            cell = ws.cell(row=row, column=c)
            cell.font = font_section
            cell.fill = fill_section
    elif style == "total":
        for c in range(1, 27):
            cell = ws.cell(row=row, column=c)
            cell.font = font_total
            cell.fill = fill_total
            cell.border = thin_border
    elif style == "grand_total":
        for c in range(1, 27):
            cell = ws.cell(row=row, column=c)
            cell.font = font_grand_total
            cell.fill = fill_grand
            cell.border = thick_border
    elif style == "sub":
        ws.cell(row=row, column=1).font = font_sub
        for c in range(2, 27):
            cell = ws.cell(row=row, column=c)
            cell.font = font_sub
    elif style == "sub_grey":
        ws.cell(row=row, column=1).font = font_sub_grey
        for c in range(2, 27):
            cell = ws.cell(row=row, column=c)
            cell.font = font_sub_grey
    elif style == "margin":
        ws.cell(row=row, column=1).font = font_sub_grey
        for c in range(2, 27):
            cell = ws.cell(row=row, column=c)
            cell.font = font_sub_grey
            cell.number_format = "0.0%"

    return row


def write_label_row(ws, row, label, style="section"):
    """Write a label-only row (no values)."""
    ws.cell(row=row, column=1, value=label)
    if style == "header":
        for c in range(1, 27):
            cell = ws.cell(row=row, column=c)
            cell.font = font_header
            cell.fill = fill_header
    elif style == "section":
        for c in range(1, 27):
            cell = ws.cell(row=row, column=c)
            cell.font = font_section
            cell.fill = fill_section
    return row


# ═══════════════════════════════════════════
# SHEET 1: P&L
# ═══════════════════════════════════════════
ws = wb.active
ws.title = "P&L"
ws.sheet_properties.tabColor = AMBER

# Column widths
ws.column_dimensions["A"].width = 28
for col in range(2, 27):
    ws.column_dimensions[get_column_letter(col)].width = 10

# Freeze panes
ws.freeze_panes = "B4"

# Title
ws.cell(row=1, column=1, value="alche — 24-Month P&L Projection").font = font_title
ws.cell(row=2, column=1, value="Pre-Seed | EUR 500K @ EUR 2.5M Cap | All 11 decisions applied").font = font_note

# Header row
r = 3
ws.cell(row=r, column=1, value="Budget P&L (EUR)")
for i in range(24):
    ws.cell(row=r, column=i+2, value=f"M{i+1}")
ws.cell(row=r, column=26, value="TOTAL")
for c in range(1, 27):
    cell = ws.cell(row=r, column=c)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center

# ─── SUBSCRIBER METRICS ───
r = 4
write_label_row(ws, r, "SUBSCRIBER METRICS", "header")
r += 1
write_row(ws, r, "  New Subs Added", "new_subs", "sub_grey")
r += 1
write_row(ws, r, "  Churned Subs", "churned", "sub_grey")
r += 1
write_row(ws, r, "  Total Subscribers", "total_subs", "total")
r += 1
write_row(ws, r, "    Core (52%)", "core_subs", "sub_grey")
r += 1
write_row(ws, r, "    Pro (38%)", "pro_subs", "sub_grey")
r += 1
write_row(ws, r, "    Premium (10%)", "prem_subs", "sub_grey")

# ─── INCOME ───
r += 2
write_label_row(ws, r, "INCOME", "header")

r += 1
write_label_row(ws, r, "Products & Services", "section")
r += 1
write_row(ws, r, "  Potions — units", "pot_units", "sub_grey")
r += 1
write_row(ws, r, "  Potions — EUR", "pot_revenue", "sub")
r += 1
write_row(ws, r, "  LED Therapy — sessions", "led_sessions", "sub_grey")
r += 1
write_row(ws, r, "  LED Therapy — EUR", "led_revenue", "sub")
r += 1
write_row(ws, r, "  Smoothie Bar — units", "smoothie_units", "sub_grey")
r += 1
write_row(ws, r, "  Smoothie Bar — EUR", "smoothie_revenue", "sub")
r += 1
write_row(ws, r, "  Events — attendees", "event_attendees", "sub_grey")
r += 1
write_row(ws, r, "  Events — EUR", "event_revenue", "sub")
r += 1
write_row(ws, r, "  Doctor SaaS — clinics", "doc_subs", "sub_grey")
r += 1
write_row(ws, r, "  Doctor SaaS — EUR", "doc_revenue", "sub")
r += 1
# Restaurant line kept for transparency (all zeros)
write_row(ws, r, "  Restaurant (Phase 2 — EUR 0)", "rest_revenue", "sub_grey")

r += 1
write_label_row(ws, r, "Subscriptions", "section")
r += 1
write_row(ws, r, "  Core (EUR 19/mo)", "sub_core", "sub")
r += 1
write_row(ws, r, "  Pro (EUR 49/mo)", "sub_pro", "sub")
r += 1
write_row(ws, r, "  Premium (EUR 99/mo)", "sub_prem", "sub")

r += 1
write_row(ws, r, "TOTAL INCOME", "total_income", "grand_total")
income_total_row = r

# ─── COGS ───
r += 2
write_label_row(ws, r, "COST OF GOODS SOLD (COGS)", "header")
r += 1
write_row(ws, r, "  Potions COGS (EUR 23.50/unit)", "cogs_potions", "sub")
r += 1
write_row(ws, r, "  Waste/Spoilage (5.5% blended)", "cogs_waste", "sub")
r += 1
write_row(ws, r, "  3PL Fulfillment (EUR 2.50/unit)", "cogs_3pl", "sub")
r += 1
write_row(ws, r, "  LED Practitioner (EUR 25/session)", "cogs_led", "sub")
r += 1
write_row(ws, r, "  Smoothie Bar COGS (35%)", "cogs_smoothie", "sub")
r += 1
write_row(ws, r, "  Events & Ticketing", "cogs_events", "sub")
r += 1
write_row(ws, r, "  Doctor Verification (EUR 25/new)", "cogs_docs", "sub")
r += 1
write_row(ws, r, "  AI API — User-Facing (EUR 0.35/sub)", "cogs_ai", "sub")
r += 1
write_row(ws, r, "  Stripe Processing (2.9%)", "cogs_stripe", "sub")

r += 1
write_row(ws, r, "TOTAL COGS", "total_cogs", "grand_total")
cogs_total_row = r

# ─── GROSS PROFIT ───
r += 2
write_row(ws, r, "GROSS PROFIT", "gross_profit", "grand_total")
gp_row = r
r += 1
write_row(ws, r, "  Gross Margin %", [months[m]["gross_margin"]/100 for m in range(24)], "margin", "0.0%")

# ─── OPERATIONS ───
r += 2
write_label_row(ws, r, "OPERATING EXPENSES", "header")

# Payroll
r += 1
write_label_row(ws, r, "Payroll", "section")
r += 1
write_row(ws, r, "  Timu (CEO) — EUR 5,082", "pay_founder1", "sub")
r += 1
write_row(ws, r, "  Daria (COO) — EUR 5,082", "pay_founder2", "sub")
r += 1
write_row(ws, r, "  Space Staff (M4+) — EUR 1,501", "pay_space_staff", "sub")
r += 1
write_row(ws, r, "  Partner Mgr (M6+) — EUR 2,440", "pay_mgr", "sub")
r += 1
write_row(ws, r, "  Growth Marketer (M12+) — EUR 5,490", "pay_growth", "sub")
r += 1
write_row(ws, r, "  Tech/Ops Lead (M18+) — EUR 7,930", "pay_cto", "sub")
r += 1
write_row(ws, r, "  Payroll Total", "pay_total", "total")

# R&D
r += 1
write_label_row(ws, r, "R&D", "section")
r += 1
write_row(ws, r, "  App Maintenance", "rd_app_maint", "sub")
r += 1
write_row(ws, r, "  UX/UI Design", "rd_ux", "sub")
r += 1
write_row(ws, r, "  R&D Total", "rd_total", "total")

# S&M
r += 1
write_label_row(ws, r, "Sales & Marketing", "section")
r += 1
write_row(ws, r, "  Paid Ads", "sm_paid_ads", "sub")
r += 1
write_row(ws, r, "  Content Creation", "sm_content", "sub")
r += 1
write_row(ws, r, "  Influencer Seeding (M7+)", "sm_influencer", "sub")
r += 1
write_row(ws, r, "  Menu Analysis (D10: EUR 0)", [0]*24, "sub_grey")
r += 1
write_row(ws, r, "  S&M Total", "sm_total", "total")

# G&A
r += 1
write_label_row(ws, r, "General & Admin", "section")
r += 1
write_row(ws, r, "  Rent (co-work M1-3 / space M4+)", "ga_rent", "sub")
r += 1
write_row(ws, r, "  Steuerberater", "ga_steuer", "sub")
r += 1
write_row(ws, r, "  Insurance", "ga_insurance", "sub")
r += 1
write_row(ws, r, "  Permits & Licenses", "ga_permits", "sub")
r += 1
write_row(ws, r, "  G&A Total", "ga_total", "total")

# Management & APIs
r += 1
write_label_row(ws, r, "Management & APIs", "section")
r += 1
write_row(ws, r, "  SaaS Stack (D9: itemized)", "mg_saas", "sub")
r += 1
write_row(ws, r, "  Claude Max (internal AI)", "mg_claude", "sub")
r += 1
write_row(ws, r, "  Terra API (D8: deferred)", [0]*24, "sub_grey")
r += 1
write_row(ws, r, "  Management Total", "mg_total", "total")

# Travel
r += 1
write_label_row(ws, r, "Travel", "section")
r += 1
write_row(ws, r, "  Investor Travel", "tr_investor", "sub")
r += 1
write_row(ws, r, "  Conferences (M7+)", "tr_conference", "sub")
r += 1
write_row(ws, r, "  Local Transport", "tr_local", "sub")
r += 1
write_row(ws, r, "  Travel Total", "tr_total", "total")

r += 1
write_row(ws, r, "TOTAL OPEX", "total_opex", "grand_total")
opex_total_row = r

# ─── EBITDA ───
r += 2
write_row(ws, r, "EBITDA", "ebitda", "grand_total")
ebitda_row = r

# ─── CAPEX ───
r += 2
write_label_row(ws, r, "CAPITAL EXPENDITURES (CapEx)", "header")
r += 1
write_row(ws, r, "  GmbH Formation (M1)", "capex_gmbh", "sub")
r += 1
write_row(ws, r, "  Legal Opinion HWG (M1)", "capex_legal", "sub")
r += 1
write_row(ws, r, "  R&D Stability Testing (M1)", "capex_rd", "sub")
r += 1
write_row(ws, r, "  Space Buildout (M3)", "capex_buildout", "sub")
r += 1
write_row(ws, r, "  Kitchen Equipment (M3)", "capex_kitchen", "sub")
r += 1
write_row(ws, r, "  LED Therapy Devices (M3)", "capex_led_devices", "sub")
r += 1
write_row(ws, r, "TOTAL CAPEX", "total_capex", "grand_total")
capex_total_row = r

# ─── TOTAL P&L ───
r += 2
write_row(ws, r, "NET P&L (EBITDA - CapEx)", "total_pnl", "grand_total")
pnl_row = r

# ─── CUMULATIVE CASH ───
r += 1
cumulative = []
cash = 500000  # EUR 500K funding
for m in range(24):
    cash += months[m]["total_pnl"]
    cumulative.append(round(cash))
write_row(ws, r, "Cumulative Cash (EUR 500K start)", cumulative, "total")
cash_row = r

# ─── KEY METRICS ───
r += 2
write_label_row(ws, r, "KEY METRICS", "header")
r += 1
write_row(ws, r, "  Monthly Burn Rate", [abs(months[m]["total_pnl"]) if months[m]["total_pnl"] < 0 else 0 for m in range(24)], "sub")
r += 1
write_row(ws, r, "  Revenue per Subscriber", [round(months[m]["total_income"] / months[m]["total_subs"], 2) if months[m]["total_subs"] > 0 else 0 for m in range(24)], "sub", "#,##0.00")
r += 1
write_row(ws, r, "  COGS per Subscriber", [round(months[m]["total_cogs"] / months[m]["total_subs"], 2) if months[m]["total_subs"] > 0 else 0 for m in range(24)], "sub", "#,##0.00")
r += 1
write_row(ws, r, "  Subscription MRR", "sub_total", "sub")
r += 1
# Physical revenue share
phys_rev = [months[m]["pot_revenue"] + months[m]["led_revenue"] + months[m]["smoothie_revenue"] + months[m]["event_revenue"] for m in range(24)]
write_row(ws, r, "  Physical Revenue", phys_rev, "sub")
r += 1
dig_rev = [months[m]["sub_total"] + months[m]["doc_revenue"] for m in range(24)]
write_row(ws, r, "  Digital Revenue", dig_rev, "sub")


# ═══════════════════════════════════════════
# SHEET 2: ASSUMPTIONS
# ═══════════════════════════════════════════
ws2 = wb.create_sheet("Assumptions")
ws2.sheet_properties.tabColor = SAGE
ws2.column_dimensions["A"].width = 32
ws2.column_dimensions["B"].width = 16
ws2.column_dimensions["C"].width = 45
ws2.column_dimensions["D"].width = 3
ws2.column_dimensions["E"].width = 32
ws2.column_dimensions["F"].width = 16
ws2.column_dimensions["G"].width = 45

ws2.cell(row=1, column=1, value="alche — P&L Assumptions").font = font_title
ws2.cell(row=2, column=1, value="All values in EUR unless noted. Decisions D1-D11 applied.").font = font_note

def write_assumption(ws, row, col, label, value, note=""):
    ws.cell(row=row, column=col, value=label).font = font_assumption_label
    ws.cell(row=row, column=col+1, value=value).font = font_assumption_val
    if note:
        ws.cell(row=row, column=col+2, value=note).font = font_note

# Left column
r = 4
ws2.cell(row=r, column=1, value="SUBSCRIPTION MODEL").font = font_section
ws2.cell(row=r, column=1).fill = fill_section
ws2.cell(row=r, column=2).fill = fill_section
ws2.cell(row=r, column=3).fill = fill_section
r += 1
write_assumption(ws2, r, 1, "Core tier price", "EUR 19/mo", "52% of subs")
r += 1
write_assumption(ws2, r, 1, "Pro tier price", "EUR 49/mo", "38% of subs")
r += 1
write_assumption(ws2, r, 1, "Premium tier price", "EUR 99/mo", "10% of subs")
r += 1
write_assumption(ws2, r, 1, "Monthly churn rate", "8%", "Locked data from research")
r += 1
write_assumption(ws2, r, 1, "New subs/mo M1-3", 3, "Pre-launch beta")
r += 1
write_assumption(ws2, r, 1, "New subs/mo M4-6", 15, "Product launch phase")
r += 1
write_assumption(ws2, r, 1, "New subs/mo M7-9", 22, "Growth acceleration")
r += 1
write_assumption(ws2, r, 1, "New subs/mo M10-12", 28, "Marketing ramp")
r += 1
write_assumption(ws2, r, 1, "New subs/mo M13-18", 32, "Scaling with team")
r += 1
write_assumption(ws2, r, 1, "New subs/mo M19-24", 38, "Growth marketer effect")

r += 2
ws2.cell(row=r, column=1, value="PRODUCT PRICING & COGS").font = font_section
ws2.cell(row=r, column=1).fill = fill_section
ws2.cell(row=r, column=2).fill = fill_section
ws2.cell(row=r, column=3).fill = fill_section
r += 1
write_assumption(ws2, r, 1, "Potion price", "EUR 49", "Pending R10 validation")
r += 1
write_assumption(ws2, r, 1, "Potion COGS", "EUR 23.50", "R10: 52% margin (was EUR 11.50/76%)")
r += 1
write_assumption(ws2, r, 1, "LED session price", "EUR 45", "Red light therapy")
r += 1
write_assumption(ws2, r, 1, "LED practitioner cost", "EUR 25", "44% margin")
r += 1
write_assumption(ws2, r, 1, "Smoothie avg price", "EUR 9.67", "From R2 Cost Analysis PDF")
r += 1
write_assumption(ws2, r, 1, "Smoothie COGS %", "35%", "From R2 — 65% gross margin")
r += 1
write_assumption(ws2, r, 1, "Event ticket price", "EUR 35", "Community events")
r += 1
write_assumption(ws2, r, 1, "Event fixed cost", "EUR 250", "Per event")
r += 1
write_assumption(ws2, r, 1, "Ticketing fee", "7%", "Luma/Eventbrite cut")
r += 1
write_assumption(ws2, r, 1, "Doctor SaaS fee", "EUR 99/mo", "D1: was EUR 150")
r += 1
write_assumption(ws2, r, 1, "Doctor verify cost", "EUR 25", "Per new doctor")
r += 1
write_assumption(ws2, r, 1, "Doctor cap", "20 clinics", "D1: max 20 for M1-24")
r += 1
write_assumption(ws2, r, 1, "Waste — capsules", "4%", "D7: category-weighted")
r += 1
write_assumption(ws2, r, 1, "Waste — powders", "7%", "D7: category-weighted")
r += 1
write_assumption(ws2, r, 1, "Waste — blended", "5.5%", "Used in P&L model")
r += 1
write_assumption(ws2, r, 1, "3PL fulfillment", "EUR 2.50/unit", "Potions only (no CGM)")
r += 1
write_assumption(ws2, r, 1, "Stripe fee", "2.9%", "On all revenue")

# Right column — Operational
r2 = 4
ws2.cell(row=r2, column=5, value="OPERATIONS").font = font_section
ws2.cell(row=r2, column=5).fill = fill_section
ws2.cell(row=r2, column=6).fill = fill_section
ws2.cell(row=r2, column=7).fill = fill_section
r2 += 1
write_assumption(ws2, r2, 5, "Founder gross salary (each)", "EUR 4,166", "× 1.22 = EUR 5,082 loaded")
r2 += 1
write_assumption(ws2, r2, 5, "Partner Mgr (M6+)", "EUR 2,000", "× 1.22 = EUR 2,440 loaded")
r2 += 1
write_assumption(ws2, r2, 5, "Growth Marketer (M12+)", "EUR 4,500", "× 1.22 = EUR 5,490 loaded")
r2 += 1
write_assumption(ws2, r2, 5, "Tech/Ops Lead (M18+)", "EUR 6,500", "× 1.22 = EUR 7,930 loaded")
r2 += 1
write_assumption(ws2, r2, 5, "Space Staff (M4+)", "EUR 1,230", "× 1.22 = EUR 1,501 loaded (part-time)")
r2 += 1
write_assumption(ws2, r2, 5, "Employer multiplier", "1.22x", "German social contributions")
r2 += 1
write_assumption(ws2, r2, 5, "App maintenance retainer", "EUR 1,500/mo", "External agency")
r2 += 1
write_assumption(ws2, r2, 5, "UX/UI design retainer", "EUR 800/mo", "Flagged: R6 says EUR 99K/24mo needed")
r2 += 1
write_assumption(ws2, r2, 5, "Claude Max (internal)", "EUR 210/mo", "D6: team AI budget")
r2 += 1
write_assumption(ws2, r2, 5, "AI API per user (COGS)", "EUR 0.35/mo", "D6: hybrid Claude+Gemini (R9)")
r2 += 1
write_assumption(ws2, r2, 5, "Rent M1-3 (co-working)", "EUR 750/mo", "Before space opens")
r2 += 1
write_assumption(ws2, r2, 5, "Rent M4+ (physical space)", "EUR 2,500/mo", "Wellness space with smoothie bar")
r2 += 1
write_assumption(ws2, r2, 5, "Steuerberater", "EUR 400/mo", "Tax advisor")
r2 += 1
write_assumption(ws2, r2, 5, "Insurance", "EUR 200/mo", "Business insurance")
r2 += 1
write_assumption(ws2, r2, 5, "SaaS Stack M1-6", "EUR 338/mo", "D9: Google, Figma, Notion, Crisp")
r2 += 1
write_assumption(ws2, r2, 5, "SaaS Stack M7-12", "EUR 500/mo", "D9: + tools as team grows")
r2 += 1
write_assumption(ws2, r2, 5, "SaaS Stack M13+", "EUR 750/mo", "D9: full stack")

r2 += 2
ws2.cell(row=r2, column=5, value="CapEx (D5)").font = font_section
ws2.cell(row=r2, column=5).fill = fill_section
ws2.cell(row=r2, column=6).fill = fill_section
ws2.cell(row=r2, column=7).fill = fill_section
r2 += 1
write_assumption(ws2, r2, 5, "GmbH Formation (M1)", "EUR 2,500", "One-off")
r2 += 1
write_assumption(ws2, r2, 5, "Legal Opinion HWG (M1)", "EUR 3,500", "One-off")
r2 += 1
write_assumption(ws2, r2, 5, "R&D Stability Testing (M1)", "EUR 5,000", "One-off, moved from OpEx")
r2 += 1
write_assumption(ws2, r2, 5, "Space Buildout (M3)", "EUR 15,000", "One-off")
r2 += 1
write_assumption(ws2, r2, 5, "Kitchen Equipment (M3)", "EUR 20,000", "One-off, smoothie bar")
r2 += 1
write_assumption(ws2, r2, 5, "LED Therapy Devices (M3)", "EUR 5,000", "One-off")
r2 += 1
write_assumption(ws2, r2, 5, "TOTAL CapEx", "EUR 51,000", "10.2% of EUR 500K funding")


# ═══════════════════════════════════════════
# SHEET 3: DECISIONS LOG
# ═══════════════════════════════════════════
ws3 = wb.create_sheet("Decisions")
ws3.sheet_properties.tabColor = TERRA
ws3.column_dimensions["A"].width = 8
ws3.column_dimensions["B"].width = 28
ws3.column_dimensions["C"].width = 50
ws3.column_dimensions["D"].width = 25
ws3.column_dimensions["E"].width = 35

ws3.cell(row=1, column=1, value="alche — Decision Log (D1-D11)").font = font_title
ws3.cell(row=2, column=1, value="All decisions made in P&L session, 2026-02-23").font = font_note

# Headers
for col, header in enumerate(["#", "Decision", "Choice Made", "Old Value", "P&L Impact"], 1):
    cell = ws3.cell(row=4, column=col, value=header)
    cell.font = font_header
    cell.fill = fill_header

decisions = [
    ["D1", "Doctor Revenue Model", "C. SaaS only EUR 99/mo × 20 clinics", "EUR 150/mo × 865 clinics", "Revenue drops ~75%; legally safe"],
    ["D2", "Restaurant Revenue", "B. Phase 2 trigger (500+ users)", "EUR 1.50/cover from M6", "EUR 0 for all 24 months"],
    ["D3", "CGM Hardware", "A. Software-Only BYOD", "EUR 130/unit with EUR 70 COGS", "Revenue removed; margin improves to 80-90%"],
    ["D4", "Physical Space", "A. Add smoothie bar + space", "EUR 0 (excluded)", "New revenue + EUR 2,500/mo rent + CapEx"],
    ["D5", "CapEx Section", "Yes, include CapEx", "EUR 0 (all OpEx)", "EUR 51,000 total CapEx"],
    ["D6", "AI Cost Model", "Both internal + user-facing", "EUR 210 + EUR 499 Terra", "EUR 210 + EUR 0.35/user COGS (R9)"],
    ["D7", "Product Waste", "A. Category-weighted", "Flat 5%", "5.5% blended (capsules 4%, powders 7%)"],
    ["D8", "Terra CGM API", "Defer — use free HealthKit", "EUR 499/mo from M6", "Saves EUR 9,481 over 24mo"],
    ["D9", "SaaS Stack", "Itemized per R8", "EUR 350→850→1800", "EUR 338→500→750"],
    ["D10", "Menu Analysis", "EUR 0 — founders do it", "EUR 500/mo", "Saves EUR 12,000 over 24mo"],
    ["D11", "Retail Products", "EUR 49 price validated; COGS EUR 23.50", "EUR 11.50 COGS", "Margin drops 76% → 52%; break-even shifts"],
]

for i, dec in enumerate(decisions):
    row = 5 + i
    for col, val in enumerate(dec, 1):
        cell = ws3.cell(row=row, column=col, value=val)
        cell.font = font_sub
        if col == 1:
            cell.font = Font(name="Calibri", size=10, bold=True, color=TERRA)
        cell.border = thin_border


# ═══════════════════════════════════════════
# SHEET 4: GROWTH CURVES
# ═══════════════════════════════════════════
ws4 = wb.create_sheet("Growth Curves")
ws4.sheet_properties.tabColor = SAGE
ws4.column_dimensions["A"].width = 24
for col in range(2, 27):
    ws4.column_dimensions[get_column_letter(col)].width = 9

ws4.cell(row=1, column=1, value="alche — Monthly Growth Curves").font = font_title

# Header
r = 3
ws4.cell(row=r, column=1, value="Metric")
for i in range(24):
    ws4.cell(row=r, column=i+2, value=f"M{i+1}")
ws4.cell(row=r, column=26, value="M24")
for c in range(1, 27):
    ws4.cell(row=r, column=c).font = font_header
    ws4.cell(row=r, column=c).fill = fill_header

curves = {
    "Potion Units": pot_units,
    "LED Sessions": led_sessions,
    "Smoothie Units": smoothie_units,
    "Event Attendees": [events[m][0] * events[m][1] for m in range(24)],
    "Doctor Clinics": doc_saas,
    "Restaurant Covers": rest_covers,
    "Total Subscribers": [months[m]["total_subs"] for m in range(24)],
    "Marketing Spend": marketing,
    "Paid Ads": marketing,
}

r = 4
for name, values in curves.items():
    ws4.cell(row=r, column=1, value=name).font = font_sub
    for col, val in enumerate(values, start=2):
        ws4.cell(row=r, column=col, value=val).font = font_sub
        ws4.cell(row=r, column=col).alignment = align_right
    r += 1


# ═══════════════════════════════════════════
# SUMMARY STATS (back to P&L sheet)
# ═══════════════════════════════════════════
# Add summary box at the top right of P&L
ws_pnl = wb["P&L"]
summary_col = 28  # Column AB

ws_pnl.column_dimensions[get_column_letter(summary_col)].width = 28
ws_pnl.column_dimensions[get_column_letter(summary_col+1)].width = 16

stats = [
    ("KEY SUMMARY", ""),
    ("Funding", "EUR 500,000"),
    ("Valuation Cap", "EUR 2,500,000"),
    ("Total CapEx", f"EUR {sum(months[m]['total_capex'] for m in range(24)):,}"),
    ("", ""),
    ("Y1 Revenue", f"EUR {sum(months[m]['total_income'] for m in range(12)):,}"),
    ("Y2 Revenue", f"EUR {sum(months[m]['total_income'] for m in range(12, 24)):,}"),
    ("24-Month Revenue", f"EUR {sum(months[m]['total_income'] for m in range(24)):,}"),
    ("", ""),
    ("Y1 Net P&L", f"EUR {sum(months[m]['total_pnl'] for m in range(12)):,}"),
    ("Y2 Net P&L", f"EUR {sum(months[m]['total_pnl'] for m in range(12, 24)):,}"),
    ("24-Month Net P&L", f"EUR {sum(months[m]['total_pnl'] for m in range(24)):,}"),
    ("", ""),
    ("M1 Burn", f"EUR {abs(months[0]['total_pnl']):,}"),
    ("M12 Burn", f"EUR {abs(months[11]['total_pnl']):,}"),
    ("M24 Subs", f"{months[23]['total_subs']}"),
    ("M24 MRR", f"EUR {months[23]['sub_total']:,}"),
    ("", ""),
    ("Cash Remaining M24", f"EUR {cumulative[23]:,}"),
    ("Runway (months)", "24" if cumulative[23] > 0 else f"~{next((i+1 for i in range(24) if cumulative[i] < 0), 24)}"),
]

# Find break-even month
be_month = "N/A"
for i in range(24):
    if months[i]["ebitda"] >= 0:
        be_month = f"M{i+1}"
        break

stats.append(("EBITDA Break-even", be_month))

for i, (label, value) in enumerate(stats):
    r = 3 + i
    if label == "KEY SUMMARY":
        ws_pnl.cell(row=r, column=summary_col, value=label).font = font_section
        ws_pnl.cell(row=r, column=summary_col).fill = fill_section
        ws_pnl.cell(row=r, column=summary_col+1).fill = fill_section
    elif label:
        ws_pnl.cell(row=r, column=summary_col, value=label).font = font_assumption_label
        ws_pnl.cell(row=r, column=summary_col+1, value=value).font = font_assumption_val
        ws_pnl.cell(row=r, column=summary_col+1).alignment = align_right


# ─── SAVE ───
output_path = "/Users/timoel/Downloads/pitchbook/pnl/alche-pnl-24mo.xlsx"
wb.save(output_path)
print(f"Excel P&L saved to: {output_path}")

# Print summary
y1_rev = sum(months[m]["total_income"] for m in range(12))
y2_rev = sum(months[m]["total_income"] for m in range(12, 24))
y1_pnl = sum(months[m]["total_pnl"] for m in range(12))
y2_pnl = sum(months[m]["total_pnl"] for m in range(12, 24))
total_capex = sum(months[m]["total_capex"] for m in range(24))

print(f"\n{'='*50}")
print(f"alche P&L SUMMARY (All 11 Decisions Applied)")
print(f"{'='*50}")
print(f"Y1 Revenue:      EUR {y1_rev:>10,}")
print(f"Y2 Revenue:      EUR {y2_rev:>10,}")
print(f"24-Mo Revenue:   EUR {y1_rev + y2_rev:>10,}")
print(f"{'─'*50}")
print(f"Y1 Net P&L:      EUR {y1_pnl:>10,}")
print(f"Y2 Net P&L:      EUR {y2_pnl:>10,}")
print(f"24-Mo Net P&L:   EUR {y1_pnl + y2_pnl:>10,}")
print(f"{'─'*50}")
print(f"Total CapEx:     EUR {total_capex:>10,}")
print(f"Cash M24:        EUR {cumulative[23]:>10,}")
print(f"M24 Subscribers: {months[23]['total_subs']:>10}")
print(f"M24 MRR:         EUR {months[23]['sub_total']:>10,}")
print(f"EBITDA Break-even: {be_month}")
print(f"{'─'*50}")
print(f"M1 Burn:         EUR {abs(months[0]['total_pnl']):>10,}")
print(f"M6 Burn:         EUR {abs(months[5]['total_pnl']):>10,}")
print(f"M12 Burn:        EUR {abs(months[11]['total_pnl']):>10,}")
print(f"M18 Burn:        EUR {abs(months[17]['total_pnl']):>10,}")
print(f"M24 P&L:         EUR {months[23]['total_pnl']:>10,}")
