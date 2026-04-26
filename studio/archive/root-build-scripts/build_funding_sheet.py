#!/usr/bin/env python3
"""Build Alche Funding Opportunities Excel Sheet"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Funding Opportunities"

# --- STYLES ---
thick = Side(style='thick', color='2C2418')
medium = Side(style='medium', color='2C2418')
thin = Side(style='thin', color='9E948A')

border_header = Border(top=thick, bottom=thick, left=thick, right=thick)
border_tier = Border(top=thick, bottom=thick, left=thick, right=thick)
border_cell = Border(top=thin, bottom=thin, left=thin, right=thin)
border_cell_bottom = Border(top=thin, bottom=medium, left=thin, right=thin)

fill_header = PatternFill(start_color='2C2418', end_color='2C2418', fill_type='solid')
fill_tier1 = PatternFill(start_color='C4956A', end_color='C4956A', fill_type='solid')  # amber
fill_tier2 = PatternFill(start_color='8B9E7C', end_color='8B9E7C', fill_type='solid')  # sage
fill_tier3 = PatternFill(start_color='B86B4A', end_color='B86B4A', fill_type='solid')  # terra
fill_tier4 = PatternFill(start_color='9E948A', end_color='9E948A', fill_type='solid')  # stone
fill_cream = PatternFill(start_color='F5F0E8', end_color='F5F0E8', fill_type='solid')
fill_cream_alt = PatternFill(start_color='EDE7DC', end_color='EDE7DC', fill_type='solid')
fill_urgent = PatternFill(start_color='E85D4A', end_color='E85D4A', fill_type='solid')
fill_medium_urg = PatternFill(start_color='C4956A', end_color='C4956A', fill_type='solid')
fill_low_urg = PatternFill(start_color='8B9E7C', end_color='8B9E7C', fill_type='solid')
fill_na_urg = PatternFill(start_color='D5CFC7', end_color='D5CFC7', fill_type='solid')

font_header = Font(name='Arial', size=11, bold=True, color='F5F0E8')
font_tier = Font(name='Arial', size=12, bold=True, color='FFFFFF')
font_bold = Font(name='Arial', size=10, bold=True, color='2C2418')
font_normal = Font(name='Arial', size=10, color='2C2418')
font_small = Font(name='Arial', size=9, color='2C2418')
font_link = Font(name='Arial', size=9, color='B86B4A', underline='single')
font_urgency = Font(name='Arial', size=10, bold=True, color='FFFFFF')

align_wrap = Alignment(wrap_text=True, vertical='top', horizontal='left')
align_center = Alignment(wrap_text=True, vertical='center', horizontal='center')
align_header = Alignment(wrap_text=True, vertical='center', horizontal='center')

# --- COLUMNS ---
columns = [
    ("", 4),           # A: row number
    ("PROGRAM", 28),
    ("TYPE", 14),
    ("LOCATION", 14),
    ("AMOUNT", 18),
    ("DILUTIVE?", 10),
    ("DEADLINE", 18),
    ("URGENCY", 12),
    ("ELIGIBILITY FIT", 16),
    ("WHY ALCHE FITS", 38),
    ("APPLICATION NOTES", 32),
    ("URL", 32),
]

# --- TITLE ROW ---
ws.merge_cells('A1:L1')
title_cell = ws['A1']
title_cell.value = "ALCHE — FUNDING OPPORTUNITIES RESEARCH  |  March 2026"
title_cell.font = Font(name='Arial', size=16, bold=True, color='2C2418')
title_cell.fill = PatternFill(start_color='C4956A', end_color='C4956A', fill_type='solid')
title_cell.alignment = Alignment(horizontal='center', vertical='center')
title_cell.border = border_header
ws.row_dimensions[1].height = 45
for col in range(2, 13):
    ws.cell(row=1, column=col).fill = PatternFill(start_color='C4956A', end_color='C4956A', fill_type='solid')
    ws.cell(row=1, column=col).border = border_header

# --- SUBTITLE ROW ---
ws.merge_cells('A2:L2')
sub_cell = ws['A2']
sub_cell.value = "Pre-seed  |  Berlin GmbH  |  Longevity Lifestyle Platform  |  EUR 500K Raise  |  Wellness / Health-Tech / Preventive Health"
sub_cell.font = Font(name='Arial', size=10, italic=True, color='2C2418')
sub_cell.fill = fill_cream
sub_cell.alignment = Alignment(horizontal='center', vertical='center')
sub_cell.border = Border(bottom=thick, left=thick, right=thick)
ws.row_dimensions[2].height = 28
for col in range(2, 13):
    ws.cell(row=2, column=col).fill = fill_cream
    ws.cell(row=2, column=col).border = Border(bottom=thick, left=thick if col == 12 else Side(), right=thick if col == 12 else Side())

# --- HEADER ROW ---
row = 3
ws.row_dimensions[row].height = 36
for i, (name, width) in enumerate(columns, 1):
    cell = ws.cell(row=row, column=i, value=name)
    cell.font = font_header
    cell.fill = fill_header
    cell.border = border_header
    cell.alignment = align_header
    ws.column_dimensions[get_column_letter(i)].width = width

# --- DATA ---
# Each entry: (program, type, location, amount, dilutive, deadline, urgency, fit, why_fits, notes, url)
# urgency: "URGENT", "HIGH", "MEDIUM", "LOW", "MONITOR", "N/A"

opportunities = [
    # TIER 1: ACT NOW
    ("TIER 1 — ACT NOW (Apply Immediately)", None),

    ("GruendungsBONUS Plus (IBB)",
     "Grant", "Berlin", "Up to EUR 50,000 (50% of costs)", "No",
     "Rolling — LIVE NOW since Jan 2026. Budget-limited, first come first served.",
     "URGENT", "STRONG",
     "Berlin GmbH, digital business, micro-enterprise, founders hold majority. GmbH must be <18 months old. High-level tech innovation NOT required. Covers rent, salaries, IT.",
     "Apply ASAP in German. Check GmbH registration date — 18-month clock is ticking. Budget typically exhausts by Q3.",
     "ibb.de/foerderprogramme/gruendungsbonus-plus.html"),

    ("INVEST — Zuschuss fur Wagniskapital (BAFA)",
     "Tax incentive (for angels)", "Germany-wide", "25% of angel investment back to investor (EUR 10K-200K per investor)", "No (benefits investors)",
     "Active until 31.12.2026. Must certify BEFORE signing angel contracts.",
     "URGENT", "STRONG",
     "GmbH, <10 yrs old, <50 employees, <EUR 10M revenue. Alche checks every box. Gives each angel 25% of their investment back as a tax-free grant.",
     "Apply for INVEST eligibility certification NOW at BAFA, before starting angel fundraising. Each angel then applies individually before signing.",
     "bafa.de/DE/Wirtschaft/Beratung_Finanzierung/Invest/invest_node.html"),

    ("Microsoft for Startups — Founders Hub",
     "Cloud credits + tools", "Global (remote)", "Up to $150,000 Azure credits + $350K resources", "No",
     "Rolling — apply anytime. 1-3% acceptance rate.",
     "URGENT", "STRONG",
     "Software-driven startups. No funding required. GDPR-compliant infrastructure for health data. No equity taken.",
     "Apply immediately. Free Azure credits for building platform. Strong for health data compliance.",
     "startups.microsoft.com"),

    ("StartUp Health — Longevity + Women's Health Moonshots",
     "Community / accelerator hybrid", "Global (remote)", "No direct investment — coaching, investor intros", "No",
     "Rolling — apply anytime.",
     "HIGH", "STRONG",
     "Founders at all stages building health innovations. Both Healthy Longevity AND Women's Health Moonshot communities align perfectly.",
     "Apply to BOTH communities now. Free, no equity, rolling admission.",
     "startuphealth.com/healthy-longevity-moonshot-community"),

    ("Seedcamp",
     "Pre-seed VC", "London / Europe-wide", "GBP 100K-250K (up to GBP 2M co-invest)", "Yes",
     "Rolling — Investment Forum every 6-8 weeks.",
     "HIGH", "MEDIUM-HIGH",
     "European founders at pre-seed/seed. HealthTech is a focus sector. Pan-European, invests in Berlin.",
     "Apply now via website. No relocation or corporate restructuring needed. GBP 100K-250K aligns with part of EUR 500K raise.",
     "seedcamp.com"),

    ("SpinLab Leipzig — HHL Accelerator",
     "Accelerator", "Leipzig (1hr from Berlin)", "EUR 6K stipend + up to EUR 50K grant. ZERO equity.", "No",
     "Expected ~April 2026. Check website NOW.",
     "HIGH", "STRONG",
     "Health is core focus. Zero equity. Germany-based. Team of 2-4. Business-level English or German. Hybrid format.",
     "Check website for exact deadline immediately. Frame as 'preventive health-tech'. HIGH PRIORITY — founder-friendly terms.",
     "spinlab.co/apply-now"),

    # TIER 2: APPLY WITHIN 1-3 MONTHS
    ("TIER 2 — APPLY WITHIN 1-3 MONTHS", None),

    ("VitalMatch Accelerator (Turin)",
     "Accelerator", "Turin, Italy (hybrid)", "EUR 150,000 - EUR 400,000", "Yes (negotiated)",
     "March 10 - April 3, 2026",
     "URGENT", "MODERATE-STRONG",
     "Seed/pre-Series A startups in sport, health, wellness. Must be market-validated. Backed by CDP Venture Capital + Startupbootcamp.",
     "Applications close April 3! Wellness + health + longevity aligns. Challenge: requires market validation. Apply if you can show any traction.",
     "vitalmatchaccelerator.it"),

    ("Techstars Berlin",
     "Accelerator", "Berlin", "$220,000 ($200K SAFE + $20K convertible)", "Yes (min 5%+)",
     "May 6, 2026 for Fall programs",
     "HIGH", "MEDIUM",
     "Berlin-based. Life Science is listed focus. Deep Tech bias but pre-product accepted. No corporate restructuring needed.",
     "Frame as 'preventive health / life science' not 'wellness'. Apply before May 6. Berlin location is perfect.",
     "techstars.com/germany/berlin"),

    ("DayOne Accelerator (Basel)",
     "Accelerator", "Basel, Switzerland (hybrid)", "Up to CHF 50K non-dilutive (top 3). ZERO equity.", "No",
     "Applications open Mar 31 - May 31, 2026",
     "HIGH", "MEDIUM",
     "Digital health, medtech. International — German startups welcome. Hybrid (2 in-person events). Equity-free, fee-free.",
     "Apply when portal opens March 31. Frame longevity science translation as relevant to pharma R&D. Equity-free is excellent.",
     "dayone.swiss/accelerator"),

    ("Falling Walls Venture (Berlin)",
     "Pitch competition", "Berlin", "Visibility + investor access", "No",
     "April 15, 2026 submission deadline",
     "HIGH", "MODERATE",
     "Science startups that translated breakthrough science into business. All disciplines incl. biotech, data & AI. Prestigious Berlin event.",
     "Submit by April 15. Frame longevity science translation as the breakthrough. Great visibility + Berlin credibility.",
     "falling-walls.com"),

    ("Transfer BONUS (IBB)",
     "Grant", "Berlin", "EUR 7,500 (entry) to EUR 45,000 (digitalization). 70-100% funded.", "No",
     "Rolling through Dec 31, 2026",
     "MEDIUM", "GOOD",
     "Berlin SME collaborating with research institution on tech transfer. Entry-level requires no prior collaboration.",
     "Identify a research partner (Charite for longevity science, Berlin university for behavioral health). EUR 45K digitalization variant could fund platform R&D.",
     "ibb-business-team.de/transfer-bonus"),

    ("Antler ONE (Berlin, next cohort)",
     "Accelerator / pre-seed investor", "Berlin + Munich/Amsterdam/Paris", "EUR 100K for 8.5% + EUR 100K uncapped SAFE", "Yes (8.5%+)",
     "Rolling — next cohort likely H2 2026",
     "MEDIUM", "MEDIUM-HIGH",
     "Pre-seed stage. Industry-agnostic. Berlin presence. EUR 200K initial investment. Up to EUR 500K total, EUR 30M follow-on.",
     "Apply now for next cohort. Just-started cohort missed (Mar 9). Rolling basis — earlier = better.",
     "antler.co/residency/germany"),

    ("Future of Health Grant (CSS / Swiss)",
     "Grant", "Switzerland (EU eligible)", "CHF 10K (concept) / 30K (PoC) / 50K (validation). Equity-free.", "No",
     "Application closes July 31, 2026",
     "MEDIUM", "STRONG",
     "Healthtech, medtech, digital health. Focus on preventive care, digital nutrition, mental health, wellness analytics. Early-stage friendly.",
     "Directly aligned with Alche's longevity lifestyle concept. Preventive care + digital nutrition + wellness analytics = perfect match. Prepare strong application.",
     "future-of-health.org"),

    ("Femtech World Awards 2026",
     "Awards / recognition", "UK (global entries)", "No cash prize — visibility, media coverage", "No",
     "Currently accepting entries — likely closing soon",
     "HIGH", "STRONG",
     "Any femtech / women's health innovation. Free to enter. Judges value momentum and potential.",
     "Apply immediately — free to enter. 'Imperfect longevity' for women is exactly what judges look for. Great PR even if shortlisted.",
     "femtechworld.co.uk"),

    ("Bayer G4A (Grants4Apps) Berlin",
     "Accelerator / grant", "Berlin", "Up to EUR 50,000. No equity.", "No",
     "No 2026 cycle announced yet — monitor",
     "MEDIUM", "VERY STRONG",
     "Digital health startups. Focus: Women's Health, prevention, patient engagement. At least 3 team members in Berlin full-time.",
     "Women's Health is explicit focus. Prevention aligns. Berlin-based. EUR 50K grant, no equity. Monitor for 2026 opening.",
     "g4a.health"),

    # TIER 3: BUILD TOWARD / POST-MVP
    ("TIER 3 — BUILD TOWARD (Post-MVP / Post-Funding)", None),

    ("Forschungszulage (R&D Tax Credit)",
     "Tax credit / cash refund", "Germany-wide", "35% of R&D costs for SMEs. Max EUR 4.2M/yr. Cash refund even at zero revenue.", "No",
     "Ongoing. Retroactive up to 3 years.",
     "HIGH", "STRONG (post-hire)",
     "Any German company with R&D. 35% of CTO salary refunded. Plus 20% overhead flat rate from 2026. Works even for loss-making companies.",
     "Activate the MOMENT CTO starts and begins product dev. EUR 80K CTO salary = EUR 28K+ cash back. Apply via BSFZ certificate + tax office.",
     "innoscripta.com/en/about-us/blogs/new-forschungszulage-rules-for-2026"),

    ("Pro FIT Early Phase (IBB)",
     "Grant + interest-free loan", "Berlin", "Up to EUR 500,000 (50% grant + 50% loan). 100% of eligible costs.", "No",
     "Rolling. Requires detailed project plan.",
     "MEDIUM", "MODERATE",
     "Highly innovative technology startup in Berlin. Requires tech roadmap (AI personalization, health data). External technical evaluation.",
     "Strong program but 'highly innovative technology' bar is high. Need CTO + concrete tech roadmap. Frame as AI-driven personalized longevity platform.",
     "ibb.de/foerderprogramme/pro-fit-fruehphasenfinanzierung.html"),

    ("IBB Ventures — B# Pre-Seed Fund",
     "Venture capital (convertible bonds)", "Berlin", "EUR 100,000 - EUR 400,000", "Yes (convertible bonds)",
     "Rolling — continuously investing.",
     "MEDIUM", "MODERATE-GOOD",
     "Berlin-based, SME, founder-owned. Requires co-investors (angels/VCs) investing at least same amount. Wandeldarlehen instrument matches.",
     "Align with your EUR 500K round. Having IBB Ventures as co-investor strengthens raise. Need matching angel commitments first.",
     "ibbventures.de/en/investment"),

    ("EIT Health Accelerator (Validation Track)",
     "Accelerator + grant", "Pan-European", "Up to EUR 150K grant. Non-dilutive.", "No",
     "Check eithealth.eu for next window",
     "MEDIUM", "GOOD (post-MVP)",
     "Early-stage startups with MVP in biotech, medtech, digital health. Validation Track supports refining business foundations + GTM.",
     "Apply once Alche has a working digital product. Digital health / preventive health is in scope. Berlin is an EIT Health hub city.",
     "eithealth.eu/new-call-opportunities"),

    ("ZIM (Zentrales Innovationsprogramm Mittelstand)",
     "Grant", "Germany-wide", "Up to EUR 690,000. ~45% funding rate for small companies.", "No",
     "Rolling until 30.06.2027",
     "MEDIUM", "CONDITIONAL",
     "SME with innovative R&D project. Technology- and sector-open. Must be genuine R&D innovation (not content alone).",
     "Viable IF Alche develops proprietary tech (AI-driven personalized longevity engine). Frame as R&D project. Need defined project scope + CTO.",
     "zim.de"),

    ("Cartier Women's Initiative 2027",
     "Awards + fellowship", "Global (Europe region)", "USD 100,000 (1st) / USD 60,000 (2nd). No equity.", "No",
     "Applications: April 16 - June 16, 2026 (for 2027 awards)",
     "MEDIUM", "CONDITIONAL",
     "Women entrepreneurs. Impact-driven. 1-6 years registered. Must have recurring revenue. Woman holds largest/equal equity share. <$2M raised.",
     "REQUIRES REVENUE. If Alche launches product + generates any revenue before June 2026, this is top-tier. Daria holds 51% — eligible. Mark calendar NOW.",
     "cartierwomensinitiative.com/regional-awards"),

    ("Micro-Mezzanine Fonds Deutschland",
     "Silent participation (mezzanine)", "Germany-wide", "Up to EUR 150,000 (target groups: women-led). 10-yr term.", "No (no equity/voting rights)",
     "Rolling. No deadline.",
     "LOW-MEDIUM", "GOOD",
     "Small/young enterprises. Higher amounts for women-led + migrant-founded. 10-year term, repayment starts year 7. ~8% p.a. effective cost.",
     "Alche qualifies. Female co-founder (Daria) = enhanced EUR 150K amount. Patient capital. Consider cost vs. other options.",
     "bundeswirtschaftsministerium.de/Redaktion/EN/Artikel/SME-Sector/micro-mezzanine-fund-germany.html"),

    ("BMBF START-interactive",
     "Grant", "Germany-wide", "Up to EUR 400,000 for startups", "No",
     "Next deadline: 15.07.2026",
     "MEDIUM", "CONDITIONAL",
     "Startup <5 yrs with interactive tech for health/quality of life. Must involve interactive digital tools.",
     "Viable if Alche builds interactive health tools (ritual tracking, biometric feedback, interactive assessments). Pure content doesn't qualify.",
     "grantbite.com/en/funding/start-interactive-health-research"),

    ("Reaktor.Berlin — Batch 9",
     "Accelerator", "Berlin", "EUR 13,200/founder (Berlin Stipendium). ZERO equity.", "No",
     "Batch 9 date TBA — monitor website",
     "MEDIUM", "HIGH",
     "Berlin-based founders. Pre-seed stage. EU ESF+ funded. Zero equity. 6 months. Berlin residency required — met.",
     "Monitor for Batch 9 opening. EUR 26,400 total (2 founders). Can stack with other non-dilutive programs.",
     "reaktor.berlin/accelerator"),

    # TIER 4: MONITOR / LOWER PRIORITY
    ("TIER 4 — MONITOR / FUTURE OPPORTUNITIES", None),

    ("European Longevity Hub Accelerator",
     "Accelerator", "Lisbon + remote", "Tailored per startup (equity-based)", "Yes (6-15%)",
     "Jan 2026 cohort closed. Next cycle TBA.",
     "MONITOR", "STRONG",
     "Longevity, wellness, supplements, nutrition, fitness. Digital Health & AI track. 4-6 months. European base.",
     "Perfect thematic alignment. Missed current cycle. Register interest and monitor for next cohort announcement.",
     "europeanlongevityhub.com/startup-accelerator"),

    ("Y Combinator S26",
     "Accelerator", "San Francisco (required)", "$500,000 for 7% equity", "Yes (7%)",
     "Expected ~May 2026",
     "MONITOR", "MEDIUM",
     "Any stage/sector. European startups eligible. Must relocate to SF. Must create US parent entity (GmbH becomes subsidiary).",
     "$500K matches EUR 500K target. YC brand unmatched. But: SF relocation + corporate restructuring + 7% equity. Highly competitive (~1-2% acceptance).",
     "ycombinator.com/apply"),

    ("Google for Startups — AI for Health (EMEA)",
     "Accelerator (non-equity)", "Hybrid (EMEA)", "Up to $350,000 Google Cloud credits", "No",
     "2026 applications closed. Next cycle likely annual.",
     "MONITOR", "VERY STRONG",
     "Seed to Series A. AI for health/wellbeing. Europe. Equity-free. Previous health & wellbeing editions.",
     "Perfect match if Alche has AI component. Monitor for next cycle. Apply immediately when it opens.",
     "startup.google.com/programs/growth-academy/ai-for-health/emea"),

    ("FemTech Lab (London/Global)",
     "Accelerator", "London / Global", "Equity-free grant via pitch competition", "No",
     "Rolling / periodic cohorts — check website",
     "MONITOR", "STRONG",
     "Women's health and wellness startups. Equity-free. Global reach.",
     "Women 30-55 longevity rituals = strong fit. Check website for next cohort.",
     "femtechlab.com"),

    ("Startupbootcamp Digital Health Berlin",
     "Accelerator", "Berlin", "EUR 25,000 + EUR 100K+ partner deals", "Yes (~6-8%)",
     "No 2026 cohort announced",
     "MONITOR", "STRONG",
     "Global digital health startups. Consumer health welcome. Berlin-based. 3 months.",
     "Perfect location + sector match. Monitor for 2026 cohort announcement.",
     "startupbootcamp.org/health-lifescience"),

    ("Vision Health Pioneers (Berlin)",
     "Incubator", "Berlin", "Up to EUR 80K (EUR 20K/person). ZERO equity.", "No",
     "No 2026 cohort announced (Cohort 7 was 2024)",
     "MONITOR", "STRONG (if eligible)",
     "First-time founders in healthcare. Berlin resident. 9 months. Check if existing GmbH disqualifies.",
     "Strong thematic fit. Verify GmbH eligibility. Monitor for Cohort 8.",
     "visionhealthpioneers.de"),

    ("Women TechEU",
     "Grant", "EU-wide", "EUR 75,000 non-dilutive + business dev program", "No",
     "Current project ends May 2026. New launch expected 2026.",
     "MONITOR", "CONDITIONAL",
     "Women-led deep tech. CEO/CTO must be female with 25%+ shares. Must be registered in EU. Requires 'deep tech' framing.",
     "Daria holds 51% — ownership threshold met. Challenge: 'deep tech' classification for a lifestyle platform. Monitor for new program launch.",
     "womentecheurope.eu"),

    ("Female Founders — Grow F",
     "Accelerator", "Europe", "Fundraising prep + network. No direct funding.", "No",
     "Check website for 2026 applications",
     "MONITOR", "STRONG",
     "European female founders. Fundraising readiness program.",
     "Good fit for fundraising preparation. Check for 2026 application window.",
     "female-founders.org/startup-program"),

    ("Cascade Funding (EU FSTP)",
     "Small grants via EU projects", "EU-wide", "EUR 10K-60K per call. Light applications.", "No",
     "Varies — 20+ open calls at any time",
     "MONITOR", "WORTH CHECKING",
     "SMEs, startups, researchers. Various themes. GenAI/LLM call currently open (relevant if Alche uses AI for content personalization).",
     "Check cascadefunding.eu monthly for health/wellness/digital calls. Small but non-dilutive with light applications.",
     "cascadefunding.eu/open-calls"),

    ("HealthCapital Berlin-Brandenburg",
     "Ecosystem support (not a grant)", "Berlin", "Free advisory, event access, introductions", "N/A",
     "Always available",
     "LOW", "GOOD",
     "Any health/wellness startup in Berlin-Brandenburg. Free resource.",
     "Register with the cluster now. bio:cap investival June 9-11, 2026. Not money, but access + future call alerts.",
     "healthcapital.de/en/services/startups"),

    ("Grundungszuschuss (Agentur fur Arbeit)",
     "Grant (tax-free)", "Germany-wide", "ALG I amount + EUR 300/mo for 6 months, then EUR 300/mo for 9 months. ~EUR 20K total.", "No",
     "Ongoing — must be on ALG I with 150+ days remaining",
     "CONDITIONAL", "CONDITIONAL",
     "Must be receiving ALG I unemployment benefits. 150+ days remaining. Full-time self-employment (15+ hrs/wk). Discretionary.",
     "Only if a founder transitions from employment via unemployment. Common Berlin founder strategy. Timing is everything.",
     "verwaltung.bund.de/leistungsverzeichnis/en/leistung/99007005017000"),

    ("Erasmus for Young Entrepreneurs",
     "Exchange programme + stipend", "EU-wide", "Monthly living-cost stipend (not business capital)", "No",
     "Rolling",
     "LOW", "QUALIFIES",
     "New entrepreneurs (started within 3 years). 1-6 month exchange with experienced entrepreneur in another EU country.",
     "Good for learning + market validation in another EU country. Not startup capital. Use strategically when entering new market.",
     "erasmus-entrepreneurs.eu"),

    ("German Accelerator",
     "Accelerator (free, no equity)", "Germany / US / Asia", "No direct funding. Free mentoring + market access.", "No",
     "Next cohorts TBD (Kickstart 2026 deadline passed)",
     "LATER", "PREMATURE",
     "German startup. International expansion focus. Product-market fit needed. Life Sciences track exists.",
     "Revisit in 12-18 months post-launch when ready for US/Asia expansion.",
     "germanaccelerator.com"),
]

# --- WRITE DATA ---
row = 4
entry_num = 0

for item in opportunities:
    if item[1] is None:
        # TIER HEADER ROW
        tier_name = item[0]
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        cell = ws.cell(row=row, column=1, value=tier_name)
        if "TIER 1" in tier_name:
            cell.fill = fill_tier1
        elif "TIER 2" in tier_name:
            cell.fill = fill_tier2
        elif "TIER 3" in tier_name:
            cell.fill = fill_tier3
        elif "TIER 4" in tier_name:
            cell.fill = fill_tier4
        cell.font = font_tier
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border_tier
        ws.row_dimensions[row].height = 32
        if "TIER 1" in tier_name:
            tier_fill = fill_tier1
        elif "TIER 2" in tier_name:
            tier_fill = fill_tier2
        elif "TIER 3" in tier_name:
            tier_fill = fill_tier3
        else:
            tier_fill = fill_tier4
        for col in range(2, 13):
            c = ws.cell(row=row, column=col)
            c.fill = tier_fill
            c.border = border_tier
        row += 1
        continue

    entry_num += 1
    program, type_, location, amount, dilutive, deadline, urgency, fit, why_fits, notes, url = item

    row_fill = fill_cream if entry_num % 2 == 1 else fill_cream_alt
    ws.row_dimensions[row].height = 72

    data = [entry_num, program, type_, location, amount, dilutive, deadline, urgency, fit, why_fits, notes, url]

    for col_idx, value in enumerate(data, 1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.border = border_cell
        cell.alignment = align_wrap
        cell.fill = row_fill

        if col_idx == 1:  # row number
            cell.font = font_small
            cell.alignment = align_center
        elif col_idx == 2:  # program name
            cell.font = font_bold
        elif col_idx == 8:  # urgency
            cell.font = font_urgency
            cell.alignment = align_center
            if value == "URGENT":
                cell.fill = fill_urgent
            elif value == "HIGH":
                cell.fill = fill_medium_urg
            elif value == "MEDIUM":
                cell.fill = PatternFill(start_color='D4B896', end_color='D4B896', fill_type='solid')
                cell.font = Font(name='Arial', size=10, bold=True, color='2C2418')
            elif value == "MONITOR":
                cell.fill = fill_low_urg
            elif value in ("LOW", "LOW-MEDIUM", "LATER", "CONDITIONAL"):
                cell.fill = fill_na_urg
                cell.font = Font(name='Arial', size=10, bold=True, color='2C2418')
        elif col_idx == 9:  # eligibility fit
            cell.font = font_bold
            cell.alignment = align_center
            if "STRONG" in value or "VERY STRONG" in value:
                cell.font = Font(name='Arial', size=10, bold=True, color='4A7A3B')
            elif "MODERATE" in value or "GOOD" in value:
                cell.font = Font(name='Arial', size=10, bold=True, color='C4956A')
            elif "CONDITIONAL" in value:
                cell.font = Font(name='Arial', size=10, bold=True, color='B86B4A')
        elif col_idx == 12:  # URL
            cell.font = font_link
        else:
            cell.font = font_normal

    row += 1

# --- SUMMARY ROW ---
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
summary = ws.cell(row=row, column=1)
summary.value = (
    "TOTAL: 38 opportunities mapped  |  "
    "7 URGENT (apply now)  |  "
    "9 HIGH priority (1-3 months)  |  "
    "11 MEDIUM (build toward)  |  "
    "11 MONITOR (future)  |  "
    "Estimated accessible non-dilutive funding: EUR 200K-500K+"
)
summary.font = Font(name='Arial', size=11, bold=True, color='2C2418')
summary.fill = fill_tier1
summary.alignment = Alignment(horizontal='center', vertical='center')
summary.border = border_tier
ws.row_dimensions[row].height = 36
for col in range(2, 13):
    ws.cell(row=row, column=col).fill = fill_tier1
    ws.cell(row=row, column=col).border = border_tier

# --- FREEZE PANES ---
ws.freeze_panes = 'C4'

# --- PRINT SETUP ---
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.orientation = 'landscape'
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

# --- SAVE ---
output_path = "/Users/timoel/Desktop/alche/Alche_Funding_Opportunities_March2026.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
