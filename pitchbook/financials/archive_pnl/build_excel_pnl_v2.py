#!/usr/bin/env python3
"""
alche P&L Excel Generator v2 — Formula-Based, Persona-Audited
═══════════════════════════════════════════════════════════════
Changes from v1:
  - All computed cells use EXCEL FORMULAS (not hardcoded values)
  - Assumptions in named cells on dedicated sheet
  - Growth curves as data inputs on dedicated sheet
  - Each product section: units → price/unit → revenue (stacked, thick separators)
  - 11 new line items from 7-persona audit:
    CRITICAL: Utilities, Cleaning, BG Insurance, Contingency
    IMPORTANT: Refunds, IHK Berlin, Recruitment, POS, Depreciation, Packaging
  - D11 COGS updated to EUR 23.50 per R10
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ─── STYLE DEFINITIONS ───
DEEP = "2C2418"
AMBER = "C4956A"
SAGE = "8B9E7C"
TERRA = "B86B4A"
WHITE = "FFFFFF"
HEADER_BG = "3D3225"
SECTION_BG = "EDE7DC"
TOTAL_BG = "E5DDD3"
GRAND_BG = "D8CFC3"
PRODUCT_SEP = "C4956A"  # amber for product separators
LIGHT_BG = "FEFCF9"

f_title = Font(name="Calibri", size=16, bold=True, color=DEEP)
f_subtitle = Font(name="Calibri", size=9, italic=True, color="9E948A")
f_header = Font(name="Calibri", size=10, bold=True, color=WHITE)
f_section = Font(name="Calibri", size=10, bold=True, color=DEEP)
f_sub = Font(name="Calibri", size=10, color=DEEP)
f_dim = Font(name="Calibri", size=10, color="8A7F73")
f_unit = Font(name="Calibri", size=9, color="8A7F73")  # for unit/price rows
f_total = Font(name="Calibri", size=10, bold=True, color=DEEP)
f_grand = Font(name="Calibri", size=11, bold=True, color=DEEP)
f_note = Font(name="Calibri", size=9, italic=True, color="9E948A")
f_label = Font(name="Calibri", size=10, color="5C5244")
f_val = Font(name="Calibri", size=10, bold=True, color=DEEP)
f_decision = Font(name="Calibri", size=10, color=TERRA)
f_pct = Font(name="Calibri", size=9, color="8A7F73")

fill_header = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
fill_section = PatternFill(start_color=SECTION_BG, end_color=SECTION_BG, fill_type="solid")
fill_total = PatternFill(start_color=TOTAL_BG, end_color=TOTAL_BG, fill_type="solid")
fill_grand = PatternFill(start_color=GRAND_BG, end_color=GRAND_BG, fill_type="solid")

align_r = Alignment(horizontal="right", vertical="center")
align_l = Alignment(horizontal="left", vertical="center")
align_c = Alignment(horizontal="center", vertical="center")

thin_bottom = Border(bottom=Side(style="thin", color="D8CFC3"))
thick_bottom = Border(bottom=Side(style="medium", color=PRODUCT_SEP))
thick_top_bottom = Border(
    top=Side(style="medium", color=DEEP),
    bottom=Side(style="medium", color=DEEP),
)
product_sep = Border(bottom=Side(style="thick", color=PRODUCT_SEP))

NUM_FMT = '#,##0'
PCT_FMT = '0.0%'
EUR_FMT = '#,##0'

# ═══════════════════════════════════════════
# SHEET 1: ASSUMPTIONS
# ═══════════════════════════════════════════
ws_a = wb.active
ws_a.title = "Assumptions"
ws_a.sheet_properties.tabColor = SAGE
ws_a.column_dimensions["A"].width = 30
ws_a.column_dimensions["B"].width = 14
ws_a.column_dimensions["C"].width = 40

ws_a.cell(row=1, column=1, value="alche — P&L Assumptions").font = f_title
ws_a.cell(row=2, column=1, value="Change any value here → P&L sheet auto-recalculates").font = f_subtitle

# We'll place assumptions in column B, starting at row 4
# The P&L formulas will reference these cells as: Assumptions!B{row}
# Layout: A=label, B=value, C=note

def put_assumption(row, label, value, note=""):
    ws_a.cell(row=row, column=1, value=label).font = f_label
    c = ws_a.cell(row=row, column=2, value=value)
    c.font = f_val
    c.alignment = align_r
    if note:
        ws_a.cell(row=row, column=3, value=note).font = f_note

def section_header(row, label):
    for col in range(1, 4):
        c = ws_a.cell(row=row, column=col)
        c.fill = fill_section
    ws_a.cell(row=row, column=1, value=label).font = f_section

# ─── Subscription Model (rows 4-14) ───
r = 4
section_header(r, "SUBSCRIPTION MODEL"); r += 1
put_assumption(r, "Core tier price (EUR/mo)", 19, "Entry tier"); r += 1  # B5
put_assumption(r, "Pro tier price (EUR/mo)", 49, "Sweet spot"); r += 1   # B6
put_assumption(r, "Premium tier price (EUR/mo)", 99, "Highest ARPU"); r += 1  # B7
put_assumption(r, "Monthly churn rate", 0.08, "8% — locked data"); r += 1  # B8
ws_a.cell(row=r-1, column=2).number_format = '0%'
put_assumption(r, "Core tier share", 0.52, "52%"); r += 1  # B9
ws_a.cell(row=r-1, column=2).number_format = '0%'
put_assumption(r, "Pro tier share", 0.38, "38%"); r += 1   # B10
ws_a.cell(row=r-1, column=2).number_format = '0%'
put_assumption(r, "Premium tier share", 0.10, "10%"); r += 1  # B11
ws_a.cell(row=r-1, column=2).number_format = '0%'

# ─── Subscriber Growth (rows 16-22) ───
r = 13
section_header(r, "SUBSCRIBER GROWTH (new subs/month)"); r += 1
put_assumption(r, "M1-M3", 3, "Pre-launch beta"); r += 1   # B14
put_assumption(r, "M4-M6", 15, "Product launch"); r += 1    # B15
put_assumption(r, "M7-M9", 22, "Growth acceleration"); r += 1   # B16
put_assumption(r, "M10-M12", 28, "Marketing ramp"); r += 1  # B17
put_assumption(r, "M13-M18", 32, "Scaling"); r += 1         # B18
put_assumption(r, "M19-M24", 38, "Growth marketer effect"); r += 1  # B19

# ─── Product Pricing (rows 21-38) ───
r = 21
section_header(r, "PRODUCT PRICING"); r += 1
put_assumption(r, "Potion retail price (EUR)", 49, "R10 validated"); r += 1       # B22
put_assumption(r, "Potion COGS (EUR/unit)", 23.50, "R10: wholesale 48% of retail"); r += 1  # B23
put_assumption(r, "LED session price (EUR)", 45, "Red light therapy"); r += 1      # B24
put_assumption(r, "LED practitioner cost (EUR)", 25, "Per session"); r += 1        # B25
put_assumption(r, "Smoothie avg price (EUR)", 9.67, "R2 Cost Analysis PDF"); r += 1  # B26
put_assumption(r, "Smoothie COGS rate", 0.35, "35% ingredients+packaging (R2)"); r += 1  # B27
ws_a.cell(row=r-1, column=2).number_format = '0%'
put_assumption(r, "Event ticket price (EUR)", 35, "Community events"); r += 1      # B28
put_assumption(r, "Event fixed cost (EUR)", 250, "Venue per event"); r += 1        # B29
put_assumption(r, "Ticketing platform fee", 0.07, "7% Luma/Eventbrite"); r += 1   # B30
ws_a.cell(row=r-1, column=2).number_format = '0%'
put_assumption(r, "Doctor SaaS fee (EUR/mo)", 99, "D1: was EUR 150"); r += 1       # B31
put_assumption(r, "Doctor verify cost (EUR)", 25, "Per new doctor"); r += 1        # B32
put_assumption(r, "Doctor max clinics", 20, "D1: cap for M1-24"); r += 1           # B33

# ─── Cost Rates (rows 36-42) ───
r = 35
section_header(r, "COST RATES"); r += 1
put_assumption(r, "Waste — blended rate", 0.055, "D7: capsules 4%, powders 7%"); r += 1  # B36
ws_a.cell(row=r-1, column=2).number_format = '0.0%'
put_assumption(r, "3PL fulfillment (EUR/unit)", 2.50, "Potions only"); r += 1      # B37
put_assumption(r, "Stripe fee rate", 0.029, "2.9%"); r += 1                        # B38
ws_a.cell(row=r-1, column=2).number_format = '0.0%'
put_assumption(r, "AI API per user (EUR/mo)", 0.35, "D6: hybrid Claude+Gemini R9"); r += 1  # B39
put_assumption(r, "Refund/returns reserve", 0.015, "1.5% of product revenue"); r += 1  # B40
ws_a.cell(row=r-1, column=2).number_format = '0.0%'

# ─── Payroll (rows 44-52) ───
r = 42
section_header(r, "PAYROLL"); r += 1
put_assumption(r, "Founder gross salary (each)", 4166, "EUR/mo"); r += 1           # B43
put_assumption(r, "Employer multiplier", 1.22, "German social contributions"); r += 1  # B44
put_assumption(r, "Space staff gross (M4+)", 1230, "Part-time"); r += 1            # B45
put_assumption(r, "Partner Mgr gross (M6+)", 2000, "EUR/mo"); r += 1              # B46
put_assumption(r, "Growth Marketer gross (M12+)", 4500, "EUR/mo"); r += 1         # B47
put_assumption(r, "Tech/Ops Lead gross (M18+)", 6500, "EUR/mo"); r += 1           # B48
put_assumption(r, "BG insurance rate", 0.015, "~1.5% of gross payroll (mandatory)"); r += 1  # B49
ws_a.cell(row=r-1, column=2).number_format = '0.0%'

# ─── OpEx Fixed Costs (rows 52-68) ───
r = 51
section_header(r, "FIXED OPERATING COSTS"); r += 1
put_assumption(r, "App maintenance retainer", 1500, "External agency"); r += 1     # B52
put_assumption(r, "UX/UI design retainer", 800, "R6 flags: EUR 99K/24mo needed"); r += 1  # B53
put_assumption(r, "Claude Max (internal AI)", 210, "D6: team budget"); r += 1      # B54
put_assumption(r, "SaaS Stack M1-6", 338, "D9: Google, Figma, Notion, Crisp"); r += 1  # B55
put_assumption(r, "SaaS Stack M7-12", 500, "D9: + tools"); r += 1                # B56
put_assumption(r, "SaaS Stack M13+", 750, "D9: full stack"); r += 1              # B57
put_assumption(r, "Rent M1-3 (co-working)", 750, "Before space opens"); r += 1    # B58
put_assumption(r, "Rent M4+ (physical space)", 2500, "Wellness space + smoothie bar"); r += 1  # B59
put_assumption(r, "Steuerberater", 400, "Tax advisor"); r += 1                    # B60
put_assumption(r, "Business insurance", 200, "General liability"); r += 1          # B61
put_assumption(r, "Utilities M4+ (Strom/Gas/Wasser)", 400, "Physical space"); r += 1  # B62
put_assumption(r, "Cleaning/maintenance M4+", 300, "Physical space"); r += 1       # B63
put_assumption(r, "POS system M4+", 50, "Cash register + software"); r += 1       # B64
put_assumption(r, "IHK Berlin (annual ÷ 12)", 25, "Mandatory chamber membership"); r += 1  # B65
put_assumption(r, "Content creation tools", 150, "Canva, stock, etc."); r += 1    # B66
put_assumption(r, "Influencer seeding M7+", 200, "Product seeding"); r += 1       # B67
put_assumption(r, "Investor travel", 150, "Monthly"); r += 1                       # B68
put_assumption(r, "Conferences M7+", 100, "Industry events"); r += 1              # B69
put_assumption(r, "Local transport", 50, "BVG"); r += 1                            # B70
put_assumption(r, "Contingency buffer rate", 0.10, "10% of OpEx (pre-contingency)"); r += 1  # B71
ws_a.cell(row=r-1, column=2).number_format = '0%'
put_assumption(r, "Recruitment M17 (CTO search)", 5000, "One-off"); r += 1        # B72

# ─── CapEx (rows 74-82) ───
r = 74
section_header(r, "CAPITAL EXPENDITURES (D5)"); r += 1
put_assumption(r, "GmbH Formation (M1)", 2500, "One-off"); r += 1                 # B75
put_assumption(r, "Legal Opinion HWG (M1)", 3500, "One-off"); r += 1              # B76
put_assumption(r, "R&D Stability Testing (M1)", 5000, "One-off"); r += 1          # B77
put_assumption(r, "Space Buildout (M3)", 15000, "One-off"); r += 1                # B78
put_assumption(r, "Kitchen Equipment (M3)", 20000, "Smoothie bar"); r += 1        # B79
put_assumption(r, "LED Therapy Devices (M3)", 5000, "One-off"); r += 1            # B80
put_assumption(r, "CapEx depreciation (months)", 36, "Linear 3-year"); r += 1     # B81


# ═══════════════════════════════════════════
# SHEET 2: GROWTH CURVES
# ═══════════════════════════════════════════
ws_g = wb.create_sheet("Growth Curves")
ws_g.sheet_properties.tabColor = "8B9E7C"
ws_g.column_dimensions["A"].width = 26
for col in range(2, 27):
    ws_g.column_dimensions[get_column_letter(col)].width = 9

ws_g.cell(row=1, column=1, value="alche — Monthly Growth Curves (Input Data)").font = f_title
ws_g.cell(row=2, column=1, value="Edit these numbers to change unit forecasts. P&L formulas reference these cells.").font = f_subtitle

# Header row (row 3)
ws_g.cell(row=3, column=1, value="Metric").font = f_header
ws_g.cell(row=3, column=1).fill = fill_header
for i in range(24):
    c = ws_g.cell(row=3, column=i+2, value=f"M{i+1}")
    c.font = f_header
    c.fill = fill_header
    c.alignment = align_c

# Growth curve data
curves = {
    4: ("Potion Units", [0,0,0,50,100,150,250,350,450,600,750,900,1100,1300,1500,1800,2100,2400,2700,3000,3400,3800,4200,4600]),
    5: ("LED Sessions", [0,0,0,10,20,35,50,65,80,100,120,140,160,180,200,220,240,260,280,300,300,300,300,300]),
    6: ("Smoothie Units", [0,0,0,100,150,250,350,450,550,650,750,850,950,1050,1150,1250,1350,1450,1550,1650,1750,1850,1950,2000]),
    7: ("Event Count", [0,0,0,1,1,2,2,3,3,3,3,3,3,3,4,4,4,4,4,5,5,5,5,5]),
    8: ("Event Avg Attendees", [0,0,0,12,15,15,18,18,18,20,20,20,22,22,22,22,24,24,25,25,25,25,25,25]),
    9: ("Doctor Clinics", [0,0,0,0,0,2,5,8,10,12,15,17,18,19,20,20,20,20,20,20,20,20,20,20]),
    10: ("Marketing Spend (EUR)", [50,50,100,200,400,600,800,800,1000,1200,1400,1400,1800,1800,2000,2000,2200,2200,2500,2500,2800,2800,3000,3000]),
}

for row, (label, values) in curves.items():
    ws_g.cell(row=row, column=1, value=label).font = f_sub
    for col, val in enumerate(values, start=2):
        c = ws_g.cell(row=row, column=col, value=val)
        c.font = f_sub
        c.alignment = align_r
        c.number_format = NUM_FMT

# New doctors per month (row 11) — computed from row 9
ws_g.cell(row=11, column=1, value="New Doctors (delta)").font = f_dim
ws_g.cell(row=11, column=2, value=0).font = f_dim  # M1
for col in range(3, 26):
    c = ws_g.cell(row=11, column=col)
    c.font = f_dim
    prev = get_column_letter(col-1)
    cur = get_column_letter(col)
    c.value = f"=MAX(0,{cur}9-{prev}9)"
    c.number_format = NUM_FMT

# Subscriber model (rows 13-18) — formulas with churn
ws_g.cell(row=13, column=1, value="SUBSCRIBER MODEL").font = f_section
ws_g.cell(row=13, column=1).fill = fill_section
for col in range(2, 26):
    ws_g.cell(row=13, column=col).fill = fill_section

# Row 14: New subs rate (from assumptions based on month)
ws_g.cell(row=14, column=1, value="New Subs Rate").font = f_dim
for i in range(24):
    col = i + 2
    m = i + 1  # month number 1-24
    # Use IF to pick the right growth rate based on month
    formula = (f'=IF({m}<=3,Assumptions!B14,'
               f'IF({m}<=6,Assumptions!B15,'
               f'IF({m}<=9,Assumptions!B16,'
               f'IF({m}<=12,Assumptions!B17,'
               f'IF({m}<=18,Assumptions!B18,'
               f'Assumptions!B19)))))')
    ws_g.cell(row=14, column=col, value=formula).font = f_dim
    ws_g.cell(row=14, column=col).number_format = NUM_FMT

# Row 15: Churned subs
ws_g.cell(row=15, column=1, value="Churned Subs").font = f_dim
ws_g.cell(row=15, column=2, value=0)  # M1: no previous subs to churn
for col in range(3, 26):
    prev_col = get_column_letter(col - 1)
    ws_g.cell(row=15, column=col, value=f"=ROUND({prev_col}16*Assumptions!B8,0)").font = f_dim
    ws_g.cell(row=15, column=col).number_format = NUM_FMT

# Row 16: Total Subscribers
ws_g.cell(row=16, column=1, value="Total Subscribers").font = f_total
# M1: just new subs
ws_g.cell(row=16, column=2, value=f"=MAX(0,B14-B15)").font = f_total
ws_g.cell(row=16, column=2).number_format = NUM_FMT
for col in range(3, 26):
    prev_col = get_column_letter(col - 1)
    cur_col = get_column_letter(col)
    ws_g.cell(row=16, column=col, value=f"=MAX(0,{prev_col}16-{cur_col}15+{cur_col}14)").font = f_total
    ws_g.cell(row=16, column=col).number_format = NUM_FMT

# Row 17: Core subs, Row 18: Pro subs, Row 19: Premium subs
ws_g.cell(row=17, column=1, value="  Core Subs (52%)").font = f_dim
ws_g.cell(row=18, column=1, value="  Pro Subs (38%)").font = f_dim
ws_g.cell(row=19, column=1, value="  Premium Subs (10%)").font = f_dim
for col in range(2, 26):
    cur = get_column_letter(col)
    ws_g.cell(row=17, column=col, value=f"=ROUND({cur}16*Assumptions!B9,0)").font = f_dim
    ws_g.cell(row=17, column=col).number_format = NUM_FMT
    ws_g.cell(row=18, column=col, value=f"=ROUND({cur}16*Assumptions!B10,0)").font = f_dim
    ws_g.cell(row=18, column=col).number_format = NUM_FMT
    ws_g.cell(row=19, column=col, value=f"=MAX(0,{cur}16-{cur}17-{cur}18)").font = f_dim
    ws_g.cell(row=19, column=col).number_format = NUM_FMT


# ═══════════════════════════════════════════
# SHEET 3: P&L (FORMULA-BASED)
# ═══════════════════════════════════════════
ws = wb.create_sheet("P&L")
ws.sheet_properties.tabColor = AMBER
ws.column_dimensions["A"].width = 32
for col in range(2, 27):
    ws.column_dimensions[get_column_letter(col)].width = 11
ws.freeze_panes = "B5"

# ─── Title ───
ws.cell(row=1, column=1, value="alche — 24-Month P&L").font = f_title
ws.cell(row=2, column=1, value="Pre-Seed | EUR 500K @ EUR 2.5M Cap | All decisions applied | Formula-based").font = f_subtitle
ws.cell(row=3, column=1, value="All amounts in EUR. Click any cell to see the formula.").font = f_note

# ─── Column headers (row 4) ───
ws.cell(row=4, column=1, value="Budget P&L (EUR)")
for i in range(24):
    ws.cell(row=4, column=i+2, value=f"M{i+1}")
ws.cell(row=4, column=26, value="TOTAL")
for c in range(1, 27):
    cell = ws.cell(row=4, column=c)
    cell.font = f_header
    cell.fill = fill_header
    cell.alignment = align_c

# ─── Helper functions ───
GC = "'Growth Curves'"  # sheet reference
AS = "Assumptions"

def style_row(ws, row, style, cols=range(1, 27)):
    """Apply visual style to a row."""
    for c in cols:
        cell = ws.cell(row=row, column=c)
        if style == "header":
            cell.font = f_header
            cell.fill = fill_header
        elif style == "section":
            cell.font = f_section
            cell.fill = fill_section
        elif style == "total":
            cell.font = f_total
            cell.fill = fill_total
            cell.border = thin_bottom
        elif style == "grand":
            cell.font = f_grand
            cell.fill = fill_grand
            cell.border = thick_top_bottom
        elif style == "sub":
            cell.font = f_sub
        elif style == "dim":
            cell.font = f_dim
        elif style == "unit":
            cell.font = f_unit
        elif style == "product_sep":
            cell.border = product_sep

def write_label(row, label, style="sub"):
    ws.cell(row=row, column=1, value=label)
    style_row(ws, row, style)

def write_formula_row(row, label, formulas, style="sub", fmt=NUM_FMT, total_formula=None):
    """Write a row with label + 24 formulas + total column."""
    ws.cell(row=row, column=1, value=label)
    for i, formula in enumerate(formulas):
        c = ws.cell(row=row, column=i+2, value=formula)
        c.alignment = align_r
        c.number_format = fmt
    # Total column (col 26) — sum of B:Y
    if total_formula:
        ws.cell(row=row, column=26, value=total_formula)
    else:
        cols_range = f"B{row}:Y{row}"
        ws.cell(row=row, column=26, value=f"=SUM({cols_range})")
    ws.cell(row=row, column=26).alignment = align_r
    ws.cell(row=row, column=26).number_format = fmt
    style_row(ws, row, style)

def write_values_row(row, label, values, style="sub", fmt=NUM_FMT):
    """Write a row with hardcoded values (for reference data)."""
    ws.cell(row=row, column=1, value=label)
    for i, val in enumerate(values):
        c = ws.cell(row=row, column=i+2, value=val)
        c.alignment = align_r
        c.number_format = fmt
    ws.cell(row=row, column=26, value=f"=SUM(B{row}:Y{row})")
    ws.cell(row=row, column=26).alignment = align_r
    ws.cell(row=row, column=26).number_format = fmt
    style_row(ws, row, style)

def sep_row(row):
    """Add thick amber separator."""
    for c in range(1, 27):
        ws.cell(row=row, column=c).border = product_sep

def month_col(m):
    """Get column letter for month m (0-indexed)."""
    return get_column_letter(m + 2)

def gc_ref(row, col):
    """Reference to Growth Curves sheet."""
    return f"{GC}!{get_column_letter(col)}{row}"

def a_ref(row):
    """Reference to Assumptions sheet cell B{row}."""
    return f"{AS}!B{row}"

# ═══════════════════════════════════════════
# P&L ROWS
# ═══════════════════════════════════════════
ROW = 5

# ─── SUBSCRIBER METRICS ───
write_label(ROW, "SUBSCRIBER METRICS", "header"); ROW += 1
# New subs
formulas = [f"={gc_ref(14, i+2)}" for i in range(24)]
write_formula_row(ROW, "  New Subs Added", formulas, "dim"); ROW += 1
# Churned
formulas = [f"={gc_ref(15, i+2)}" for i in range(24)]
write_formula_row(ROW, "  Churned Subs", formulas, "dim"); ROW += 1
# Total subs
formulas = [f"={gc_ref(16, i+2)}" for i in range(24)]
write_formula_row(ROW, "  Total Subscribers", formulas, "total"); ROW += 1
# Tier breakdown
formulas = [f"={gc_ref(17, i+2)}" for i in range(24)]
write_formula_row(ROW, "    Core (52%)", formulas, "dim"); ROW += 1
formulas = [f"={gc_ref(18, i+2)}" for i in range(24)]
write_formula_row(ROW, "    Pro (38%)", formulas, "dim"); ROW += 1
formulas = [f"={gc_ref(19, i+2)}" for i in range(24)]
write_formula_row(ROW, "    Premium (10%)", formulas, "dim"); ROW += 1

ROW += 1  # blank row

# ═══════════════════════════════════════════
# INCOME
# ═══════════════════════════════════════════
write_label(ROW, "INCOME", "header"); INCOME_START = ROW; ROW += 1

# ─── POTIONS ───
write_label(ROW, "  Potions", "section"); ROW += 1
# Units
pot_units_row = ROW
formulas = [f"={gc_ref(4, i+2)}" for i in range(24)]
write_formula_row(ROW, "    Units shipped", formulas, "unit"); ROW += 1
# Price/unit
write_formula_row(ROW, "    Price/unit", [f"={a_ref(22)}"]*24, "unit", '#,##0.00'); ROW += 1
# Revenue
pot_rev_row = ROW
formulas = [f"={month_col(i)}{pot_units_row}*{a_ref(22)}" for i in range(24)]
write_formula_row(ROW, "    Revenue", formulas, "sub"); ROW += 1
sep_row(ROW - 1)

# ─── LED THERAPY ───
write_label(ROW, "  LED Therapy", "section"); ROW += 1
led_units_row = ROW
formulas = [f"={gc_ref(5, i+2)}" for i in range(24)]
write_formula_row(ROW, "    Sessions", formulas, "unit"); ROW += 1
write_formula_row(ROW, "    Price/session", [f"={a_ref(24)}"]*24, "unit", '#,##0.00'); ROW += 1
led_rev_row = ROW
formulas = [f"={month_col(i)}{led_units_row}*{a_ref(24)}" for i in range(24)]
write_formula_row(ROW, "    Revenue", formulas, "sub"); ROW += 1
sep_row(ROW - 1)

# ─── SMOOTHIE BAR ───
write_label(ROW, "  Smoothie Bar", "section"); ROW += 1
sm_units_row = ROW
formulas = [f"={gc_ref(6, i+2)}" for i in range(24)]
write_formula_row(ROW, "    Units sold", formulas, "unit"); ROW += 1
write_formula_row(ROW, "    Price/smoothie", [f"={a_ref(26)}"]*24, "unit", '#,##0.00'); ROW += 1
sm_rev_row = ROW
formulas = [f"=ROUND({month_col(i)}{sm_units_row}*{a_ref(26)},0)" for i in range(24)]
write_formula_row(ROW, "    Revenue", formulas, "sub"); ROW += 1
sep_row(ROW - 1)

# ─── EVENTS ───
write_label(ROW, "  Community Events", "section"); ROW += 1
ev_count_row = ROW
formulas = [f"={gc_ref(7, i+2)}" for i in range(24)]
write_formula_row(ROW, "    Event count", formulas, "unit"); ROW += 1
ev_attend_row = ROW
formulas = [f"={gc_ref(8, i+2)}" for i in range(24)]
write_formula_row(ROW, "    Avg attendees/event", formulas, "unit"); ROW += 1
ev_total_attend_row = ROW
formulas = [f"={month_col(i)}{ev_count_row}*{month_col(i)}{ev_attend_row}" for i in range(24)]
write_formula_row(ROW, "    Total attendees", formulas, "unit"); ROW += 1
write_formula_row(ROW, "    Price/ticket", [f"={a_ref(28)}"]*24, "unit", '#,##0.00'); ROW += 1
ev_rev_row = ROW
formulas = [f"={month_col(i)}{ev_total_attend_row}*{a_ref(28)}" for i in range(24)]
write_formula_row(ROW, "    Revenue", formulas, "sub"); ROW += 1
sep_row(ROW - 1)

# ─── DOCTOR SAAS ───
write_label(ROW, "  Doctor SaaS (D1: EUR 99/mo)", "section"); ROW += 1
doc_subs_row = ROW
formulas = [f"={gc_ref(9, i+2)}" for i in range(24)]
write_formula_row(ROW, "    Clinics listed", formulas, "unit"); ROW += 1
write_formula_row(ROW, "    Fee/month", [f"={a_ref(31)}"]*24, "unit", '#,##0.00'); ROW += 1
doc_rev_row = ROW
formulas = [f"={month_col(i)}{doc_subs_row}*{a_ref(31)}" for i in range(24)]
write_formula_row(ROW, "    Revenue", formulas, "sub"); ROW += 1
sep_row(ROW - 1)

# ─── RESTAURANT (Phase 2 — EUR 0) ───
rest_rev_row = ROW
write_values_row(ROW, "  Restaurant (Phase 2 — EUR 0)", [0]*24, "dim"); ROW += 1
sep_row(ROW - 1)

# ─── SUBSCRIPTIONS ───
write_label(ROW, "  Subscriptions", "section"); ROW += 1
sub_core_row = ROW
formulas = [f"={gc_ref(17, i+2)}*{a_ref(5)}" for i in range(24)]
write_formula_row(ROW, "    Core (EUR 19/mo)", formulas, "sub"); ROW += 1
sub_pro_row = ROW
formulas = [f"={gc_ref(18, i+2)}*{a_ref(6)}" for i in range(24)]
write_formula_row(ROW, "    Pro (EUR 49/mo)", formulas, "sub"); ROW += 1
sub_prem_row = ROW
formulas = [f"={gc_ref(19, i+2)}*{a_ref(7)}" for i in range(24)]
write_formula_row(ROW, "    Premium (EUR 99/mo)", formulas, "sub"); ROW += 1
sub_total_row = ROW
formulas = [f"={month_col(i)}{sub_core_row}+{month_col(i)}{sub_pro_row}+{month_col(i)}{sub_prem_row}" for i in range(24)]
write_formula_row(ROW, "    Subscription MRR", formulas, "total"); ROW += 1
sep_row(ROW - 1)

# ─── TOTAL INCOME ───
ROW += 1
total_income_row = ROW
# Sum of all revenue rows
rev_rows = [pot_rev_row, led_rev_row, sm_rev_row, ev_rev_row, doc_rev_row, rest_rev_row, sub_total_row]
formulas = ["+".join(f"{month_col(i)}{r}" for r in rev_rows) for i in range(24)]
formulas = [f"={f}" for f in formulas]
write_formula_row(ROW, "TOTAL INCOME", formulas, "grand"); ROW += 1

ROW += 1

# ═══════════════════════════════════════════
# COGS
# ═══════════════════════════════════════════
write_label(ROW, "COST OF GOODS SOLD (COGS)", "header"); ROW += 1

# ─── Potions COGS ───
write_label(ROW, "  Potions COGS", "section"); ROW += 1
cogs_pot_row = ROW
formulas = [f"=ROUND({month_col(i)}{pot_units_row}*{a_ref(23)},0)" for i in range(24)]
write_formula_row(ROW, "    Product cost (EUR 23.50/unit)", formulas, "sub"); ROW += 1
cogs_waste_row = ROW
formulas = [f"=ROUND({month_col(i)}{pot_rev_row}*{a_ref(36)},0)" for i in range(24)]
write_formula_row(ROW, "    Waste/spoilage (5.5%)", formulas, "sub"); ROW += 1
cogs_3pl_row = ROW
formulas = [f"=IF({i+1}>=4,ROUND({month_col(i)}{pot_units_row}*{a_ref(37)},0),0)" for i in range(24)]
write_formula_row(ROW, "    3PL fulfillment (EUR 2.50/unit)", formulas, "sub"); ROW += 1
sep_row(ROW - 1)

# ─── LED COGS ───
cogs_led_row = ROW
formulas = [f"=ROUND({month_col(i)}{led_units_row}*{a_ref(25)},0)" for i in range(24)]
write_formula_row(ROW, "  LED Practitioner (EUR 25/session)", formulas, "sub"); ROW += 1
sep_row(ROW - 1)

# ─── Smoothie COGS ───
write_label(ROW, "  Smoothie Bar COGS", "section"); ROW += 1
cogs_sm_row = ROW
formulas = [f"=ROUND({month_col(i)}{sm_rev_row}*{a_ref(27)},0)" for i in range(24)]
write_formula_row(ROW, "    Ingredients + packaging (35%)", formulas, "sub"); ROW += 1
sep_row(ROW - 1)

# ─── Events COGS ───
cogs_ev_row = ROW
formulas = [f"=ROUND({month_col(i)}{ev_count_row}*{a_ref(29)}+{month_col(i)}{ev_rev_row}*{a_ref(30)},0)" for i in range(24)]
write_formula_row(ROW, "  Events & Ticketing", formulas, "sub"); ROW += 1
sep_row(ROW - 1)

# ─── Doctor COGS ───
cogs_doc_row = ROW
formulas = [f"=ROUND({gc_ref(11, i+2)}*{a_ref(32)},0)" for i in range(24)]
write_formula_row(ROW, "  Doctor Verification (EUR 25/new)", formulas, "sub"); ROW += 1

# ─── AI API COGS ───
cogs_ai_row = ROW
formulas = [f"=ROUND({gc_ref(16, i+2)}*{a_ref(39)},0)" for i in range(24)]
write_formula_row(ROW, "  AI API — User-Facing (EUR 0.35/sub)", formulas, "sub"); ROW += 1

# ─── Refund Reserve (NEW from persona audit) ───
cogs_refund_row = ROW
product_rev_sum = "+".join(f"{month_col(i)}{r}" for r in [pot_rev_row, sm_rev_row, ev_rev_row])
formulas = [f"=ROUND(({month_col(i)}{pot_rev_row}+{month_col(i)}{sm_rev_row}+{month_col(i)}{ev_rev_row})*{a_ref(40)},0)" for i in range(24)]
write_formula_row(ROW, "  Refund/Returns Reserve (1.5%)", formulas, "sub"); ROW += 1

# ─── Stripe ───
cogs_stripe_row = ROW
formulas = [f"=ROUND({month_col(i)}{total_income_row}*{a_ref(38)},0)" for i in range(24)]
write_formula_row(ROW, "  Stripe Processing (2.9%)", formulas, "sub"); ROW += 1

# ─── TOTAL COGS ───
ROW += 1
total_cogs_row = ROW
cogs_rows = [cogs_pot_row, cogs_waste_row, cogs_3pl_row, cogs_led_row, cogs_sm_row,
             cogs_ev_row, cogs_doc_row, cogs_ai_row, cogs_refund_row, cogs_stripe_row]
formulas = [f"={'+'.join(f'{month_col(i)}{r}' for r in cogs_rows)}" for i in range(24)]
write_formula_row(ROW, "TOTAL COGS", formulas, "grand"); ROW += 1

# ─── GROSS PROFIT ───
ROW += 1
gp_row = ROW
formulas = [f"={month_col(i)}{total_income_row}-{month_col(i)}{total_cogs_row}" for i in range(24)]
write_formula_row(ROW, "GROSS PROFIT", formulas, "grand"); ROW += 1
# Gross margin %
gm_row = ROW
formulas = [f"=IF({month_col(i)}{total_income_row}>0,{month_col(i)}{gp_row}/{month_col(i)}{total_income_row},0)" for i in range(24)]
write_formula_row(ROW, "  Gross Margin %", formulas, "dim", PCT_FMT,
                  total_formula=f"=IF(Z{total_income_row}>0,Z{gp_row}/Z{total_income_row},0)"); ROW += 1

ROW += 1

# ═══════════════════════════════════════════
# OPERATING EXPENSES
# ═══════════════════════════════════════════
write_label(ROW, "OPERATING EXPENSES", "header"); ROW += 1

# ─── PAYROLL ───
write_label(ROW, "  Payroll", "section"); ROW += 1
pay_f1_row = ROW
formulas = [f"=ROUND({a_ref(43)}*{a_ref(44)},0)"]*24
write_formula_row(ROW, "    Timu (CEO)", formulas, "sub"); ROW += 1
pay_f2_row = ROW
write_formula_row(ROW, "    Daria (COO)", formulas, "sub"); ROW += 1
pay_space_row = ROW
formulas = [f"=IF({i+1}>=4,ROUND({a_ref(45)}*{a_ref(44)},0),0)" for i in range(24)]
write_formula_row(ROW, "    Space Staff (M4+)", formulas, "sub"); ROW += 1
pay_mgr_row = ROW
formulas = [f"=IF({i+1}>=6,ROUND({a_ref(46)}*{a_ref(44)},0),0)" for i in range(24)]
write_formula_row(ROW, "    Partner Mgr (M6+)", formulas, "sub"); ROW += 1
pay_growth_row = ROW
formulas = [f"=IF({i+1}>=12,ROUND({a_ref(47)}*{a_ref(44)},0),0)" for i in range(24)]
write_formula_row(ROW, "    Growth Marketer (M12+)", formulas, "sub"); ROW += 1
pay_cto_row = ROW
formulas = [f"=IF({i+1}>=18,ROUND({a_ref(48)}*{a_ref(44)},0),0)" for i in range(24)]
write_formula_row(ROW, "    Tech/Ops Lead (M18+)", formulas, "sub"); ROW += 1
# BG Insurance (NEW — mandatory)
pay_bg_row = ROW
pay_rows_for_bg = [pay_f1_row, pay_f2_row, pay_space_row, pay_mgr_row, pay_growth_row, pay_cto_row]
formulas = [f"=ROUND(({'+'.join(f'{month_col(i)}{r}' for r in pay_rows_for_bg)})*{a_ref(49)},0)" for i in range(24)]
write_formula_row(ROW, "    BG Insurance (1.5% payroll)", formulas, "sub"); ROW += 1
# Payroll total
pay_total_row = ROW
all_pay = pay_rows_for_bg + [pay_bg_row]
formulas = [f"={'+'.join(f'{month_col(i)}{r}' for r in all_pay)}" for i in range(24)]
write_formula_row(ROW, "    Payroll Total", formulas, "total"); ROW += 1
sep_row(ROW - 1)

# ─── R&D ───
write_label(ROW, "  R&D", "section"); ROW += 1
rd_app_row = ROW
formulas = [f"={a_ref(52)}"]*24
write_formula_row(ROW, "    App Maintenance", formulas, "sub"); ROW += 1
rd_ux_row = ROW
formulas = [f"={a_ref(53)}"]*24
write_formula_row(ROW, "    UX/UI Design", formulas, "sub"); ROW += 1
rd_total_row = ROW
formulas = [f"={month_col(i)}{rd_app_row}+{month_col(i)}{rd_ux_row}" for i in range(24)]
write_formula_row(ROW, "    R&D Total", formulas, "total"); ROW += 1
sep_row(ROW - 1)

# ─── S&M ───
write_label(ROW, "  Sales & Marketing", "section"); ROW += 1
sm_ads_row = ROW
formulas = [f"={gc_ref(10, i+2)}" for i in range(24)]
write_formula_row(ROW, "    Paid Ads", formulas, "sub"); ROW += 1
sm_content_row = ROW
formulas = [f"={a_ref(66)}"]*24
write_formula_row(ROW, "    Content Creation", formulas, "sub"); ROW += 1
sm_inf_row = ROW
formulas = [f"=IF({i+1}>=7,{a_ref(67)},0)" for i in range(24)]
write_formula_row(ROW, "    Influencer Seeding (M7+)", formulas, "sub"); ROW += 1
sm_total_row = ROW
formulas = [f"={month_col(i)}{sm_ads_row}+{month_col(i)}{sm_content_row}+{month_col(i)}{sm_inf_row}" for i in range(24)]
write_formula_row(ROW, "    S&M Total", formulas, "total"); ROW += 1
sep_row(ROW - 1)

# ─── G&A ───
write_label(ROW, "  General & Admin", "section"); ROW += 1
ga_rent_row = ROW
formulas = [f"=IF({i+1}<=3,{a_ref(58)},{a_ref(59)})" for i in range(24)]
write_formula_row(ROW, "    Rent (co-work M1-3 / space M4+)", formulas, "sub"); ROW += 1
ga_steuer_row = ROW
formulas = [f"={a_ref(60)}"]*24
write_formula_row(ROW, "    Steuerberater", formulas, "sub"); ROW += 1
ga_ins_row = ROW
formulas = [f"={a_ref(61)}"]*24
write_formula_row(ROW, "    Business Insurance", formulas, "sub"); ROW += 1
# NEW: Utilities
ga_util_row = ROW
formulas = [f"=IF({i+1}>=4,{a_ref(62)},0)" for i in range(24)]
write_formula_row(ROW, "    Utilities (M4+)", formulas, "sub"); ROW += 1
# NEW: Cleaning
ga_clean_row = ROW
formulas = [f"=IF({i+1}>=4,{a_ref(63)},0)" for i in range(24)]
write_formula_row(ROW, "    Cleaning/Maintenance (M4+)", formulas, "sub"); ROW += 1
# NEW: POS
ga_pos_row = ROW
formulas = [f"=IF({i+1}>=4,{a_ref(64)},0)" for i in range(24)]
write_formula_row(ROW, "    POS System (M4+)", formulas, "sub"); ROW += 1
# NEW: IHK
ga_ihk_row = ROW
formulas = [f"={a_ref(65)}"]*24
write_formula_row(ROW, "    IHK Berlin", formulas, "sub"); ROW += 1
# Permits
ga_permit_row = ROW
formulas = [f"=IF({i+1}=3,500,0)" for i in range(24)]
write_formula_row(ROW, "    Permits & Licenses (M3)", formulas, "sub"); ROW += 1
ga_total_row = ROW
ga_rows = [ga_rent_row, ga_steuer_row, ga_ins_row, ga_util_row, ga_clean_row, ga_pos_row, ga_ihk_row, ga_permit_row]
formulas = [f"={'+'.join(f'{month_col(i)}{r}' for r in ga_rows)}" for i in range(24)]
write_formula_row(ROW, "    G&A Total", formulas, "total"); ROW += 1
sep_row(ROW - 1)

# ─── Management & APIs ───
write_label(ROW, "  Management & APIs", "section"); ROW += 1
mg_saas_row = ROW
formulas = [f"=IF({i+1}<=6,{a_ref(55)},IF({i+1}<=12,{a_ref(56)},{a_ref(57)}))" for i in range(24)]
write_formula_row(ROW, "    SaaS Stack (D9)", formulas, "sub"); ROW += 1
mg_claude_row = ROW
formulas = [f"={a_ref(54)}"]*24
write_formula_row(ROW, "    Claude Max (internal)", formulas, "sub"); ROW += 1
mg_total_row = ROW
formulas = [f"={month_col(i)}{mg_saas_row}+{month_col(i)}{mg_claude_row}" for i in range(24)]
write_formula_row(ROW, "    Management Total", formulas, "total"); ROW += 1
sep_row(ROW - 1)

# ─── Travel ───
write_label(ROW, "  Travel", "section"); ROW += 1
tr_inv_row = ROW
formulas = [f"={a_ref(68)}"]*24
write_formula_row(ROW, "    Investor Travel", formulas, "sub"); ROW += 1
tr_conf_row = ROW
formulas = [f"=IF({i+1}>=7,{a_ref(69)},0)" for i in range(24)]
write_formula_row(ROW, "    Conferences (M7+)", formulas, "sub"); ROW += 1
tr_local_row = ROW
formulas = [f"={a_ref(70)}"]*24
write_formula_row(ROW, "    Local Transport", formulas, "sub"); ROW += 1
tr_total_row = ROW
formulas = [f"={month_col(i)}{tr_inv_row}+{month_col(i)}{tr_conf_row}+{month_col(i)}{tr_local_row}" for i in range(24)]
write_formula_row(ROW, "    Travel Total", formulas, "total"); ROW += 1
sep_row(ROW - 1)

# ─── Recruitment (NEW) ───
recruit_row = ROW
formulas = [f"=IF({i+1}=17,{a_ref(72)},0)" for i in range(24)]
write_formula_row(ROW, "  Recruitment (CTO search M17)", formulas, "sub"); ROW += 1

# ─── TOTAL OPEX (pre-contingency) ───
ROW += 1
opex_pre_row = ROW
opex_section_rows = [pay_total_row, rd_total_row, sm_total_row, ga_total_row, mg_total_row, tr_total_row, recruit_row]
formulas = [f"={'+'.join(f'{month_col(i)}{r}' for r in opex_section_rows)}" for i in range(24)]
write_formula_row(ROW, "  OpEx (pre-contingency)", formulas, "total"); ROW += 1

# ─── Contingency Buffer (NEW — 10%) ───
contingency_row = ROW
formulas = [f"=ROUND({month_col(i)}{opex_pre_row}*{a_ref(71)},0)" for i in range(24)]
write_formula_row(ROW, "  Contingency Buffer (10%)", formulas, "sub"); ROW += 1

# ─── TOTAL OPEX (with contingency) ───
ROW += 1
total_opex_row = ROW
formulas = [f"={month_col(i)}{opex_pre_row}+{month_col(i)}{contingency_row}" for i in range(24)]
write_formula_row(ROW, "TOTAL OPEX", formulas, "grand"); ROW += 1

# ═══════════════════════════════════════════
# EBITDA
# ═══════════════════════════════════════════
ROW += 1
ebitda_row = ROW
formulas = [f"={month_col(i)}{gp_row}-{month_col(i)}{total_opex_row}" for i in range(24)]
write_formula_row(ROW, "EBITDA", formulas, "grand"); ROW += 1

ROW += 1

# ═══════════════════════════════════════════
# CapEx
# ═══════════════════════════════════════════
write_label(ROW, "CAPITAL EXPENDITURES (CapEx)", "header"); ROW += 1
capex_gmbh_row = ROW
formulas = [f"=IF({i+1}=1,{a_ref(75)},0)" for i in range(24)]
write_formula_row(ROW, "  GmbH Formation (M1)", formulas, "sub"); ROW += 1
capex_legal_row = ROW
formulas = [f"=IF({i+1}=1,{a_ref(76)},0)" for i in range(24)]
write_formula_row(ROW, "  Legal Opinion HWG (M1)", formulas, "sub"); ROW += 1
capex_rd_row = ROW
formulas = [f"=IF({i+1}=1,{a_ref(77)},0)" for i in range(24)]
write_formula_row(ROW, "  R&D Stability Testing (M1)", formulas, "sub"); ROW += 1
capex_build_row = ROW
formulas = [f"=IF({i+1}=3,{a_ref(78)},0)" for i in range(24)]
write_formula_row(ROW, "  Space Buildout (M3)", formulas, "sub"); ROW += 1
capex_kitchen_row = ROW
formulas = [f"=IF({i+1}=3,{a_ref(79)},0)" for i in range(24)]
write_formula_row(ROW, "  Kitchen Equipment (M3)", formulas, "sub"); ROW += 1
capex_led_row = ROW
formulas = [f"=IF({i+1}=3,{a_ref(80)},0)" for i in range(24)]
write_formula_row(ROW, "  LED Devices (M3)", formulas, "sub"); ROW += 1

# ─── Depreciation (NEW — linear over 36 months) ───
dep_row = ROW
total_capex_refs = "+".join(f"{a_ref(r)}" for r in [75, 76, 77, 78, 79, 80])
# Depreciation starts M1 for M1 items, M3 for M3 items
formulas = []
for i in range(24):
    m = i + 1
    # M1 items depreciate from M1: (GmbH + Legal + R&D) / 36
    # M3 items depreciate from M3: (Buildout + Kitchen + LED) / 36
    if m >= 3:
        formulas.append(f"=ROUND(({a_ref(75)}+{a_ref(76)}+{a_ref(77)}+{a_ref(78)}+{a_ref(79)}+{a_ref(80)})/{a_ref(81)},0)")
    elif m >= 1:
        formulas.append(f"=ROUND(({a_ref(75)}+{a_ref(76)}+{a_ref(77)})/{a_ref(81)},0)")
    else:
        formulas.append(0)
write_formula_row(ROW, "  Depreciation (linear / 36mo)", formulas, "dim"); ROW += 1

# Total CapEx (cash outflow only, not depreciation)
ROW += 1
total_capex_row = ROW
capex_cash_rows = [capex_gmbh_row, capex_legal_row, capex_rd_row, capex_build_row, capex_kitchen_row, capex_led_row]
formulas = [f"={'+'.join(f'{month_col(i)}{r}' for r in capex_cash_rows)}" for i in range(24)]
write_formula_row(ROW, "TOTAL CAPEX (cash)", formulas, "grand"); ROW += 1

# ═══════════════════════════════════════════
# NET P&L
# ═══════════════════════════════════════════
ROW += 1
net_pnl_row = ROW
formulas = [f"={month_col(i)}{ebitda_row}-{month_col(i)}{total_capex_row}" for i in range(24)]
write_formula_row(ROW, "NET P&L (EBITDA - CapEx)", formulas, "grand"); ROW += 1

# ─── Cumulative Cash ───
cum_cash_row = ROW
# M1: 500000 + net P&L
ws.cell(row=ROW, column=1, value="Cumulative Cash (EUR 500K start)")
ws.cell(row=ROW, column=2, value=f"=500000+B{net_pnl_row}")
ws.cell(row=ROW, column=2).number_format = NUM_FMT
ws.cell(row=ROW, column=2).alignment = align_r
for i in range(1, 24):
    col = i + 2
    prev_col = get_column_letter(col - 1)
    cur_col = get_column_letter(col)
    ws.cell(row=ROW, column=col, value=f"={prev_col}{cum_cash_row}+{cur_col}{net_pnl_row}")
    ws.cell(row=ROW, column=col).number_format = NUM_FMT
    ws.cell(row=ROW, column=col).alignment = align_r
ws.cell(row=ROW, column=26, value=f"=Y{cum_cash_row}")  # final balance
ws.cell(row=ROW, column=26).number_format = NUM_FMT
style_row(ws, ROW, "total"); ROW += 1

ROW += 1

# ═══════════════════════════════════════════
# KEY METRICS
# ═══════════════════════════════════════════
write_label(ROW, "KEY METRICS", "header"); ROW += 1
# Monthly burn
formulas = [f"=IF({month_col(i)}{net_pnl_row}<0,ABS({month_col(i)}{net_pnl_row}),0)" for i in range(24)]
write_formula_row(ROW, "  Monthly Burn (if loss)", formulas, "dim"); ROW += 1
# Revenue per sub
formulas = [f"=IF({gc_ref(16, i+2)}>0,ROUND({month_col(i)}{total_income_row}/{gc_ref(16, i+2)},2),0)" for i in range(24)]
write_formula_row(ROW, "  Revenue / Subscriber", formulas, "dim", '#,##0.00'); ROW += 1
# Subscription MRR
formulas = [f"={month_col(i)}{sub_total_row}" for i in range(24)]
write_formula_row(ROW, "  Subscription MRR", formulas, "dim"); ROW += 1
# Gross margin %
formulas = [f"={month_col(i)}{gm_row}" for i in range(24)]
write_formula_row(ROW, "  Gross Margin %", formulas, "dim", PCT_FMT,
                  total_formula=f"=Z{gm_row}"); ROW += 1


# ═══════════════════════════════════════════
# SHEET 4: DECISIONS
# ═══════════════════════════════════════════
ws3 = wb.create_sheet("Decisions")
ws3.sheet_properties.tabColor = TERRA
ws3.column_dimensions["A"].width = 8
ws3.column_dimensions["B"].width = 28
ws3.column_dimensions["C"].width = 50
ws3.column_dimensions["D"].width = 25
ws3.column_dimensions["E"].width = 40

ws3.cell(row=1, column=1, value="alche — Decision Log (D1-D11)").font = f_title
ws3.cell(row=2, column=1, value="All decisions from P&L session, 2026-02-23").font = f_subtitle

for col, h in enumerate(["#", "Decision", "Choice Made", "Old Value", "P&L Impact"], 1):
    c = ws3.cell(row=4, column=col, value=h)
    c.font = f_header
    c.fill = fill_header

decisions = [
    ["D1", "Doctor Revenue", "C. SaaS EUR 99/mo × 20 clinics", "EUR 150/mo × 865", "Revenue -75%; legally safe (StGB §299a)"],
    ["D2", "Restaurant Revenue", "B. Phase 2 (500+ users)", "EUR 1.50/cover M6+", "EUR 0 all 24 months"],
    ["D3", "CGM Hardware", "A. Software-Only BYOD", "EUR 130 × units", "Removed; margin 80-90%"],
    ["D4", "Physical Space", "A. Smoothie bar + space", "EUR 0", "New rev + EUR 2,500/mo rent + CapEx"],
    ["D5", "CapEx Section", "Yes, include", "All OpEx", "EUR 51,000 total"],
    ["D6", "AI Cost Model", "Both internal + user-facing", "EUR 210 + EUR 499 Terra", "EUR 210 + EUR 0.35/sub (R9)"],
    ["D7", "Product Waste", "Category-weighted", "Flat 5%", "5.5% blended"],
    ["D8", "Terra CGM API", "Defer — free HealthKit", "EUR 499/mo M6+", "Saves EUR 9,481"],
    ["D9", "SaaS Stack", "Itemized (R8)", "EUR 350→1800", "EUR 338→500→750"],
    ["D10", "Menu Analysis", "EUR 0 — founders do it", "EUR 500/mo", "Saves EUR 12,000"],
    ["D11", "Retail Products", "EUR 49 validated; COGS EUR 23.50", "EUR 11.50 COGS", "Margin 76% → 52% (R10)"],
]

for i, dec in enumerate(decisions):
    row = 5 + i
    for col, val in enumerate(dec, 1):
        c = ws3.cell(row=row, column=col, value=val)
        c.font = f_sub
        if col == 1:
            c.font = Font(name="Calibri", size=10, bold=True, color=TERRA)
        c.border = thin_bottom

# ─── NEW LINES FROM PERSONA AUDIT ───
row = 18
ws3.cell(row=row, column=1, value="NEW LINES ADDED (7-Persona Audit)").font = f_section
ws3.cell(row=row, column=1).fill = fill_section
for c in range(2, 6):
    ws3.cell(row=row, column=c).fill = fill_section
row += 1

new_lines = [
    ["P1", "Utilities (Strom/Gas/Wasser)", "EUR 400/mo from M4", "Operations Manager", "Can't run smoothie bar without electricity"],
    ["P2", "Cleaning/Maintenance", "EUR 300/mo from M4", "Operations Manager", "Physical space upkeep"],
    ["P3", "BG Insurance", "~1.5% of gross payroll", "Steuerberater", "MANDATORY German employer insurance"],
    ["P4", "Contingency Buffer", "10% of OpEx", "Angel Investor", "Every serious investor expects this"],
    ["P5", "Refund/Returns Reserve", "1.5% of product revenue", "CFO/FP&A", "Physical product returns"],
    ["P6", "IHK Berlin", "EUR 25/mo (EUR 300/yr)", "Steuerberater", "Mandatory chamber membership"],
    ["P7", "Recruitment Cost", "EUR 5,000 one-off M17", "HR Manager", "CTO job posting / headhunter"],
    ["P8", "POS System", "EUR 50/mo from M4", "Operations Manager", "Cash register for space"],
    ["P9", "Depreciation", "CapEx / 36 months linear", "CFO/FP&A", "Tax-accurate P&L presentation"],
    ["P10", "Permits & Licenses", "EUR 500 one-off M3", "Food Inspector", "Lebensmittelüberwachung"],
    ["P11", "Space Staff", "EUR 1,501/mo from M4", "Operations Manager", "Part-time for smoothie bar"],
]

for col, h in enumerate(["#", "Line Item", "Amount", "Source Persona", "Rationale"], 1):
    c = ws3.cell(row=row, column=col, value=h)
    c.font = f_header
    c.fill = fill_header
row += 1

for item in new_lines:
    for col, val in enumerate(item, 1):
        c = ws3.cell(row=row, column=col, value=val)
        c.font = f_sub
        if col == 1:
            c.font = Font(name="Calibri", size=10, bold=True, color=SAGE)
        c.border = thin_bottom
    row += 1


# ═══════════════════════════════════════════
# MOVE P&L to first position
# ═══════════════════════════════════════════
wb.move_sheet("P&L", offset=-2)


# ─── SAVE ───
output_path = "/Users/timoel/Downloads/pitchbook/pnl/alche-pnl-24mo-v2.xlsx"
wb.save(output_path)
print(f"Excel P&L v2 saved to: {output_path}")
print(f"Sheets: {wb.sheetnames}")
print(f"P&L rows: {ROW}")
print(f"\nOpen in Excel and click any cell to see the formula.")
print(f"Change any value on the Assumptions sheet → P&L auto-recalculates.")
