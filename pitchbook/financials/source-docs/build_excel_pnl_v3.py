#!/usr/bin/env python3
"""
alche P&L Excel Generator v3 — Full Persona-Audited, Multi-Sheet Model
═══════════════════════════════════════════════════════════════════════
v3.3 changes:
  - Ablöse (key money): EUR 50K M3 CapEx, amortized 36mo (hava parası)
  - Kaution (security deposit): EUR 7.5K M3 cash outflow (Cash Flow only, refundable)
  - Total CapEx: EUR 98K → EUR 148K

Changes from v2:
  - All 7 persona audit findings incorporated (P1-P7)
  - Employer multiplier: 1.22x → 1.25x (P2/P7)
  - Insurance split: EUR 200 → EUR 650/mo (P6: D&O, cyber, product, recall)
  - DSB (Data Protection Officer): EUR 400/mo mandatory (P6)
  - Nebenkosten: EUR 325/mo from M4 (P3)
  - Utilities: EUR 400 → EUR 500/mo (P3)
  - Buildout: EUR 15,000 → EUR 25,000 (P3)
  - Recruitment CTO: EUR 5,000 → EUR 18,000 (P7)
  - NEW regulatory lines: GEMA, Rundfunkbeitrag, Verpackungsgesetz, KSK, waste disposal, WiFi, etc.
  - NEW CapEx: DPIA, MDR opinion, fire safety, signage, HACCP, hygiene, food safety equip
  - CGM commission model (D3 updated) — EUR 15/user/mo from M7; Premium 3-mo gate or add-on; zero COGS
  - Tax provision: ~30% when cumulative P&L positive (P2)
  - Cash Flow Statement sheet (P5: EUR 218K gap)
  - Unit Economics Dashboard sheet
  - Scenario Summary sheet (Base/Bull/Bear)
  - Decisions sheet expanded with persona findings
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ─── STYLE DEFINITIONS ───
DEEP = "2C2418"
AMBER = "C4956A"
SAGE = "8B9E7C"
TERRA = "B86B4A"
ROSE = "C47A8A"
WHITE = "FFFFFF"
HEADER_BG = "3D3225"
SECTION_BG = "EDE7DC"
TOTAL_BG = "E5DDD3"
GRAND_BG = "D8CFC3"
PRODUCT_SEP = "C4956A"
LIGHT_BG = "FEFCF9"
CGM_GREY = "B0A89E"  # for greyed-out CGM row

f_title = Font(name="Calibri", size=16, bold=True, color=DEEP)
f_subtitle = Font(name="Calibri", size=9, italic=True, color="9E948A")
f_header = Font(name="Calibri", size=10, bold=True, color=WHITE)
f_section = Font(name="Calibri", size=10, bold=True, color=DEEP)
f_sub = Font(name="Calibri", size=10, color=DEEP)
f_dim = Font(name="Calibri", size=10, color="8A7F73")
f_unit = Font(name="Calibri", size=9, color="8A7F73")
f_total = Font(name="Calibri", size=10, bold=True, color=DEEP)
f_grand = Font(name="Calibri", size=11, bold=True, color=DEEP)
f_note = Font(name="Calibri", size=9, italic=True, color="9E948A")
f_label = Font(name="Calibri", size=10, color="5C5244")
f_val = Font(name="Calibri", size=10, bold=True, color=DEEP)
f_decision = Font(name="Calibri", size=10, color=TERRA)
f_pct = Font(name="Calibri", size=9, color="8A7F73")
f_cgm = Font(name="Calibri", size=10, italic=True, color=CGM_GREY)  # CGM greyed

fill_header = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
fill_section = PatternFill(start_color=SECTION_BG, end_color=SECTION_BG, fill_type="solid")
fill_total = PatternFill(start_color=TOTAL_BG, end_color=TOTAL_BG, fill_type="solid")
fill_grand = PatternFill(start_color=GRAND_BG, end_color=GRAND_BG, fill_type="solid")
fill_cgm = PatternFill(start_color="F0EDE8", end_color="F0EDE8", fill_type="solid")

align_r = Alignment(horizontal="right", vertical="center")
align_l = Alignment(horizontal="left", vertical="center")
align_c = Alignment(horizontal="center", vertical="center")

thin_bottom = Border(bottom=Side(style="thin", color="D8CFC3"))
thick_bottom = Border(bottom=Side(style="medium", color=PRODUCT_SEP))
thick_top_bottom = Border(
    top=Side(style="medium", color=DEEP),
    bottom=Side(style="medium", color=DEEP),
)
product_sep = Border(bottom=Side(style="thick", color=PRODUCT_SEP))

NUM_FMT = '#,##0'
PCT_FMT = '0.0%'
EUR_FMT = '€#,##0'


# ═══════════════════════════════════════════
# SHEET 1: ASSUMPTIONS
# ═══════════════════════════════════════════
ws_a = wb.active
ws_a.title = "Assumptions"
ws_a.sheet_properties.tabColor = SAGE
ws_a.column_dimensions["A"].width = 34
ws_a.column_dimensions["B"].width = 14
ws_a.column_dimensions["C"].width = 48

ws_a.cell(row=1, column=1, value="alche — P&L Assumptions v3").font = f_title
ws_a.cell(row=2, column=1, value="Change any value here → P&L auto-recalculates. All persona findings incorporated.").font = f_subtitle

def put_assumption(row, label, value, note=""):
    ws_a.cell(row=row, column=1, value=label).font = f_label
    c = ws_a.cell(row=row, column=2, value=value)
    c.font = f_val
    c.alignment = align_r
    if note:
        ws_a.cell(row=row, column=3, value=note).font = f_note

def section_header_a(row, label):
    for col in range(1, 4):
        c = ws_a.cell(row=row, column=col)
        c.fill = fill_section
    ws_a.cell(row=row, column=1, value=label).font = f_section

# ─── Subscription Model (rows 4-11) ───
r = 4
section_header_a(r, "SUBSCRIPTION MODEL"); r += 1
put_assumption(r, "Core tier price (EUR/mo)", 19, "Entry tier"); r += 1              # B5
put_assumption(r, "Pro tier price (EUR/mo)", 49, "Sweet spot"); r += 1               # B6
put_assumption(r, "Premium tier price (EUR/mo)", 99, "Highest ARPU"); r += 1         # B7
put_assumption(r, "Monthly churn rate", 0.08, "8% — locked data"); r += 1            # B8
ws_a.cell(row=r-1, column=2).number_format = '0%'
put_assumption(r, "Core tier share", 0.52, "52%"); r += 1                            # B9
ws_a.cell(row=r-1, column=2).number_format = '0%'
put_assumption(r, "Pro tier share", 0.38, "38%"); r += 1                             # B10
ws_a.cell(row=r-1, column=2).number_format = '0%'
put_assumption(r, "Premium tier share", 0.10, "10%"); r += 1                         # B11
ws_a.cell(row=r-1, column=2).number_format = '0%'

# ─── Subscriber Growth (rows 13-19) ───
r = 13
section_header_a(r, "SUBSCRIBER GROWTH (new subs/month)"); r += 1
put_assumption(r, "M1-M3", 0, "Pre-launch — no app yet"); r += 1                    # B14
put_assumption(r, "M4-M6", 15, "Product launch"); r += 1                            # B15
put_assumption(r, "M7-M9", 22, "Growth acceleration"); r += 1                       # B16
put_assumption(r, "M10-M12", 28, "Marketing ramp"); r += 1                          # B17
put_assumption(r, "M13-M18", 32, "Scaling"); r += 1                                 # B18
put_assumption(r, "M19-M24", 38, "Scaling phase"); r += 1                           # B19

# ─── Product Pricing (rows 21-33) ───
r = 21
section_header_a(r, "PRODUCT PRICING"); r += 1
put_assumption(r, "Retail product price (EUR)", 49, "R10 validated"); r += 1           # B22
put_assumption(r, "Retail product COGS (EUR/unit)", 23.50, "R10: wholesale ~48% of retail"); r += 1  # B23
put_assumption(r, "LED session price (EUR)", 0, "Free with smoothie purchase"); r += 1  # B24
put_assumption(r, "LED practitioner cost (EUR)", 0, "Self-service — no practitioner"); r += 1  # B25
put_assumption(r, "Smoothie avg price (EUR)", 9.67, "R2 Cost Analysis"); r += 1       # B26
put_assumption(r, "Smoothie COGS rate", 0.35, "35% ingredients+packaging (R2)"); r += 1  # B27
ws_a.cell(row=r-1, column=2).number_format = '0%'
put_assumption(r, "Event ticket price (EUR)", 35, "Community events"); r += 1         # B28
put_assumption(r, "Event fixed cost (EUR)", 250, "Venue per event"); r += 1           # B29
put_assumption(r, "Ticketing platform fee", 0.07, "7% Luma/Eventbrite"); r += 1       # B30
ws_a.cell(row=r-1, column=2).number_format = '0%'
put_assumption(r, "Doctor SaaS fee (EUR/mo)", 99, "D1: StGB 299a safe"); r += 1       # B31
put_assumption(r, "Doctor verify cost (EUR)", 25, "Per new doctor"); r += 1           # B32
put_assumption(r, "Doctor max clinics", 20, "Cap M1-24"); r += 1                     # B33

# ─── Cost Rates (rows 35-41) ───
r = 35
section_header_a(r, "COST RATES"); r += 1
put_assumption(r, "Waste — blended rate", 0.055, "D7: capsules 4%, powders 7%"); r += 1  # B36
ws_a.cell(row=r-1, column=2).number_format = '0.0%'
put_assumption(r, "3PL fulfillment (EUR/unit)", 2.50, "Retail products only"); r += 1         # B37
put_assumption(r, "Stripe fee rate", 0.029, "2.9%"); r += 1                          # B38
ws_a.cell(row=r-1, column=2).number_format = '0.0%'
put_assumption(r, "AI API per user (EUR/mo)", 0.35, "D6: hybrid Claude+Gemini"); r += 1  # B39
put_assumption(r, "Refund/returns reserve", 0.015, "1.5% of product revenue"); r += 1  # B40
ws_a.cell(row=r-1, column=2).number_format = '0.0%'
put_assumption(r, "Bad debt provision", 0.005, "P5: 0.5% of total revenue"); r += 1   # B41
ws_a.cell(row=r-1, column=2).number_format = '0.0%'

# ─── Payroll (rows 43-52) ───
r = 43
section_header_a(r, "PAYROLL"); r += 1
put_assumption(r, "Founder gross salary (each)", 4166, "EUR 50K/yr (P7: viable)"); r += 1  # B44
put_assumption(r, "Employer multiplier", 1.25, "P2/P7: incl. U1/U2/U3 (was 1.22)"); r += 1  # B45
put_assumption(r, "Space staff gross (M4+)", 1230, "Midijob ~25 hrs/wk"); r += 1     # B46
put_assumption(r, "Partner Mgr gross (M6+)", 2000, "EUR/mo"); r += 1                 # B47
put_assumption(r, "CTO gross (M18+)", 6500, "EUR 78K — P7: below 25th pctile"); r += 1  # B48
put_assumption(r, "BG insurance rate", 0.005, "P7: ~0.5% for office (was 1.5%)"); r += 1  # B49
ws_a.cell(row=r-1, column=2).number_format = '0.0%'
put_assumption(r, "KSK levy rate", 0.049, "P2: 4.9% on creative freelancer fees"); r += 1  # B50
ws_a.cell(row=r-1, column=2).number_format = '0.0%'
put_assumption(r, "Recruitment CTO (M17)", 18000, "P7: headhunter 20-30%"); r += 1   # B51
put_assumption(r, "CTO onboarding (M18)", 3500, "P7: equipment + setup"); r += 1     # B52

# ─── Fixed Operating Costs (rows 55-83) ───
r = 55
section_header_a(r, "FIXED OPERATING COSTS"); r += 1
put_assumption(r, "App maintenance retainer", 1500, "External agency"); r += 1        # B56
put_assumption(r, "Claude Max (internal AI)", 210, "D6: team budget"); r += 1         # B57
put_assumption(r, "SaaS Stack M1-6", 338, "D9: Google, Figma, Notion, Crisp"); r += 1  # B58
put_assumption(r, "SaaS Stack M7-12", 500, "D9: + tools"); r += 1                    # B59
put_assumption(r, "SaaS Stack M13+", 750, "D9: full stack"); r += 1                  # B60
put_assumption(r, "Rent M1-3 (co-working)", 750, "Before space opens"); r += 1        # B61
put_assumption(r, "Rent M4+ (physical space)", 2500, "Kaltmiete"); r += 1             # B62
put_assumption(r, "Nebenkosten M4+", 325, "P3: EUR 4-5/sqm commercial"); r += 1      # B63
put_assumption(r, "Steuerberater", 400, "Tax advisor monthly"); r += 1                # B64
# Insurance split (P6)
put_assumption(r, "General liability insurance", 200, "Betriebshaftpflicht"); r += 1  # B65
put_assumption(r, "D&O insurance", 100, "P6: mandatory with investors"); r += 1       # B66
put_assumption(r, "Cyber insurance", 150, "P6: health data — Art. 9 GDPR"); r += 1   # B67
put_assumption(r, "Product liability insurance", 100, "P6: supplements"); r += 1      # B68
put_assumption(r, "Product recall insurance M4+", 100, "P4: for NEM retail"); r += 1  # B69
# Utilities & space ops (P3)
put_assumption(r, "Utilities M4+ (Strom/Gas/Wasser)", 500, "P3: was EUR 400"); r += 1  # B70
put_assumption(r, "Cleaning/maintenance M4+", 300, "P3: daily cleaning"); r += 1      # B71
put_assumption(r, "Maintenance reserve M4+", 200, "P3: equipment repairs"); r += 1    # B72
put_assumption(r, "Waste disposal M4+", 120, "P3: BSR commercial"); r += 1            # B73
put_assumption(r, "WiFi / internet M4+", 50, "P3: business fiber"); r += 1            # B74
put_assumption(r, "POS system M4+", 50, "Cash register + TSE"); r += 1                # B75
# Compliance & regulatory (P2/P4/P6)
put_assumption(r, "DSB (ext. DPO) monthly", 400, "P6: MANDATORY Art. 37 GDPR"); r += 1  # B76
put_assumption(r, "CMP tool (cookie consent)", 25, "P6: TDDDG required"); r += 1     # B77
put_assumption(r, "Food safety compliance M4+", 80, "P4: HACCP + monitoring"); r += 1  # B78
put_assumption(r, "IHK Berlin (annual ÷ 12)", 10, "P2: EUR 120/yr (was EUR 25)"); r += 1  # B79
put_assumption(r, "GEMA M4+ (annual ÷ 12)", 31, "P2: EUR 370/yr for music"); r += 1  # B80
put_assumption(r, "Rundfunkbeitrag", 6, "P2: EUR 6.12/mo for 0-8 employees"); r += 1  # B81
put_assumption(r, "Verpackungsgesetz (annual ÷ 12)", 6, "P2: LUCID + dual system"); r += 1  # B82

# ─── Marketing (rows 85-90) ───
r = 85
section_header_a(r, "SALES & MARKETING"); r += 1
put_assumption(r, "Content creation tools", 150, "Canva, stock, etc."); r += 1        # B86
put_assumption(r, "Influencer seeding M7+", 200, "Product seeding"); r += 1           # B87
put_assumption(r, "Investor travel", 150, "Monthly"); r += 1                          # B88
put_assumption(r, "Conferences M7+", 100, "Industry events"); r += 1                  # B89
put_assumption(r, "Local transport", 50, "BVG"); r += 1                               # B90

# ─── Buffers (rows 92-94) ───
r = 92
section_header_a(r, "BUFFERS"); r += 1
put_assumption(r, "Contingency buffer rate", 0.10, "10% of OpEx (pre-contingency)"); r += 1  # B93
ws_a.cell(row=r-1, column=2).number_format = '0%'
put_assumption(r, "Bank fees", 35, "P5: Geschaeftskonto monthly"); r += 1             # B94

# ─── Tax (rows 96-99) ───
r = 96
section_header_a(r, "TAX (P2: Berlin GmbH)"); r += 1
put_assumption(r, "Gewerbesteuer rate (Berlin)", 0.1435, "P2: 3.5% × 410% Hebesatz"); r += 1  # B97
ws_a.cell(row=r-1, column=2).number_format = '0.00%'
put_assumption(r, "KSt + Soli rate", 0.15825, "P2: 15% + 5.5% Soli"); r += 1        # B98
ws_a.cell(row=r-1, column=2).number_format = '0.000%'
put_assumption(r, "Combined tax rate (simplified)", 0.30, "P2: ~30.18% rounded down"); r += 1  # B99
ws_a.cell(row=r-1, column=2).number_format = '0%'

# ─── CapEx (rows 101-116) ───
r = 101
section_header_a(r, "CAPITAL EXPENDITURES"); r += 1
put_assumption(r, "GmbH Formation (M1)", 2500, "Notary + registration"); r += 1       # B102
put_assumption(r, "Legal Opinion HWG (M1)", 3500, "D3/R3"); r += 1                   # B103
put_assumption(r, "R&D Stability Testing (M1)", 5000, "Supplement formulation"); r += 1  # B104
put_assumption(r, "DPIA (M1)", 7500, "P6: MANDATORY Art. 35 GDPR"); r += 1           # B105
put_assumption(r, "MDR Classification Opinion (M1)", 7500, "P6: MANDATORY pre-launch"); r += 1  # B106
put_assumption(r, "Privacy Policy + ToS (M1)", 5000, "P6: Art. 13/14 GDPR"); r += 1  # B107
put_assumption(r, "App Design & Development (M1)", 15000, "CapEx: UX/UI capitalized"); r += 1  # B108
put_assumption(r, "Space Buildout (M3)", 25000, "P3: was EUR 15K"); r += 1            # B109
put_assumption(r, "Kitchen Equipment (M3)", 20000, "Smoothie bar"); r += 1            # B110
put_assumption(r, "LED Therapy Devices (M3)", 5000, "MITO LIGHT panels"); r += 1      # B111
put_assumption(r, "Fire Safety Equipment (M3)", 300, "P3: extinguishers + signage"); r += 1  # B112
put_assumption(r, "Signage & Branding (M3)", 500, "P3: window + exterior"); r += 1    # B113
put_assumption(r, "HACCP Plan Setup (M3)", 1000, "P4: consultant"); r += 1            # B114
put_assumption(r, "Hygiene Training (M3)", 60, "P4: IfSG §43 × 4 persons"); r += 1   # B115
put_assumption(r, "Food Safety Equipment (M3)", 500, "P4: thermometers + sensors"); r += 1  # B116
put_assumption(r, "Ablöse — Key Money (M3)", 50000, "Hava parası — commercial lease takeover"); r += 1  # B117
put_assumption(r, "Kaution — Security Deposit (M3)", 7500, "3 months Kaltmiete (refundable, NOT P&L)"); r += 1  # B118
put_assumption(r, "Depreciation period (months)", 36, "Linear 3-year"); r += 1        # B119

# ─── Scenario Multipliers (rows 120-125) ───
r = 120
section_header_a(r, "SCENARIO MULTIPLIERS"); r += 1
put_assumption(r, "Bull — revenue multiplier", 1.3, "Optimistic growth"); r += 1      # B121
put_assumption(r, "Bull — churn multiplier", 0.75, "Lower churn"); r += 1             # B122
put_assumption(r, "Bear — revenue multiplier", 0.6, "Slower growth"); r += 1          # B123
put_assumption(r, "Bear — churn multiplier", 1.5, "Higher churn"); r += 1             # B124
put_assumption(r, "Bear — CAC multiplier", 2.0, "Higher acquisition cost"); r += 1    # B125


# ═══════════════════════════════════════════
# SHEET 2: GROWTH CURVES
# ═══════════════════════════════════════════
ws_g = wb.create_sheet("Growth Curves")
ws_g.sheet_properties.tabColor = "8B9E7C"
ws_g.column_dimensions["A"].width = 26
for col in range(2, 27):
    ws_g.column_dimensions[get_column_letter(col)].width = 9

ws_g.cell(row=1, column=1, value="alche — Monthly Growth Curves (Input Data)").font = f_title
ws_g.cell(row=2, column=1, value="Edit numbers to change unit forecasts. P&L formulas reference these cells.").font = f_subtitle

# Header row (row 3)
ws_g.cell(row=3, column=1, value="Metric").font = f_header
ws_g.cell(row=3, column=1).fill = fill_header
for i in range(24):
    c = ws_g.cell(row=3, column=i+2, value=f"M{i+1}")
    c.font = f_header
    c.fill = fill_header
    c.alignment = align_c

# Growth curve data
curves = {
    4: ("Retail Units", [0,0,0,50,100,150,250,350,450,600,750,900,1100,1300,1500,1800,2100,2400,2700,3000,3400,3800,4200,4600]),
    5: ("LED Sessions", [0,0,0,10,20,35,50,65,80,100,120,140,160,180,200,220,240,260,280,300,300,300,300,300]),
    6: ("Smoothie Units", [0,0,0,100,150,250,350,450,550,650,750,850,950,1050,1150,1250,1350,1450,1550,1650,1750,1850,1950,2000]),
    7: ("Event Count", [0,0,0,1,1,2,2,3,3,3,3,3,3,3,4,4,4,4,4,5,5,5,5,5]),
    8: ("Event Avg Attendees", [0,0,0,12,15,15,18,18,18,20,20,20,22,22,22,22,24,24,25,25,25,25,25,25]),
    9: ("Doctor Clinics", [0,0,0,0,0,2,5,8,10,12,15,17,18,19,20,20,20,20,20,20,20,20,20,20]),
    10: ("Marketing Spend (EUR)", [50,50,100,200,400,600,800,800,1000,1200,1400,1400,1800,1800,2000,2000,2200,2200,2500,2500,2800,2800,3000,3000]),
    11: ("KSK-Eligible Freelancer Spend", [500,500,500,500,800,800,800,800,800,800,800,800,800,800,800,800,800,800,800,800,800,800,800,800]),
}

for row, (label, values) in curves.items():
    ws_g.cell(row=row, column=1, value=label).font = f_sub
    for col, val in enumerate(values, start=2):
        c = ws_g.cell(row=row, column=col, value=val)
        c.font = f_sub
        c.alignment = align_r
        c.number_format = NUM_FMT

# New doctors per month (row 12)
ws_g.cell(row=12, column=1, value="New Doctors (delta)").font = f_dim
ws_g.cell(row=12, column=2, value=0).font = f_dim
for col in range(3, 26):
    c = ws_g.cell(row=12, column=col)
    c.font = f_dim
    prev = get_column_letter(col-1)
    cur = get_column_letter(col)
    c.value = f"=MAX(0,{cur}9-{prev}9)"
    c.number_format = NUM_FMT

# Subscriber model (rows 14-20)
ws_g.cell(row=14, column=1, value="SUBSCRIBER MODEL").font = f_section
ws_g.cell(row=14, column=1).fill = fill_section
for col in range(2, 26):
    ws_g.cell(row=14, column=col).fill = fill_section

# Row 15: New subs rate
ws_g.cell(row=15, column=1, value="New Subs Rate").font = f_dim
for i in range(24):
    col = i + 2
    m = i + 1
    formula = (f'=IF({m}<=3,Assumptions!B14,'
               f'IF({m}<=6,Assumptions!B15,'
               f'IF({m}<=9,Assumptions!B16,'
               f'IF({m}<=12,Assumptions!B17,'
               f'IF({m}<=18,Assumptions!B18,'
               f'Assumptions!B19)))))')
    ws_g.cell(row=15, column=col, value=formula).font = f_dim
    ws_g.cell(row=15, column=col).number_format = NUM_FMT

# Row 16: Churned subs
ws_g.cell(row=16, column=1, value="Churned Subs").font = f_dim
ws_g.cell(row=16, column=2, value=0)
for col in range(3, 26):
    prev_col = get_column_letter(col - 1)
    ws_g.cell(row=16, column=col, value=f"=ROUND({prev_col}17*Assumptions!B8,0)").font = f_dim
    ws_g.cell(row=16, column=col).number_format = NUM_FMT

# Row 17: Total Subscribers
ws_g.cell(row=17, column=1, value="Total Subscribers").font = f_total
ws_g.cell(row=17, column=2, value=f"=MAX(0,B15-B16)").font = f_total
ws_g.cell(row=17, column=2).number_format = NUM_FMT
for col in range(3, 26):
    prev_col = get_column_letter(col - 1)
    cur_col = get_column_letter(col)
    ws_g.cell(row=17, column=col, value=f"=MAX(0,{prev_col}17-{cur_col}16+{cur_col}15)").font = f_total
    ws_g.cell(row=17, column=col).number_format = NUM_FMT

# Row 18-20: Tier breakdown
ws_g.cell(row=18, column=1, value="  Core Subs (52%)").font = f_dim
ws_g.cell(row=19, column=1, value="  Pro Subs (38%)").font = f_dim
ws_g.cell(row=20, column=1, value="  Premium Subs (10%)").font = f_dim
for col in range(2, 26):
    cur = get_column_letter(col)
    ws_g.cell(row=18, column=col, value=f"=ROUND({cur}17*Assumptions!B9,0)").font = f_dim
    ws_g.cell(row=18, column=col).number_format = NUM_FMT
    ws_g.cell(row=19, column=col, value=f"=ROUND({cur}17*Assumptions!B10,0)").font = f_dim
    ws_g.cell(row=19, column=col).number_format = NUM_FMT
    ws_g.cell(row=20, column=col, value=f"=MAX(0,{cur}17-{cur}18-{cur}19)").font = f_dim
    ws_g.cell(row=20, column=col).number_format = NUM_FMT


# ═══════════════════════════════════════════
# SHEET 3: P&L (FORMULA-BASED)
# ═══════════════════════════════════════════
ws = wb.create_sheet("P&L")
ws.sheet_properties.tabColor = AMBER
ws.column_dimensions["A"].width = 36
for col in range(2, 27):
    ws.column_dimensions[get_column_letter(col)].width = 11
ws.freeze_panes = "B5"

# ─── Title ───
ws.cell(row=1, column=1, value="alche — 24-Month P&L v3").font = f_title
ws.cell(row=2, column=1, value="Pre-Seed | EUR 500K @ EUR 2.5M Cap | All 7 persona audits applied").font = f_subtitle
ws.cell(row=3, column=1, value="All amounts in EUR net of VAT. Click any cell to see the formula.").font = f_note

# ─── Column headers (row 4) ───
ws.cell(row=4, column=1, value="Budget P&L (EUR)")
for i in range(24):
    ws.cell(row=4, column=i+2, value=f"M{i+1}")
ws.cell(row=4, column=26, value="TOTAL")
for c in range(1, 27):
    cell = ws.cell(row=4, column=c)
    cell.font = f_header
    cell.fill = fill_header
    cell.alignment = align_c

# ─── Helper functions ───
GC = "'Growth Curves'"
AS = "Assumptions"

def style_row(ws, row, style, cols=range(1, 27)):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        if style == "header":
            cell.font = f_header; cell.fill = fill_header
        elif style == "section":
            cell.font = f_section; cell.fill = fill_section
        elif style == "total":
            cell.font = f_total; cell.fill = fill_total; cell.border = thin_bottom
        elif style == "grand":
            cell.font = f_grand; cell.fill = fill_grand; cell.border = thick_top_bottom
        elif style == "sub":
            cell.font = f_sub
        elif style == "dim":
            cell.font = f_dim
        elif style == "unit":
            cell.font = f_unit
        elif style == "cgm":
            cell.font = f_cgm; cell.fill = fill_cgm

def write_label(row, label, style="sub"):
    ws.cell(row=row, column=1, value=label)
    style_row(ws, row, style)

def write_formula_row(row, label, formulas, style="sub", fmt=NUM_FMT, total_formula=None):
    ws.cell(row=row, column=1, value=label)
    for i, formula in enumerate(formulas):
        c = ws.cell(row=row, column=i+2, value=formula)
        c.alignment = align_r
        c.number_format = fmt
    if total_formula:
        ws.cell(row=row, column=26, value=total_formula)
    else:
        ws.cell(row=row, column=26, value=f"=SUM(B{row}:Y{row})")
    ws.cell(row=row, column=26).alignment = align_r
    ws.cell(row=row, column=26).number_format = fmt
    style_row(ws, row, style)

def write_values_row(row, label, values, style="sub", fmt=NUM_FMT):
    ws.cell(row=row, column=1, value=label)
    for i, val in enumerate(values):
        c = ws.cell(row=row, column=i+2, value=val)
        c.alignment = align_r
        c.number_format = fmt
    ws.cell(row=row, column=26, value=f"=SUM(B{row}:Y{row})")
    ws.cell(row=row, column=26).alignment = align_r
    ws.cell(row=row, column=26).number_format = fmt
    style_row(ws, row, style)

def sep_row(row):
    for c in range(1, 27):
        ws.cell(row=row, column=c).border = product_sep

def month_col(m):
    return get_column_letter(m + 2)

def gc_ref(row, col):
    return f"{GC}!{get_column_letter(col)}{row}"

def a_ref(row):
    return f"{AS}!B{row}"

# ═══════════════════════════════════════════
# P&L ROWS
# ═══════════════════════════════════════════
ROW = 5

# ─── SUBSCRIBER METRICS ───
write_label(ROW, "SUBSCRIBER METRICS", "header"); ROW += 1
formulas = [f"={gc_ref(15, i+2)}" for i in range(24)]
write_formula_row(ROW, "  New Subs Added", formulas, "dim"); ROW += 1
formulas = [f"={gc_ref(16, i+2)}" for i in range(24)]
write_formula_row(ROW, "  Churned Subs", formulas, "dim"); ROW += 1
formulas = [f"={gc_ref(17, i+2)}" for i in range(24)]
write_formula_row(ROW, "  Total Subscribers", formulas, "total"); ROW += 1
formulas = [f"={gc_ref(18, i+2)}" for i in range(24)]
write_formula_row(ROW, "    Core (52%)", formulas, "dim"); ROW += 1
formulas = [f"={gc_ref(19, i+2)}" for i in range(24)]
write_formula_row(ROW, "    Pro (38%)", formulas, "dim"); ROW += 1
formulas = [f"={gc_ref(20, i+2)}" for i in range(24)]
write_formula_row(ROW, "    Premium (10%)", formulas, "dim"); ROW += 1

ROW += 1

# ═══════════════════════════════════════════
# INCOME
# ═══════════════════════════════════════════
write_label(ROW, "INCOME", "header"); INCOME_START = ROW; ROW += 1

# ─── RETAIL PRODUCTS ───
write_label(ROW, "  Retail Products (D11: EUR 49 validated, COGS EUR 23.50)", "section"); ROW += 1
pot_units_row = ROW
formulas = [f"={gc_ref(4, i+2)}" for i in range(24)]
write_formula_row(ROW, "    Units sold", formulas, "unit"); ROW += 1
write_formula_row(ROW, "    Price/unit", [f"={a_ref(22)}"]*24, "unit", '#,##0.00'); ROW += 1
pot_rev_row = ROW
formulas = [f"={month_col(i)}{pot_units_row}*{a_ref(22)}" for i in range(24)]
write_formula_row(ROW, "    Revenue", formulas, "sub"); ROW += 1
sep_row(ROW - 1)

# ─── LED THERAPY ───
write_label(ROW, "  LED Therapy (Free with Smoothie)", "section"); ROW += 1
led_units_row = ROW
formulas = [f"={gc_ref(5, i+2)}" for i in range(24)]
write_formula_row(ROW, "    Sessions (self-service)", formulas, "unit"); ROW += 1
write_formula_row(ROW, "    Price/session", [f"={a_ref(24)}"]*24, "unit", '#,##0.00'); ROW += 1
led_rev_row = ROW
formulas = [f"={month_col(i)}{led_units_row}*{a_ref(24)}" for i in range(24)]
write_formula_row(ROW, "    Revenue (EUR 0 — free amenity)", formulas, "sub"); ROW += 1
sep_row(ROW - 1)

# ─── SMOOTHIE BAR ───
write_label(ROW, "  Smoothie Bar", "section"); ROW += 1
sm_units_row = ROW
formulas = [f"={gc_ref(6, i+2)}" for i in range(24)]
write_formula_row(ROW, "    Units sold", formulas, "unit"); ROW += 1
write_formula_row(ROW, "    Price/smoothie", [f"={a_ref(26)}"]*24, "unit", '#,##0.00'); ROW += 1
sm_rev_row = ROW
formulas = [f"=ROUND({month_col(i)}{sm_units_row}*{a_ref(26)},0)" for i in range(24)]
write_formula_row(ROW, "    Revenue", formulas, "sub"); ROW += 1
sep_row(ROW - 1)

# ─── EVENTS ───
write_label(ROW, "  Community Events", "section"); ROW += 1
ev_count_row = ROW
formulas = [f"={gc_ref(7, i+2)}" for i in range(24)]
write_formula_row(ROW, "    Event count", formulas, "unit"); ROW += 1
ev_attend_row = ROW
formulas = [f"={gc_ref(8, i+2)}" for i in range(24)]
write_formula_row(ROW, "    Avg attendees/event", formulas, "unit"); ROW += 1
ev_total_attend_row = ROW
formulas = [f"={month_col(i)}{ev_count_row}*{month_col(i)}{ev_attend_row}" for i in range(24)]
write_formula_row(ROW, "    Total attendees", formulas, "unit"); ROW += 1
write_formula_row(ROW, "    Price/ticket", [f"={a_ref(28)}"]*24, "unit", '#,##0.00'); ROW += 1
ev_rev_row = ROW
formulas = [f"={month_col(i)}{ev_total_attend_row}*{a_ref(28)}" for i in range(24)]
write_formula_row(ROW, "    Revenue", formulas, "sub"); ROW += 1
sep_row(ROW - 1)

# ─── DOCTOR SAAS ───
write_label(ROW, "  Doctor SaaS (D1: EUR 99/mo)", "section"); ROW += 1
doc_subs_row = ROW
formulas = [f"={gc_ref(9, i+2)}" for i in range(24)]
write_formula_row(ROW, "    Clinics listed", formulas, "unit"); ROW += 1
write_formula_row(ROW, "    Fee/month", [f"={a_ref(31)}"]*24, "unit", '#,##0.00'); ROW += 1
doc_rev_row = ROW
formulas = [f"={month_col(i)}{doc_subs_row}*{a_ref(31)}" for i in range(24)]
write_formula_row(ROW, "    Revenue", formulas, "sub"); ROW += 1
sep_row(ROW - 1)

# ─── CGM (D3: Commission Model — Premium 3-mo gate or add-on) ───
cgm_rev_row = ROW
# EUR 15/user/month commission; starts M7 (Premium 3-month gate); add-on buyers also eligible M7+
# Zero COGS — pure referral/affiliate margin; no inventory, no RMA
CGM_COMMISSION = [0, 0, 0, 0, 0, 0, 45, 90, 135, 180, 225, 270, 315, 360, 405, 450, 480, 510, 540, 570, 585, 600, 615, 630]
write_values_row(ROW, "  CGM Referral Commission (D3: EUR 15/user/mo — Premium 3-mo gate or add-on)", CGM_COMMISSION, "sub")
ROW += 1
sep_row(ROW - 1)

# ─── RESTAURANT (Phase 2 — EUR 0) ───
rest_rev_row = ROW
write_values_row(ROW, "  Restaurant (D2: Phase 2 — EUR 0)", [0]*24, "dim"); ROW += 1
sep_row(ROW - 1)

# ─── SUBSCRIPTIONS ───
write_label(ROW, "  Subscriptions", "section"); ROW += 1
sub_core_row = ROW
formulas = [f"={gc_ref(18, i+2)}*{a_ref(5)}" for i in range(24)]
write_formula_row(ROW, "    Core (EUR 19/mo)", formulas, "sub"); ROW += 1
sub_pro_row = ROW
formulas = [f"={gc_ref(19, i+2)}*{a_ref(6)}" for i in range(24)]
write_formula_row(ROW, "    Pro (EUR 49/mo)", formulas, "sub"); ROW += 1
sub_prem_row = ROW
formulas = [f"={gc_ref(20, i+2)}*{a_ref(7)}" for i in range(24)]
write_formula_row(ROW, "    Premium (EUR 99/mo)", formulas, "sub"); ROW += 1
sub_total_row = ROW
formulas = [f"={month_col(i)}{sub_core_row}+{month_col(i)}{sub_pro_row}+{month_col(i)}{sub_prem_row}" for i in range(24)]
write_formula_row(ROW, "    Subscription MRR", formulas, "total"); ROW += 1
sep_row(ROW - 1)

# ─── TOTAL INCOME ───
ROW += 1
total_income_row = ROW
rev_rows = [pot_rev_row, led_rev_row, sm_rev_row, ev_rev_row, doc_rev_row, cgm_rev_row, sub_total_row]
formulas = [f"={'+'.join(f'{month_col(i)}{r}' for r in rev_rows)}" for i in range(24)]
write_formula_row(ROW, "TOTAL INCOME", formulas, "grand"); ROW += 1

ROW += 1

# ═══════════════════════════════════════════
# COGS
# ═══════════════════════════════════════════
write_label(ROW, "COST OF GOODS SOLD (COGS)", "header"); ROW += 1

# Retail COGS
write_label(ROW, "  Retail COGS", "section"); ROW += 1
cogs_pot_row = ROW
formulas = [f"=ROUND({month_col(i)}{pot_units_row}*{a_ref(23)},0)" for i in range(24)]
write_formula_row(ROW, "    Product cost (EUR 23.50/unit)", formulas, "sub"); ROW += 1
cogs_waste_row = ROW
formulas = [f"=ROUND({month_col(i)}{pot_rev_row}*{a_ref(36)},0)" for i in range(24)]
write_formula_row(ROW, "    Waste/spoilage (5.5%)", formulas, "sub"); ROW += 1
cogs_3pl_row = ROW
formulas = [f"=IF({i+1}>=4,ROUND({month_col(i)}{pot_units_row}*{a_ref(37)},0),0)" for i in range(24)]
write_formula_row(ROW, "    3PL fulfillment (EUR 2.50/unit)", formulas, "sub"); ROW += 1
sep_row(ROW - 1)

# LED COGS
cogs_led_row = ROW
formulas = [f"=ROUND({month_col(i)}{led_units_row}*{a_ref(25)},0)" for i in range(24)]
write_formula_row(ROW, "  LED COGS (free — no practitioner)", formulas, "sub"); ROW += 1
sep_row(ROW - 1)

# Smoothie COGS
write_label(ROW, "  Smoothie Bar COGS", "section"); ROW += 1
cogs_sm_row = ROW
formulas = [f"=ROUND({month_col(i)}{sm_rev_row}*{a_ref(27)},0)" for i in range(24)]
write_formula_row(ROW, "    Ingredients + packaging (35%)", formulas, "sub"); ROW += 1
sep_row(ROW - 1)

# Events COGS
cogs_ev_row = ROW
formulas = [f"=ROUND({month_col(i)}{ev_count_row}*{a_ref(29)}+{month_col(i)}{ev_rev_row}*{a_ref(30)},0)" for i in range(24)]
write_formula_row(ROW, "  Events & Ticketing", formulas, "sub"); ROW += 1
sep_row(ROW - 1)

# Doctor COGS
cogs_doc_row = ROW
formulas = [f"=ROUND({gc_ref(12, i+2)}*{a_ref(32)},0)" for i in range(24)]
write_formula_row(ROW, "  Doctor Verification (EUR 25/new)", formulas, "sub"); ROW += 1

# AI API COGS
cogs_ai_row = ROW
formulas = [f"=ROUND({gc_ref(17, i+2)}*{a_ref(39)},0)" for i in range(24)]
write_formula_row(ROW, "  AI API — User-Facing (EUR 0.35/sub)", formulas, "sub"); ROW += 1

# Refund Reserve
cogs_refund_row = ROW
formulas = [f"=ROUND(({month_col(i)}{pot_rev_row}+{month_col(i)}{sm_rev_row}+{month_col(i)}{ev_rev_row})*{a_ref(40)},0)" for i in range(24)]
write_formula_row(ROW, "  Refund/Returns Reserve (1.5%)", formulas, "sub"); ROW += 1

# Bad Debt Provision (NEW — P5)
cogs_bad_debt_row = ROW
formulas = [f"=ROUND({month_col(i)}{total_income_row}*{a_ref(41)},0)" for i in range(24)]
write_formula_row(ROW, "  Bad Debt Provision (0.5%)", formulas, "sub"); ROW += 1

# Stripe
cogs_stripe_row = ROW
formulas = [f"=ROUND({month_col(i)}{total_income_row}*{a_ref(38)},0)" for i in range(24)]
write_formula_row(ROW, "  Stripe Processing (2.9%)", formulas, "sub"); ROW += 1

# TOTAL COGS
ROW += 1
total_cogs_row = ROW
cogs_rows = [cogs_pot_row, cogs_waste_row, cogs_3pl_row, cogs_led_row, cogs_sm_row,
             cogs_ev_row, cogs_doc_row, cogs_ai_row, cogs_refund_row, cogs_bad_debt_row, cogs_stripe_row]
formulas = [f"={'+'.join(f'{month_col(i)}{r}' for r in cogs_rows)}" for i in range(24)]
write_formula_row(ROW, "TOTAL COGS", formulas, "grand"); ROW += 1

# GROSS PROFIT
ROW += 1
gp_row = ROW
formulas = [f"={month_col(i)}{total_income_row}-{month_col(i)}{total_cogs_row}" for i in range(24)]
write_formula_row(ROW, "GROSS PROFIT", formulas, "grand"); ROW += 1
gm_row = ROW
formulas = [f"=IF({month_col(i)}{total_income_row}>0,{month_col(i)}{gp_row}/{month_col(i)}{total_income_row},0)" for i in range(24)]
write_formula_row(ROW, "  Gross Margin %", formulas, "dim", PCT_FMT,
                  total_formula=f"=IF(Z{total_income_row}>0,Z{gp_row}/Z{total_income_row},0)"); ROW += 1

ROW += 1

# ═══════════════════════════════════════════
# OPERATING EXPENSES
# ═══════════════════════════════════════════
write_label(ROW, "OPERATING EXPENSES", "header"); ROW += 1

# ─── PAYROLL ───
write_label(ROW, "  Payroll", "section"); ROW += 1
pay_f1_row = ROW
formulas = [f"=ROUND({a_ref(44)}*{a_ref(45)},0)"]*24
write_formula_row(ROW, "    Timu (CEO)", formulas, "sub"); ROW += 1
pay_f2_row = ROW
write_formula_row(ROW, "    Daria (COO)", formulas, "sub"); ROW += 1
pay_space_row = ROW
formulas = [f"=IF({i+1}>=4,ROUND({a_ref(46)}*{a_ref(45)},0),0)" for i in range(24)]
write_formula_row(ROW, "    Space Staff (M4+)", formulas, "sub"); ROW += 1
pay_mgr_row = ROW
formulas = [f"=IF({i+1}>=6,ROUND({a_ref(47)}*{a_ref(45)},0),0)" for i in range(24)]
write_formula_row(ROW, "    Partner Mgr (M6+)", formulas, "sub"); ROW += 1
pay_cto_row = ROW
formulas = [f"=IF({i+1}>=18,ROUND({a_ref(48)}*{a_ref(45)},0),0)" for i in range(24)]
write_formula_row(ROW, "    CTO (M18+)", formulas, "sub"); ROW += 1
# BG Insurance
pay_bg_row = ROW
pay_rows_for_bg = [pay_f1_row, pay_f2_row, pay_space_row, pay_mgr_row, pay_cto_row]
formulas = [f"=ROUND(({'+'.join(f'{month_col(i)}{r}' for r in pay_rows_for_bg)})*{a_ref(49)},0)" for i in range(24)]
write_formula_row(ROW, "    BG Insurance (0.5% payroll)", formulas, "sub"); ROW += 1
# KSK Levy (NEW — P2/P7)
pay_ksk_row = ROW
formulas = [f"=ROUND({gc_ref(11, i+2)}*{a_ref(50)},0)" for i in range(24)]
write_formula_row(ROW, "    KSK Levy (4.9% on freelancers)", formulas, "sub"); ROW += 1
# Payroll total
pay_total_row = ROW
all_pay = pay_rows_for_bg + [pay_bg_row, pay_ksk_row]
formulas = [f"={'+'.join(f'{month_col(i)}{r}' for r in all_pay)}" for i in range(24)]
write_formula_row(ROW, "    Payroll Total", formulas, "total"); ROW += 1
sep_row(ROW - 1)

# ─── R&D ───
write_label(ROW, "  R&D", "section"); ROW += 1
rd_app_row = ROW
formulas = [f"={a_ref(56)}"]*24
write_formula_row(ROW, "    App Maintenance", formulas, "sub"); ROW += 1
rd_total_row = ROW
formulas = [f"={month_col(i)}{rd_app_row}" for i in range(24)]
write_formula_row(ROW, "    R&D Total", formulas, "total"); ROW += 1
sep_row(ROW - 1)

# ─── S&M ───
write_label(ROW, "  Sales & Marketing", "section"); ROW += 1
sm_ads_row = ROW
formulas = [f"={gc_ref(10, i+2)}" for i in range(24)]
write_formula_row(ROW, "    Paid Ads", formulas, "sub"); ROW += 1
sm_content_row = ROW
formulas = [f"={a_ref(86)}"]*24
write_formula_row(ROW, "    Content Creation", formulas, "sub"); ROW += 1
sm_inf_row = ROW
formulas = [f"=IF({i+1}>=7,{a_ref(87)},0)" for i in range(24)]
write_formula_row(ROW, "    Influencer Seeding (M7+)", formulas, "sub"); ROW += 1
sm_total_row = ROW
formulas = [f"={month_col(i)}{sm_ads_row}+{month_col(i)}{sm_content_row}+{month_col(i)}{sm_inf_row}" for i in range(24)]
write_formula_row(ROW, "    S&M Total", formulas, "total"); ROW += 1
sep_row(ROW - 1)

# ─── G&A ───
write_label(ROW, "  General & Admin", "section"); ROW += 1
ga_rent_row = ROW
formulas = [f"=IF({i+1}<=3,{a_ref(61)},{a_ref(62)})" for i in range(24)]
write_formula_row(ROW, "    Rent (co-work M1-3 / space M4+)", formulas, "sub"); ROW += 1
# Nebenkosten (NEW — P3)
ga_neben_row = ROW
formulas = [f"=IF({i+1}>=4,{a_ref(63)},0)" for i in range(24)]
write_formula_row(ROW, "    Nebenkosten M4+ (P3)", formulas, "sub"); ROW += 1
ga_steuer_row = ROW
formulas = [f"={a_ref(64)}"]*24
write_formula_row(ROW, "    Steuerberater", formulas, "sub"); ROW += 1
# Insurance — split (P6)
ga_ins_gen_row = ROW
formulas = [f"={a_ref(65)}"]*24
write_formula_row(ROW, "    General Liability Insurance", formulas, "sub"); ROW += 1
ga_ins_do_row = ROW
formulas = [f"={a_ref(66)}"]*24
write_formula_row(ROW, "    D&O Insurance (P6)", formulas, "sub"); ROW += 1
ga_ins_cyber_row = ROW
formulas = [f"={a_ref(67)}"]*24
write_formula_row(ROW, "    Cyber Insurance (P6)", formulas, "sub"); ROW += 1
ga_ins_prod_row = ROW
formulas = [f"={a_ref(68)}"]*24
write_formula_row(ROW, "    Product Liability Insurance (P6)", formulas, "sub"); ROW += 1
ga_ins_recall_row = ROW
formulas = [f"=IF({i+1}>=4,{a_ref(69)},0)" for i in range(24)]
write_formula_row(ROW, "    Product Recall Insurance M4+ (P4)", formulas, "sub"); ROW += 1
# Utilities & space (P3)
ga_util_row = ROW
formulas = [f"=IF({i+1}>=4,{a_ref(70)},0)" for i in range(24)]
write_formula_row(ROW, "    Utilities (M4+)", formulas, "sub"); ROW += 1
ga_clean_row = ROW
formulas = [f"=IF({i+1}>=4,{a_ref(71)},0)" for i in range(24)]
write_formula_row(ROW, "    Cleaning/Maintenance (M4+)", formulas, "sub"); ROW += 1
ga_maint_row = ROW
formulas = [f"=IF({i+1}>=4,{a_ref(72)},0)" for i in range(24)]
write_formula_row(ROW, "    Maintenance Reserve M4+ (P3)", formulas, "sub"); ROW += 1
ga_waste_row = ROW
formulas = [f"=IF({i+1}>=4,{a_ref(73)},0)" for i in range(24)]
write_formula_row(ROW, "    Waste Disposal M4+ (P3)", formulas, "sub"); ROW += 1
ga_wifi_row = ROW
formulas = [f"=IF({i+1}>=4,{a_ref(74)},0)" for i in range(24)]
write_formula_row(ROW, "    WiFi/Internet M4+ (P3)", formulas, "sub"); ROW += 1
ga_pos_row = ROW
formulas = [f"=IF({i+1}>=4,{a_ref(75)},0)" for i in range(24)]
write_formula_row(ROW, "    POS System (M4+)", formulas, "sub"); ROW += 1
# Compliance (P2/P4/P6)
ga_dsb_row = ROW
formulas = [f"={a_ref(76)}"]*24
write_formula_row(ROW, "    DSB / ext. DPO (P6: mandatory)", formulas, "sub"); ROW += 1
ga_cmp_row = ROW
formulas = [f"={a_ref(77)}"]*24
write_formula_row(ROW, "    CMP Cookie Tool (P6)", formulas, "sub"); ROW += 1
ga_food_row = ROW
formulas = [f"=IF({i+1}>=4,{a_ref(78)},0)" for i in range(24)]
write_formula_row(ROW, "    Food Safety Compliance M4+ (P4)", formulas, "sub"); ROW += 1
ga_ihk_row = ROW
formulas = [f"={a_ref(79)}"]*24
write_formula_row(ROW, "    IHK Berlin (P2: EUR 120/yr)", formulas, "sub"); ROW += 1
ga_gema_row = ROW
formulas = [f"=IF({i+1}>=4,{a_ref(80)},0)" for i in range(24)]
write_formula_row(ROW, "    GEMA M4+ (P2)", formulas, "sub"); ROW += 1
ga_rundfunk_row = ROW
formulas = [f"={a_ref(81)}"]*24
write_formula_row(ROW, "    Rundfunkbeitrag (P2)", formulas, "sub"); ROW += 1
ga_verpack_row = ROW
formulas = [f"={a_ref(82)}"]*24
write_formula_row(ROW, "    Verpackungsgesetz (P2)", formulas, "sub"); ROW += 1
ga_bank_row = ROW
formulas = [f"={a_ref(94)}"]*24
write_formula_row(ROW, "    Bank Fees (P5)", formulas, "sub"); ROW += 1
# Permits
ga_permit_row = ROW
formulas = [f"=IF({i+1}=3,500,0)" for i in range(24)]
write_formula_row(ROW, "    Permits & Licenses (M3)", formulas, "sub"); ROW += 1

# G&A total
ga_total_row = ROW
ga_rows = [ga_rent_row, ga_neben_row, ga_steuer_row,
           ga_ins_gen_row, ga_ins_do_row, ga_ins_cyber_row, ga_ins_prod_row, ga_ins_recall_row,
           ga_util_row, ga_clean_row, ga_maint_row, ga_waste_row, ga_wifi_row, ga_pos_row,
           ga_dsb_row, ga_cmp_row, ga_food_row, ga_ihk_row, ga_gema_row, ga_rundfunk_row, ga_verpack_row,
           ga_bank_row, ga_permit_row]
formulas = [f"={'+'.join(f'{month_col(i)}{r}' for r in ga_rows)}" for i in range(24)]
write_formula_row(ROW, "    G&A Total", formulas, "total"); ROW += 1
sep_row(ROW - 1)

# ─── Management & APIs ───
write_label(ROW, "  Management & APIs", "section"); ROW += 1
mg_saas_row = ROW
formulas = [f"=IF({i+1}<=6,{a_ref(58)},IF({i+1}<=12,{a_ref(59)},{a_ref(60)}))" for i in range(24)]
write_formula_row(ROW, "    SaaS Stack (D9)", formulas, "sub"); ROW += 1
mg_claude_row = ROW
formulas = [f"={a_ref(57)}"]*24
write_formula_row(ROW, "    Claude Max (internal)", formulas, "sub"); ROW += 1
mg_total_row = ROW
formulas = [f"={month_col(i)}{mg_saas_row}+{month_col(i)}{mg_claude_row}" for i in range(24)]
write_formula_row(ROW, "    Management Total", formulas, "total"); ROW += 1
sep_row(ROW - 1)

# ─── Travel ───
write_label(ROW, "  Travel", "section"); ROW += 1
tr_inv_row = ROW
formulas = [f"={a_ref(88)}"]*24
write_formula_row(ROW, "    Investor Travel", formulas, "sub"); ROW += 1
tr_conf_row = ROW
formulas = [f"=IF({i+1}>=7,{a_ref(89)},0)" for i in range(24)]
write_formula_row(ROW, "    Conferences (M7+)", formulas, "sub"); ROW += 1
tr_local_row = ROW
formulas = [f"={a_ref(90)}"]*24
write_formula_row(ROW, "    Local Transport", formulas, "sub"); ROW += 1
tr_total_row = ROW
formulas = [f"={month_col(i)}{tr_inv_row}+{month_col(i)}{tr_conf_row}+{month_col(i)}{tr_local_row}" for i in range(24)]
write_formula_row(ROW, "    Travel Total", formulas, "total"); ROW += 1
sep_row(ROW - 1)

# ─── One-off People Costs ───
recruit_row = ROW
formulas = [f"=IF({i+1}=17,{a_ref(51)},0)" for i in range(24)]
write_formula_row(ROW, "  Recruitment — CTO (M17) (P7)", formulas, "sub"); ROW += 1
onboard_row = ROW
formulas = [f"=IF({i+1}=18,{a_ref(52)},0)" for i in range(24)]
write_formula_row(ROW, "  CTO Onboarding/Equipment (M18) (P7)", formulas, "sub"); ROW += 1

# ─── TOTAL OPEX (pre-contingency) ───
ROW += 1
opex_pre_row = ROW
opex_section_rows = [pay_total_row, rd_total_row, sm_total_row, ga_total_row, mg_total_row, tr_total_row, recruit_row, onboard_row]
formulas = [f"={'+'.join(f'{month_col(i)}{r}' for r in opex_section_rows)}" for i in range(24)]
write_formula_row(ROW, "  OpEx (pre-contingency)", formulas, "total"); ROW += 1

# Contingency
contingency_row = ROW
formulas = [f"=ROUND({month_col(i)}{opex_pre_row}*{a_ref(93)},0)" for i in range(24)]
write_formula_row(ROW, "  Contingency Buffer (10%)", formulas, "sub"); ROW += 1

# TOTAL OPEX
ROW += 1
total_opex_row = ROW
formulas = [f"={month_col(i)}{opex_pre_row}+{month_col(i)}{contingency_row}" for i in range(24)]
write_formula_row(ROW, "TOTAL OPEX", formulas, "grand"); ROW += 1

# ═══════════════════════════════════════════
# PROFITABILITY BRIDGE
# ═══════════════════════════════════════════
ROW += 1
write_label(ROW, "PROFITABILITY BRIDGE", "header"); ROW += 1

# Note row explaining EBITDA
ws.cell(row=ROW, column=1, value="EBITDA = Earnings Before Interest, Tax, Depreciation & Amortization").font = f_note
ROW += 1

# Repeat key totals for readability (formula references, not hardcoded)
bridge_income_row = ROW
formulas = [f"={month_col(i)}{total_income_row}" for i in range(24)]
write_formula_row(ROW, "  Total Income", formulas, "sub"); ROW += 1

bridge_cogs_row = ROW
formulas = [f"=-{month_col(i)}{total_cogs_row}" for i in range(24)]
write_formula_row(ROW, "  − Total COGS", formulas, "sub"); ROW += 1

bridge_gp_row = ROW
formulas = [f"={month_col(i)}{gp_row}" for i in range(24)]
write_formula_row(ROW, "  = Gross Profit", formulas, "total"); ROW += 1

bridge_gm_row = ROW
formulas = [f"=IF({month_col(i)}{total_income_row}>0,{month_col(i)}{gp_row}/{month_col(i)}{total_income_row},0)" for i in range(24)]
write_formula_row(ROW, "    Gross Margin %", formulas, "dim", PCT_FMT,
                  total_formula=f"=IF(Z{total_income_row}>0,Z{gp_row}/Z{total_income_row},0)"); ROW += 1

bridge_opex_row = ROW
formulas = [f"=-{month_col(i)}{total_opex_row}" for i in range(24)]
write_formula_row(ROW, "  − Total OpEx", formulas, "sub"); ROW += 1

sep_row(ROW - 1)

# EBITDA
ebitda_row = ROW
formulas = [f"={month_col(i)}{gp_row}-{month_col(i)}{total_opex_row}" for i in range(24)]
write_formula_row(ROW, "  = EBITDA", formulas, "grand"); ROW += 1

ebitda_margin_row = ROW
formulas = [f"=IF({month_col(i)}{total_income_row}>0,{month_col(i)}{ebitda_row}/{month_col(i)}{total_income_row},0)" for i in range(24)]
write_formula_row(ROW, "    EBITDA Margin %", formulas, "dim", PCT_FMT,
                  total_formula=f"=IF(Z{total_income_row}>0,Z{ebitda_row}/Z{total_income_row},0)"); ROW += 1

ROW += 1

# ═══════════════════════════════════════════
# CapEx
# ═══════════════════════════════════════════
write_label(ROW, "CAPITAL EXPENDITURES (CapEx)", "header"); ROW += 1
# M1 items
capex_gmbh_row = ROW
formulas = [f"=IF({i+1}=1,{a_ref(102)},0)" for i in range(24)]
write_formula_row(ROW, "  GmbH Formation (M1)", formulas, "sub"); ROW += 1
capex_legal_row = ROW
formulas = [f"=IF({i+1}=1,{a_ref(103)},0)" for i in range(24)]
write_formula_row(ROW, "  Legal Opinion HWG (M1)", formulas, "sub"); ROW += 1
capex_rd_row = ROW
formulas = [f"=IF({i+1}=1,{a_ref(104)},0)" for i in range(24)]
write_formula_row(ROW, "  R&D Stability Testing (M1)", formulas, "sub"); ROW += 1
capex_dpia_row = ROW
formulas = [f"=IF({i+1}=1,{a_ref(105)},0)" for i in range(24)]
write_formula_row(ROW, "  DPIA (M1) (P6: mandatory)", formulas, "sub"); ROW += 1
capex_mdr_row = ROW
formulas = [f"=IF({i+1}=1,{a_ref(106)},0)" for i in range(24)]
write_formula_row(ROW, "  MDR Classification Opinion (M1) (P6)", formulas, "sub"); ROW += 1
capex_privacy_row = ROW
formulas = [f"=IF({i+1}=1,{a_ref(107)},0)" for i in range(24)]
write_formula_row(ROW, "  Privacy Policy + ToS (M1) (P6)", formulas, "sub"); ROW += 1
capex_appdesign_row = ROW
formulas = [f"=IF({i+1}=1,{a_ref(108)},0)" for i in range(24)]
write_formula_row(ROW, "  App Design & Development (M1)", formulas, "sub"); ROW += 1
# M3 items
capex_build_row = ROW
formulas = [f"=IF({i+1}=3,{a_ref(109)},0)" for i in range(24)]
write_formula_row(ROW, "  Space Buildout (M3) (P3: was EUR 15K)", formulas, "sub"); ROW += 1
capex_kitchen_row = ROW
formulas = [f"=IF({i+1}=3,{a_ref(110)},0)" for i in range(24)]
write_formula_row(ROW, "  Kitchen Equipment (M3)", formulas, "sub"); ROW += 1
capex_led_row = ROW
formulas = [f"=IF({i+1}=3,{a_ref(111)},0)" for i in range(24)]
write_formula_row(ROW, "  LED Devices (M3)", formulas, "sub"); ROW += 1
capex_fire_row = ROW
formulas = [f"=IF({i+1}=3,{a_ref(112)},0)" for i in range(24)]
write_formula_row(ROW, "  Fire Safety Equipment (M3) (P3)", formulas, "sub"); ROW += 1
capex_sign_row = ROW
formulas = [f"=IF({i+1}=3,{a_ref(113)},0)" for i in range(24)]
write_formula_row(ROW, "  Signage & Branding (M3) (P3)", formulas, "sub"); ROW += 1
capex_haccp_row = ROW
formulas = [f"=IF({i+1}=3,{a_ref(114)},0)" for i in range(24)]
write_formula_row(ROW, "  HACCP Plan Setup (M3) (P4)", formulas, "sub"); ROW += 1
capex_hygiene_row = ROW
formulas = [f"=IF({i+1}=3,{a_ref(115)},0)" for i in range(24)]
write_formula_row(ROW, "  Hygiene Training (M3) (P4)", formulas, "sub"); ROW += 1
capex_foodequip_row = ROW
formulas = [f"=IF({i+1}=3,{a_ref(116)},0)" for i in range(24)]
write_formula_row(ROW, "  Food Safety Equipment (M3) (P4)", formulas, "sub"); ROW += 1
# Ablöse — Key Money (M3) — capitalized intangible asset, amortized over lease term
capex_ablose_row = ROW
formulas = [f"=IF({i+1}=3,{a_ref(117)},0)" for i in range(24)]
write_formula_row(ROW, "  Ablöse — Key Money (M3)", formulas, "sub"); ROW += 1

# Depreciation (includes Ablöse amortization in 36-month pool)
dep_row = ROW
formulas = []
for i in range(24):
    m = i + 1
    m1_items = f"{a_ref(102)}+{a_ref(103)}+{a_ref(104)}+{a_ref(105)}+{a_ref(106)}+{a_ref(107)}+{a_ref(108)}"
    m3_items = f"{a_ref(109)}+{a_ref(110)}+{a_ref(111)}+{a_ref(112)}+{a_ref(113)}+{a_ref(114)}+{a_ref(115)}+{a_ref(116)}+{a_ref(117)}"
    if m >= 3:
        formulas.append(f"=ROUND(({m1_items}+{m3_items})/{a_ref(119)},0)")
    elif m >= 1:
        formulas.append(f"=ROUND(({m1_items})/{a_ref(119)},0)")
    else:
        formulas.append(0)
write_formula_row(ROW, "  Depreciation (linear / 36mo)", formulas, "dim"); ROW += 1

# EBIT = EBITDA − Depreciation
ebit_row = ROW
formulas = [f"={month_col(i)}{ebitda_row}-{month_col(i)}{dep_row}" for i in range(24)]
write_formula_row(ROW, "  = EBIT (EBITDA − D&A)", formulas, "total"); ROW += 1

ebit_margin_row = ROW
formulas = [f"=IF({month_col(i)}{total_income_row}>0,{month_col(i)}{ebit_row}/{month_col(i)}{total_income_row},0)" for i in range(24)]
write_formula_row(ROW, "    EBIT Margin %", formulas, "dim", PCT_FMT,
                  total_formula=f"=IF(Z{total_income_row}>0,Z{ebit_row}/Z{total_income_row},0)"); ROW += 1

# Total CapEx (cash)
ROW += 1
total_capex_row = ROW
capex_cash_rows = [capex_gmbh_row, capex_legal_row, capex_rd_row, capex_dpia_row, capex_mdr_row, capex_privacy_row,
                   capex_appdesign_row, capex_build_row, capex_kitchen_row, capex_led_row, capex_fire_row, capex_sign_row,
                   capex_haccp_row, capex_hygiene_row, capex_foodequip_row, capex_ablose_row]
formulas = [f"={'+'.join(f'{month_col(i)}{r}' for r in capex_cash_rows)}" for i in range(24)]
write_formula_row(ROW, "TOTAL CAPEX (cash)", formulas, "grand"); ROW += 1

# ═══════════════════════════════════════════
# NET P&L (before tax)
# ═══════════════════════════════════════════
ROW += 1
net_pretax_row = ROW
formulas = [f"={month_col(i)}{ebitda_row}-{month_col(i)}{total_capex_row}" for i in range(24)]
write_formula_row(ROW, "NET P&L (before tax)", formulas, "grand"); ROW += 1

# ─── Tax Provision (NEW — P2) ───
ROW += 1
write_label(ROW, "TAX PROVISION (P2: ~30% when profitable)", "header"); ROW += 1
# Cumulative pre-tax P&L
cum_pretax_row = ROW
ws.cell(row=ROW, column=1, value="  Cumulative Pre-Tax P&L")
ws.cell(row=ROW, column=2, value=f"=B{net_pretax_row}")
ws.cell(row=ROW, column=2).number_format = NUM_FMT
ws.cell(row=ROW, column=2).alignment = align_r
for i in range(1, 24):
    col = i + 2
    prev = get_column_letter(col - 1)
    cur = get_column_letter(col)
    ws.cell(row=ROW, column=col, value=f"={prev}{cum_pretax_row}+{cur}{net_pretax_row}")
    ws.cell(row=ROW, column=col).number_format = NUM_FMT
    ws.cell(row=ROW, column=col).alignment = align_r
ws.cell(row=ROW, column=26, value=f"=Y{cum_pretax_row}")
ws.cell(row=ROW, column=26).number_format = NUM_FMT
style_row(ws, ROW, "dim"); ROW += 1

# Tax provision: 30% of monthly profit ONLY when cumulative is positive
tax_row = ROW
formulas = []
for i in range(24):
    col_letter = month_col(i)
    formulas.append(f"=IF(AND({col_letter}{net_pretax_row}>0,{col_letter}{cum_pretax_row}>0),ROUND({col_letter}{net_pretax_row}*{a_ref(99)},0),0)")
write_formula_row(ROW, "  Tax Provision (~30%)", formulas, "sub"); ROW += 1

# NET P&L (after tax)
ROW += 1
net_pnl_row = ROW
formulas = [f"={month_col(i)}{net_pretax_row}-{month_col(i)}{tax_row}" for i in range(24)]
write_formula_row(ROW, "NET P&L (after tax)", formulas, "grand"); ROW += 1

# ─── Cumulative Cash ───
cum_cash_row = ROW
ws.cell(row=ROW, column=1, value="Cumulative Cash (EUR 500K start)")
ws.cell(row=ROW, column=2, value=f"=500000+B{net_pnl_row}")
ws.cell(row=ROW, column=2).number_format = NUM_FMT
ws.cell(row=ROW, column=2).alignment = align_r
for i in range(1, 24):
    col = i + 2
    prev_col = get_column_letter(col - 1)
    cur_col = get_column_letter(col)
    ws.cell(row=ROW, column=col, value=f"={prev_col}{cum_cash_row}+{cur_col}{net_pnl_row}")
    ws.cell(row=ROW, column=col).number_format = NUM_FMT
    ws.cell(row=ROW, column=col).alignment = align_r
ws.cell(row=ROW, column=26, value=f"=Y{cum_cash_row}")
ws.cell(row=ROW, column=26).number_format = NUM_FMT
style_row(ws, ROW, "total"); ROW += 1

ROW += 1

# ═══════════════════════════════════════════
# KEY METRICS
# ═══════════════════════════════════════════
write_label(ROW, "KEY METRICS", "header"); ROW += 1
# Monthly burn
burn_row = ROW
formulas = [f"=IF({month_col(i)}{net_pnl_row}<0,ABS({month_col(i)}{net_pnl_row}),0)" for i in range(24)]
write_formula_row(ROW, "  Monthly Burn (if loss)", formulas, "dim"); ROW += 1
# Revenue per sub
arpu_row = ROW
formulas = [f"=IF({gc_ref(17, i+2)}>0,ROUND({month_col(i)}{total_income_row}/{gc_ref(17, i+2)},2),0)" for i in range(24)]
write_formula_row(ROW, "  Revenue / Subscriber (Total ARPU)", formulas, "dim", '#,##0.00'); ROW += 1
# Sub-only ARPU (P5 finding)
sub_arpu_row = ROW
formulas = [f"=IF({gc_ref(17, i+2)}>0,ROUND({month_col(i)}{sub_total_row}/{gc_ref(17, i+2)},2),0)" for i in range(24)]
write_formula_row(ROW, "  Subscription ARPU (P5: actual)", formulas, "dim", '#,##0.00'); ROW += 1
# Subscription MRR
formulas = [f"={month_col(i)}{sub_total_row}" for i in range(24)]
write_formula_row(ROW, "  Subscription MRR", formulas, "dim"); ROW += 1
# Sub revenue as % of total (P5 finding)
formulas = [f"=IF({month_col(i)}{total_income_row}>0,{month_col(i)}{sub_total_row}/{month_col(i)}{total_income_row},0)" for i in range(24)]
write_formula_row(ROW, "  Subscription % of Total Revenue", formulas, "dim", PCT_FMT,
                  total_formula=f"=IF(Z{total_income_row}>0,Z{sub_total_row}/Z{total_income_row},0)"); ROW += 1
# Gross margin %
formulas = [f"={month_col(i)}{gm_row}" for i in range(24)]
write_formula_row(ROW, "  Gross Margin %", formulas, "dim", PCT_FMT,
                  total_formula=f"=Z{gm_row}"); ROW += 1
# Months of runway
formulas = [f"=IF({month_col(i)}{burn_row}>0,ROUND({month_col(i)}{cum_cash_row}/{month_col(i)}{burn_row},1),99)" for i in range(24)]
write_formula_row(ROW, "  Months of Runway Remaining", formulas, "dim", '#,##0.0'); ROW += 1

# ─── BREAK-EVEN INDICATORS ───
ROW += 1
write_label(ROW, "BREAK-EVEN INDICATORS", "header"); ROW += 1
# EBITDA break-even flag: 1 if EBITDA > 0, else 0
ebitda_be_row = ROW
formulas = [f'=IF({month_col(i)}{ebitda_row}>0,1,0)' for i in range(24)]
write_formula_row(ROW, "  EBITDA Positive? (1=Yes)", formulas, "dim",
                  total_formula=f'=MATCH(1,B{ROW}:Y{ROW},0)'); ROW += 1
# Net P&L (after tax) break-even flag
net_be_row = ROW
formulas = [f'=IF({month_col(i)}{net_pnl_row}>0,1,0)' for i in range(24)]
write_formula_row(ROW, "  Net P&L Positive? (1=Yes)", formulas, "dim",
                  total_formula=f'=MATCH(1,B{ROW}:Y{ROW},0)'); ROW += 1
# Cumulative cash vs starting capital
cum_be_row = ROW
formulas = [f'=IF({month_col(i)}{cum_cash_row}>500000,1,0)' for i in range(24)]
write_formula_row(ROW, "  Cash > EUR 500K? (1=Yes)", formulas, "dim",
                  total_formula=f'=IFERROR(MATCH(1,B{ROW}:Y{ROW},0),"Never")'); ROW += 1
# EBITDA break-even month (first month where EBITDA > 0)
ebitda_be_month_row = ROW
ws.cell(row=ROW, column=1, value="  EBITDA Break-Even Month")
ws.cell(row=ROW, column=2, value=f'=IFERROR("M"&MATCH(1,B{ebitda_be_row}:Y{ebitda_be_row},0),"Never")')
ws.cell(row=ROW, column=2).alignment = align_r
style_row(ws, ROW, "total"); ROW += 1
# Net P&L break-even month
net_be_month_row = ROW
ws.cell(row=ROW, column=1, value="  Net P&L Break-Even Month")
ws.cell(row=ROW, column=2, value=f'=IFERROR("M"&MATCH(1,B{net_be_row}:Y{net_be_row},0),"Never")')
ws.cell(row=ROW, column=2).alignment = align_r
style_row(ws, ROW, "total"); ROW += 1


# ═══════════════════════════════════════════
# SHEET 4: CASH FLOW STATEMENT
# ═══════════════════════════════════════════
ws_cf = wb.create_sheet("Cash Flow")
ws_cf.sheet_properties.tabColor = TERRA
ws_cf.column_dimensions["A"].width = 36
for col in range(2, 27):
    ws_cf.column_dimensions[get_column_letter(col)].width = 11
ws_cf.freeze_panes = "B4"

ws_cf.cell(row=1, column=1, value="alche — Cash Flow Statement v3").font = f_title
ws_cf.cell(row=2, column=1, value="P5: EUR ~218K gap between P&L profit and actual cash. This sheet closes that gap.").font = f_subtitle

# Headers
for i in range(24):
    c = ws_cf.cell(row=3, column=i+2, value=f"M{i+1}")
    c.font = f_header; c.fill = fill_header; c.alignment = align_c
ws_cf.cell(row=3, column=1, value="Cash Flow (EUR)").font = f_header
ws_cf.cell(row=3, column=1).fill = fill_header
ws_cf.cell(row=3, column=26, value="TOTAL").font = f_header
ws_cf.cell(row=3, column=26).fill = fill_header
ws_cf.cell(row=3, column=26).alignment = align_c

PL = "'P&L'"
cf_row = 4

def cf_label(row, label, style="sub"):
    ws_cf.cell(row=row, column=1, value=label)
    for c in range(1, 27):
        cell = ws_cf.cell(row=row, column=c)
        if style == "header":
            cell.font = f_header; cell.fill = fill_header
        elif style == "section":
            cell.font = f_section; cell.fill = fill_section
        elif style == "total":
            cell.font = f_total; cell.fill = fill_total
        elif style == "grand":
            cell.font = f_grand; cell.fill = fill_grand; cell.border = thick_top_bottom
        elif style == "sub":
            cell.font = f_sub
        elif style == "dim":
            cell.font = f_dim

def cf_formula_row(row, label, formulas, style="sub", fmt=NUM_FMT):
    ws_cf.cell(row=row, column=1, value=label)
    for i, f_str in enumerate(formulas):
        c = ws_cf.cell(row=row, column=i+2, value=f_str)
        c.alignment = align_r; c.number_format = fmt
    ws_cf.cell(row=row, column=26, value=f"=SUM(B{row}:Y{row})")
    ws_cf.cell(row=row, column=26).alignment = align_r
    ws_cf.cell(row=row, column=26).number_format = fmt
    cf_label(row, label, style)

# Operating cash flow (start from EBITDA — standard format)
cf_label(cf_row, "OPERATING ACTIVITIES", "header"); cf_row += 1
# EBITDA
cf_ebitda_row = cf_row
formulas = [f"={PL}!{month_col(i)}{ebitda_row}" for i in range(24)]
cf_formula_row(cf_row, "  EBITDA", formulas, "sub"); cf_row += 1
# Tax provision (cash outflow)
cf_tax_row = cf_row
formulas = [f"=-{PL}!{month_col(i)}{tax_row}" for i in range(24)]
cf_formula_row(cf_row, "  - Tax Provision", formulas, "sub"); cf_row += 1
# Working capital: inventory build (P5 critical finding)
cf_wc_row = cf_row
# Inventory change = delta in (retail units × COGS × 2 months pipeline)
formulas = []
for i in range(24):
    if i == 0:
        formulas.append(f"=-ROUND({gc_ref(4, 2)}*{a_ref(23)}*2,0)")
    else:
        formulas.append(f"=-ROUND(MAX(0,({gc_ref(4, i+2)}-{gc_ref(4, i+1)})*{a_ref(23)}*2),0)")
cf_formula_row(cf_row, "  - Working Capital Build (inventory)", formulas, "sub"); cf_row += 1

# Operating cash flow total
cf_op_total = cf_row
formulas = [f"={month_col(i)}{cf_ebitda_row}+{month_col(i)}{cf_tax_row}+{month_col(i)}{cf_wc_row}" for i in range(24)]
for i, f_str in enumerate(formulas):
    c = ws_cf.cell(row=cf_row, column=i+2, value=f_str)
    c.alignment = align_r; c.number_format = NUM_FMT
ws_cf.cell(row=cf_row, column=26, value=f"=SUM(B{cf_row}:Y{cf_row})")
ws_cf.cell(row=cf_row, column=26).alignment = align_r; ws_cf.cell(row=cf_row, column=26).number_format = NUM_FMT
cf_label(cf_row, "  Operating Cash Flow", "total"); cf_row += 1

cf_row += 1
cf_label(cf_row, "INVESTING ACTIVITIES", "header"); cf_row += 1
# CapEx (shown as negative)
cf_capex_row = cf_row
formulas = [f"=-{PL}!{month_col(i)}{total_capex_row}" for i in range(24)]
cf_formula_row(cf_row, "  CapEx", formulas, "sub"); cf_row += 1
# Kaution — security deposit (cash outflow, NOT P&L — balance sheet asset, refundable)
cf_kaution_row = cf_row
formulas = [f"=IF({i+1}=3,-{a_ref(118)},0)" for i in range(24)]
cf_formula_row(cf_row, "  Kaution — Security Deposit (M3)", formulas, "sub"); cf_row += 1

cf_inv_total = cf_row
formulas = [f"={month_col(i)}{cf_capex_row}+{month_col(i)}{cf_kaution_row}" for i in range(24)]
for i, f_str in enumerate(formulas):
    c = ws_cf.cell(row=cf_row, column=i+2, value=f_str)
    c.alignment = align_r; c.number_format = NUM_FMT
ws_cf.cell(row=cf_row, column=26, value=f"=SUM(B{cf_row}:Y{cf_row})")
cf_label(cf_row, "  Investing Cash Flow", "total"); cf_row += 1

cf_row += 1
cf_label(cf_row, "FINANCING ACTIVITIES", "header"); cf_row += 1
cf_funding_row = cf_row
formulas = [f"=IF({i+1}=1,500000,0)" for i in range(24)]
cf_formula_row(cf_row, "  Pre-Seed Funding (M1)", formulas, "sub"); cf_row += 1

cf_fin_total = cf_row
formulas = [f"={month_col(i)}{cf_funding_row}" for i in range(24)]
for i, f_str in enumerate(formulas):
    c = ws_cf.cell(row=cf_row, column=i+2, value=f_str)
    c.alignment = align_r; c.number_format = NUM_FMT
ws_cf.cell(row=cf_row, column=26, value=f"=SUM(B{cf_row}:Y{cf_row})")
cf_label(cf_row, "  Financing Cash Flow", "total"); cf_row += 1

# Net cash flow
cf_row += 1
cf_net_cf_row = cf_row
formulas = [f"={month_col(i)}{cf_op_total}+{month_col(i)}{cf_inv_total}+{month_col(i)}{cf_fin_total}" for i in range(24)]
for i, f_str in enumerate(formulas):
    c = ws_cf.cell(row=cf_row, column=i+2, value=f_str)
    c.alignment = align_r; c.number_format = NUM_FMT
ws_cf.cell(row=cf_row, column=26, value=f"=SUM(B{cf_row}:Y{cf_row})")
ws_cf.cell(row=cf_row, column=26).alignment = align_r; ws_cf.cell(row=cf_row, column=26).number_format = NUM_FMT
cf_label(cf_row, "NET CASH FLOW", "grand"); cf_row += 1

# Cumulative cash (cash flow version)
cf_cum_row = cf_row
ws_cf.cell(row=cf_row, column=1, value="CUMULATIVE CASH BALANCE")
ws_cf.cell(row=cf_row, column=2, value=f"=B{cf_net_cf_row}")
ws_cf.cell(row=cf_row, column=2).number_format = NUM_FMT
ws_cf.cell(row=cf_row, column=2).alignment = align_r
for i in range(1, 24):
    col = i + 2
    prev = get_column_letter(col - 1)
    cur = get_column_letter(col)
    ws_cf.cell(row=cf_row, column=col, value=f"={prev}{cf_cum_row}+{cur}{cf_net_cf_row}")
    ws_cf.cell(row=cf_row, column=col).number_format = NUM_FMT
    ws_cf.cell(row=cf_row, column=col).alignment = align_r
ws_cf.cell(row=cf_row, column=26, value=f"=Y{cf_cum_row}")
ws_cf.cell(row=cf_row, column=26).number_format = NUM_FMT
cf_label(cf_row, "CUMULATIVE CASH BALANCE", "grand"); cf_row += 1

# P&L vs Cash gap
cf_row += 1
cf_label(cf_row, "P&L vs CASH GAP ANALYSIS", "header"); cf_row += 1
cf_pnl_cum = cf_row
formulas = [f"={PL}!{month_col(i)}{cum_cash_row}" for i in range(24)]
cf_formula_row(cf_row, "  P&L Cumulative Cash", formulas, "dim"); cf_row += 1
cf_actual_cash = cf_row
formulas = [f"={month_col(i)}{cf_cum_row}" for i in range(24)]
cf_formula_row(cf_row, "  Actual Cash (this sheet)", formulas, "dim"); cf_row += 1
cf_gap = cf_row
formulas = [f"={month_col(i)}{cf_actual_cash}-{month_col(i)}{cf_pnl_cum}" for i in range(24)]
cf_formula_row(cf_row, "  GAP (Working Capital Drag)", formulas, "total"); cf_row += 1


# ═══════════════════════════════════════════
# SHEET 5: UNIT ECONOMICS
# ═══════════════════════════════════════════
ws_ue = wb.create_sheet("Unit Economics")
ws_ue.sheet_properties.tabColor = ROSE
ws_ue.column_dimensions["A"].width = 36
for col in range(2, 27):
    ws_ue.column_dimensions[get_column_letter(col)].width = 11
ws_ue.freeze_panes = "B4"

ws_ue.cell(row=1, column=1, value="alche — Unit Economics Dashboard").font = f_title
ws_ue.cell(row=2, column=1, value="P1/P5: LTV:CAC, payback period, contribution margin per stream").font = f_subtitle

for i in range(24):
    c = ws_ue.cell(row=3, column=i+2, value=f"M{i+1}")
    c.font = f_header; c.fill = fill_header; c.alignment = align_c
ws_ue.cell(row=3, column=1, value="Metric").font = f_header
ws_ue.cell(row=3, column=1).fill = fill_header
ws_ue.cell(row=3, column=26, value="TOTAL").font = f_header
ws_ue.cell(row=3, column=26).fill = fill_header

ue_row = 4

def ue_formula_row(row, label, formulas, style="sub", fmt=NUM_FMT):
    ws_ue.cell(row=row, column=1, value=label)
    for i, f_str in enumerate(formulas):
        c = ws_ue.cell(row=row, column=i+2, value=f_str)
        c.alignment = align_r; c.number_format = fmt
    # For most metrics, total is average or last value
    ws_ue.cell(row=row, column=26, value=f"=Y{row}")
    ws_ue.cell(row=row, column=26).alignment = align_r
    ws_ue.cell(row=row, column=26).number_format = fmt
    for c in range(1, 27):
        cell = ws_ue.cell(row=row, column=c)
        if style == "header":
            cell.font = f_header; cell.fill = fill_header
        elif style == "section":
            cell.font = f_section; cell.fill = fill_section
        elif style == "total":
            cell.font = f_total; cell.fill = fill_total
        elif style == "sub":
            cell.font = f_sub
        elif style == "dim":
            cell.font = f_dim

# Subscription ARPU
for c in range(1, 27):
    ws_ue.cell(row=ue_row, column=c).font = f_section
    ws_ue.cell(row=ue_row, column=c).fill = fill_section
ws_ue.cell(row=ue_row, column=1, value="SUBSCRIPTION METRICS"); ue_row += 1

formulas = [f"=IF({gc_ref(17, i+2)}>0,ROUND({PL}!{month_col(i)}{sub_total_row}/{gc_ref(17, i+2)},2),0)" for i in range(24)]
ue_formula_row(ue_row, "  Subscription ARPU (EUR/sub/mo)", formulas, "sub", '#,##0.00'); ue_row += 1

# LTV at current churn
formulas = [f"=IF({gc_ref(17, i+2)}>0,ROUND(({PL}!{month_col(i)}{sub_total_row}/{gc_ref(17, i+2)})/{a_ref(8)},0),0)" for i in range(24)]
ue_formula_row(ue_row, "  Subscription LTV (ARPU / churn)", formulas, "sub"); ue_row += 1

# CAC (marketing spend / new subs)
formulas = [f"=IF({gc_ref(15, i+2)}>0,ROUND({gc_ref(10, i+2)}/{gc_ref(15, i+2)},0),0)" for i in range(24)]
ue_formula_row(ue_row, "  CAC (paid ads / new subs)", formulas, "sub"); ue_row += 1

# LTV:CAC
ue_ltv_row = ue_row - 2
ue_cac_row = ue_row - 1
formulas = [f"=IF({month_col(i)}{ue_cac_row}>0,ROUND({month_col(i)}{ue_ltv_row}/{month_col(i)}{ue_cac_row},1),0)" for i in range(24)]
ue_formula_row(ue_row, "  LTV:CAC Ratio", formulas, "total", '#,##0.0'); ue_row += 1

# Payback period
formulas = [f"=IF(AND({month_col(i)}{ue_row-3}>0,{month_col(i)}{ue_row-4}>0),ROUND({month_col(i)}{ue_cac_row}/{month_col(i)}{ue_row-4},1),0)" for i in range(24)]
ue_formula_row(ue_row, "  CAC Payback (months)", formulas, "sub", '#,##0.0'); ue_row += 1

ue_row += 1

# Per-stream contribution margins
for c in range(1, 27):
    ws_ue.cell(row=ue_row, column=c).font = f_section
    ws_ue.cell(row=ue_row, column=c).fill = fill_section
ws_ue.cell(row=ue_row, column=1, value="CONTRIBUTION MARGIN BY STREAM"); ue_row += 1

# Retail margin
formulas = [f"=IF({PL}!{month_col(i)}{pot_rev_row}>0,({PL}!{month_col(i)}{pot_rev_row}-{PL}!{month_col(i)}{cogs_pot_row}-{PL}!{month_col(i)}{cogs_waste_row}-{PL}!{month_col(i)}{cogs_3pl_row})/{PL}!{month_col(i)}{pot_rev_row},0)" for i in range(24)]
ue_formula_row(ue_row, "  Retail Contribution Margin %", formulas, "sub", PCT_FMT); ue_row += 1

# LED margin
formulas = [f"=IF({PL}!{month_col(i)}{led_rev_row}>0,({PL}!{month_col(i)}{led_rev_row}-{PL}!{month_col(i)}{cogs_led_row})/{PL}!{month_col(i)}{led_rev_row},0)" for i in range(24)]
ue_formula_row(ue_row, "  LED Therapy Contribution Margin %", formulas, "sub", PCT_FMT); ue_row += 1

# Smoothie margin
formulas = [f"=IF({PL}!{month_col(i)}{sm_rev_row}>0,({PL}!{month_col(i)}{sm_rev_row}-{PL}!{month_col(i)}{cogs_sm_row})/{PL}!{month_col(i)}{sm_rev_row},0)" for i in range(24)]
ue_formula_row(ue_row, "  Smoothie Bar Contribution Margin %", formulas, "sub", PCT_FMT); ue_row += 1

# Subscription margin (digital = ~87%)
formulas = [f"=IF({PL}!{month_col(i)}{sub_total_row}>0,({PL}!{month_col(i)}{sub_total_row}-{PL}!{month_col(i)}{cogs_ai_row})/{PL}!{month_col(i)}{sub_total_row},0)" for i in range(24)]
ue_formula_row(ue_row, "  Subscription Contribution Margin %", formulas, "sub", PCT_FMT); ue_row += 1

ue_row += 1
# Revenue mix
for c in range(1, 27):
    ws_ue.cell(row=ue_row, column=c).font = f_section
    ws_ue.cell(row=ue_row, column=c).fill = fill_section
ws_ue.cell(row=ue_row, column=1, value="REVENUE MIX (P5: only 4.7% subs at M24)"); ue_row += 1

for label, rev_ref in [("  Retail %", pot_rev_row), ("  LED %", led_rev_row),
                         ("  Smoothie %", sm_rev_row), ("  Events %", ev_rev_row),
                         ("  Doctor SaaS %", doc_rev_row), ("  Subscriptions %", sub_total_row)]:
    formulas = [f"=IF({PL}!{month_col(i)}{total_income_row}>0,{PL}!{month_col(i)}{rev_ref}/{PL}!{month_col(i)}{total_income_row},0)" for i in range(24)]
    ue_formula_row(ue_row, label, formulas, "dim", PCT_FMT); ue_row += 1


# ═══════════════════════════════════════════
# SHEET 6: SCENARIO SUMMARY
# ═══════════════════════════════════════════
ws_sc = wb.create_sheet("Scenarios")
ws_sc.sheet_properties.tabColor = "9E948A"
ws_sc.column_dimensions["A"].width = 36
ws_sc.column_dimensions["B"].width = 18
ws_sc.column_dimensions["C"].width = 18
ws_sc.column_dimensions["D"].width = 18

ws_sc.cell(row=1, column=1, value="alche — Scenario Summary").font = f_title
ws_sc.cell(row=2, column=1, value="Base = current model. Bull/Bear use multipliers from Assumptions sheet.").font = f_subtitle

# Headers
for col, h in enumerate(["Metric", "BASE", "BULL (1.3x rev)", "BEAR (0.6x rev)"], 1):
    c = ws_sc.cell(row=4, column=col, value=h)
    c.font = f_header; c.fill = fill_header; c.alignment = align_c

sc_row = 5

def sc_metric(row, label, base_formula, bull_formula, bear_formula, fmt=NUM_FMT):
    ws_sc.cell(row=row, column=1, value=label).font = f_sub
    for col, f_str in enumerate([base_formula, bull_formula, bear_formula], 2):
        c = ws_sc.cell(row=row, column=col, value=f_str)
        c.font = f_val; c.alignment = align_r; c.number_format = fmt

# Key metrics — reference P&L totals (column Z = col 26)
sc_metric(sc_row, "Total Revenue (24-mo)",
          f"={PL}!Z{total_income_row}",
          f"=ROUND({PL}!Z{total_income_row}*{a_ref(121)},0)",
          f"=ROUND({PL}!Z{total_income_row}*{a_ref(123)},0)"); sc_row += 1

sc_metric(sc_row, "Total COGS (24-mo)",
          f"={PL}!Z{total_cogs_row}",
          f"=ROUND({PL}!Z{total_cogs_row}*{a_ref(121)},0)",
          f"=ROUND({PL}!Z{total_cogs_row}*{a_ref(123)},0)"); sc_row += 1

sc_metric(sc_row, "Gross Profit (24-mo)",
          f"={PL}!Z{gp_row}",
          f"=B{sc_row-2}-B{sc_row-1}+ROUND(({PL}!Z{gp_row}-({PL}!Z{total_income_row}-{PL}!Z{total_cogs_row}))*0,0)",
          f"=ROUND({PL}!Z{gp_row}*{a_ref(123)},0)"); sc_row += 1
# Simplify: bull GP = bull rev - bull cogs
ws_sc.cell(row=sc_row-1, column=3, value=f"=C{sc_row-3}-C{sc_row-2}")

sc_metric(sc_row, "Total OpEx (24-mo)",
          f"={PL}!Z{total_opex_row}",
          f"={PL}!Z{total_opex_row}",
          f"={PL}!Z{total_opex_row}"); sc_row += 1

sc_metric(sc_row, "Total CapEx (24-mo)",
          f"={PL}!Z{total_capex_row}",
          f"={PL}!Z{total_capex_row}",
          f"={PL}!Z{total_capex_row}"); sc_row += 1

sc_metric(sc_row, "EBITDA (24-mo)",
          f"=B{sc_row-3}-B{sc_row-2}",
          f"=C{sc_row-3}-C{sc_row-2}",
          f"=D{sc_row-3}-D{sc_row-2}"); sc_row += 1

sc_metric(sc_row, "Net P&L (24-mo, pre-tax)",
          f"=B{sc_row-1}-B{sc_row-2}",
          f"=C{sc_row-1}-C{sc_row-2}",
          f"=D{sc_row-1}-D{sc_row-2}"); sc_row += 1

sc_metric(sc_row, "Estimated Cash at M24",
          f"=500000+B{sc_row-1}",
          f"=500000+C{sc_row-1}",
          f"=500000+D{sc_row-1}"); sc_row += 1

sc_row += 1
# Highlight row
for col in range(1, 5):
    ws_sc.cell(row=sc_row, column=col).fill = fill_section
ws_sc.cell(row=sc_row, column=1, value="INTERPRETATION").font = f_section
sc_row += 1

ws_sc.cell(row=sc_row, column=1, value="Bear case: shows if EUR 500K is sufficient").font = f_note
ws_sc.cell(row=sc_row, column=1).alignment = Alignment(wrap_text=True)
sc_row += 1
ws_sc.cell(row=sc_row, column=1, value="If Bear cash < 0 → raise needs to increase or scope must be reduced").font = f_note
sc_row += 1
ws_sc.cell(row=sc_row, column=1, value="Bull case: shows working capital stress at higher volume (P5)").font = f_note


# ═══════════════════════════════════════════
# SHEET 7: DECISIONS
# ═══════════════════════════════════════════
ws3 = wb.create_sheet("Decisions")
ws3.sheet_properties.tabColor = TERRA
ws3.column_dimensions["A"].width = 8
ws3.column_dimensions["B"].width = 28
ws3.column_dimensions["C"].width = 50
ws3.column_dimensions["D"].width = 25
ws3.column_dimensions["E"].width = 40

ws3.cell(row=1, column=1, value="alche — Decision Log (D1-D11) + Persona Audit").font = f_title
ws3.cell(row=2, column=1, value="All decisions from P&L sessions, 2026-02-22/23. 7 persona audits completed.").font = f_subtitle

for col, h in enumerate(["#", "Decision", "Choice Made", "Old Value", "P&L Impact"], 1):
    c = ws3.cell(row=4, column=col, value=h)
    c.font = f_header; c.fill = fill_header

decisions = [
    ["D1", "Doctor Revenue", "C. SaaS EUR 99/mo × 20 clinics", "EUR 150/mo × 865", "Revenue -75%; legally safe (StGB §299a)"],
    ["D2", "Restaurant Revenue", "B. Phase 2 (500+ users)", "EUR 1.50/cover M6+", "EUR 0 all 24 months"],
    ["D3", "CGM Hardware", "Commission — Premium 3-mo gate or add-on", "EUR 130 × units (old)", "EUR 15/user/mo referral commission; starts M7; Premium gate 3-mo; COGS = 0"],
    ["D4", "Physical Space", "A. Smoothie bar + space", "EUR 0", "New rev + EUR 2,825/mo (rent+Nebenkosten) + CapEx"],
    ["D5", "CapEx Section", "Yes, include", "All OpEx", "EUR ~148K total (v3.3: +EUR 50K Ablöse)"],
    ["D6", "AI Cost Model", "Both internal + user-facing", "EUR 210 + EUR 499 Terra", "EUR 210 + EUR 0.35/sub (R9)"],
    ["D7", "Product Waste", "Category-weighted", "Flat 5%", "5.5% blended"],
    ["D8", "Terra CGM API", "Defer — free HealthKit", "EUR 499/mo M6+", "Saves EUR 9,481"],
    ["D9", "SaaS Stack", "Itemized (R8)", "EUR 350→1800", "EUR 338→500→750"],
    ["D10", "Menu Analysis", "EUR 0 — founders do it", "EUR 500/mo", "Saves EUR 12,000"],
    ["D11", "Retail Products", "EUR 49 validated; COGS EUR 23.50", "EUR 11.50 COGS", "Margin 76% → 52% (R10)"],
]

for i, dec in enumerate(decisions):
    row = 5 + i
    for col, val in enumerate(dec, 1):
        c = ws3.cell(row=row, column=col, value=val)
        c.font = f_sub
        if col == 1:
            c.font = Font(name="Calibri", size=10, bold=True, color=TERRA)
        c.border = thin_bottom

# v3.1 / v3.2 corrections
row = 18
ws3.cell(row=row, column=1, value="MODEL CORRECTIONS (v3.1 / v3.2)").font = f_section
ws3.cell(row=row, column=1).fill = fill_section
for c in range(2, 6):
    ws3.cell(row=row, column=c).fill = fill_section
row += 1

for col, h in enumerate(["#", "Correction", "Change Made", "Previous", "Impact"], 1):
    c = ws3.cell(row=row, column=col, value=h)
    c.font = f_header; c.fill = fill_header
row += 1

corrections = [
    ["v3.1", "Growth Marketer removed", "Timu IS the growth marketer", "EUR 4,500/mo from M12", "Saves EUR 73K (M12-M24)"],
    ["v3.1", "Subscribers gated to M4+", "App launches with space", "3 new subs/mo from M1", "M1-M3: zero subs, zero sub revenue"],
    ["v3.1", "UX/UI → CapEx", "EUR 15K one-off App Design (M1)", "EUR 800/mo OpEx", "Saves EUR 19.2K OpEx, adds EUR 15K CapEx"],
    ["v3.1", "B19 note update", "'Scaling phase'", "'Growth marketer effect'", "Label only — no financial impact"],
    ["v3.2", "LED therapy → free", "EUR 0 price, self-service", "EUR 45/session, EUR 25 practitioner", "Forgoes ~EUR 151K rev, saves ~EUR 84K COGS"],
    ["v3.3", "Ablöse (key money) added", "EUR 50K M3 CapEx, amortized 36mo", "Not modeled", "CapEx EUR 98K → EUR 148K; tighter runway"],
    ["v3.3", "Kaution (security deposit)", "EUR 7.5K M3 cash outflow", "Not modeled", "Cash Flow only — refundable, not P&L"],
]

for item in corrections:
    for col, val in enumerate(item, 1):
        c = ws3.cell(row=row, column=col, value=val)
        c.font = f_sub
        if col == 1:
            c.font = Font(name="Calibri", size=10, bold=True, color=AMBER)
        c.border = thin_bottom
    row += 1

row += 1

# Persona audit summary

ws3.cell(row=row, column=1, value="7-PERSONA AUDIT RESULTS").font = f_section
ws3.cell(row=row, column=1).fill = fill_section
for c in range(2, 6):
    ws3.cell(row=row, column=c).fill = fill_section
row += 1

for col, h in enumerate(["#", "Persona", "Key Finding", "EUR Impact", "Status in v3"], 1):
    c = ws3.cell(row=row, column=col, value=h)
    c.font = f_header; c.fill = fill_header
row += 1

persona_results = [
    ["P1", "Angel Investor", "EUR 113-236K missing costs; CAC EUR 40 too low", "EUR 175K midpoint gap", "All items added"],
    ["P2", "Steuerberater", "GewSt 14.35% + KSt 15.825% = ~30% tax; IHK EUR 120/yr", "~EUR 3K/mo when profitable", "Tax provision added"],
    ["P3", "Operations Manager", "Rent is Kaltmiete; buildout EUR 25K+; utilities EUR 500", "+EUR 665-1,600/mo", "Nebenkosten, buildout, utilities updated"],
    ["P4", "Food Safety Inspector", "No Gaststättenerlaubnis needed; NMN NOT EU approved", "+EUR 2.6-5.8K one-off", "HACCP, hygiene, food safety added"],
    ["P5", "Growth CFO", "EUR 218K P&L vs cash gap; ARPU EUR 38.44 not EUR 49", "Working capital critical", "Cash flow sheet added"],
    ["P6", "Compliance Legal", "DSB mandatory; DPIA mandatory; insurance EUR 650/mo", "+EUR 29.5-53K one-off", "DSB, DPIA, MDR, insurance split added"],
    ["P7", "HR Manager", "CTO recruitment EUR 18K; employer multiplier 1.25x", "+EUR 16-27K one-off", "Recruitment, multiplier, onboarding added"],
]

for item in persona_results:
    for col, val in enumerate(item, 1):
        c = ws3.cell(row=row, column=col, value=val)
        c.font = f_sub
        if col == 1:
            c.font = Font(name="Calibri", size=10, bold=True, color=SAGE)
        c.border = thin_bottom
    row += 1


# ═══════════════════════════════════════════
# REORDER SHEETS
# ═══════════════════════════════════════════
# Reorder: P&L first, then Assumptions, Growth Curves, Cash Flow, Unit Economics, Scenarios, Decisions
desired_order = ["P&L", "Assumptions", "Growth Curves", "Cash Flow", "Unit Economics", "Scenarios", "Decisions"]
for i, name in enumerate(desired_order):
    current_idx = wb.sheetnames.index(name)
    wb.move_sheet(name, offset=i - current_idx)


# ─── SAVE ───
output_path = "/Users/timoel/Downloads/pitchbook/financials/models/alche-pnl-24mo-v3.xlsx"
wb.save(output_path)
print(f"Excel P&L v3 saved to: {output_path}")
print(f"Sheets: {wb.sheetnames}")
print(f"P&L rows: {ROW}")
print(f"\nKey changes from v2:")
print(f"  - Employer multiplier: 1.22x → 1.25x")
print(f"  - Insurance: EUR 200 → EUR 650/mo (split into 5 categories)")
print(f"  - Buildout: EUR 15K → EUR 25K")
print(f"  - Recruitment: EUR 5K → EUR 18K + EUR 3.5K onboarding")
print(f"  - NEW: DSB EUR 400/mo, Nebenkosten EUR 325/mo, Utilities EUR 500/mo")
print(f"  - NEW: Tax provision ~30% when profitable")
print(f"  - NEW: Cash Flow Statement sheet (P5 working capital gap)")
print(f"  - NEW: Unit Economics Dashboard")
print(f"  - NEW: Scenario Summary (Base/Bull/Bear)")
print(f"  - CGM: Commission model (D3 updated) — EUR 15/user/mo from M7, Premium 3-mo gate or add-on")
print(f"  - NEW: Bad debt provision, GEMA, Rundfunkbeitrag, Verpackungsgesetz, KSK")
print(f"  - CapEx: EUR 51K → ~EUR 148K (v3.3: +EUR 50K Ablöse, +DPIA, MDR, fire safety, HACCP, App Design)")
print(f"  - NEW: Ablöse EUR 50K (M3 CapEx, amortized over 36mo)")
print(f"  - NEW: Kaution EUR 7.5K (M3 cash outflow only — refundable deposit, not P&L)")
print(f"\nv3 corrections (Timu review):")
print(f"  - REMOVED: Growth Marketer hire (Timu IS the growth marketer)")
print(f"  - CHANGED: Subscribers gated to M4+ (no app before space launch)")
print(f"  - MOVED: UX/UI EUR 800/mo → CapEx EUR 15K (App Design & Development)")
print(f"  - UPDATED: B19 note 'Growth marketer effect' → 'Scaling phase'")
print(f"  - CHANGED: LED therapy → free with smoothie (B24=0), no practitioner (B25=0)")
